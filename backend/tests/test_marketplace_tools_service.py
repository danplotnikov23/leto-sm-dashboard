from __future__ import annotations

import re
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.integrations.ozon_elastic_boosting.processor import load_workbook_file
from app.services.marketplace_tools_service import MarketplaceToolsService


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _ozon_promo_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары и цены"
    sheet.append([""] * 15)
    headers = [""] * 15
    headers[1] = "SKU"
    headers[2] = "Артикул"
    headers[4] = "Название"
    headers[5] = "Категория"
    headers[7] = "Ваша цена, RUB"
    headers[10] = "Участие товара в акции"
    headers[11] = "Итоговая цена по акции, RUB"
    headers[14] = "Цена для минимального акционного бустинга, RUB"
    sheet.append(headers)
    sheet.append([""] * 15)
    for sku, article, price, promo_price in (
        ("100", "ART-100", 10_000, 9_500),
        ("101", "ART-101", 10_000, 9_500),
    ):
        values = [""] * 15
        values[1] = sku
        values[2] = article
        values[4] = f"Товар {sku}"
        values[5] = "Тест"
        values[7] = price
        values[10] = "Да"
        values[11] = promo_price
        values[14] = promo_price
        sheet.append(values)
    workbook.create_sheet("Не трогать")
    return _workbook_bytes(workbook)


def _ozon_promo_auto_add_bytes() -> bytes:
    """Mimics Ozon's "FBS auto-add" promo export: dynamic sheet/column names
    that embed the promo start date, instead of the static "Товары и цены"
    layout used by _ozon_promo_bytes()."""
    workbook = Workbook()
    workbook.active.title = "Описание"
    sheet = workbook.create_sheet("Участвуют с 07.08.2026 (MSK)")
    headers = [""] * 15
    headers[0] = "OzonID"
    headers[1] = "SKU"
    headers[2] = "Артикул"
    headers[3] = "Название"
    headers[4] = "Категория"
    headers[6] = "Ваша цена, RUB"
    headers[12] = "Рассчитанная цена для участия в акции, RUB"
    headers[13] = "Участие товара в акции с 07.08.2026 (MSK)"
    headers[14] = "Итоговая цена по акции с 07.08.2026 (MSK), RUB"
    sheet.append([""] * 15)
    sheet.append(headers)
    sheet.append([""] * 15)
    for sku, article, price, min_boost_price in (
        ("100", "ART-100", 10_000, 9_500),
        ("101", "ART-101", 10_000, 9_500),
    ):
        values = [""] * 15
        values[1] = sku
        values[2] = article
        values[3] = f"Товар {sku}"
        values[4] = "Тест"
        values[6] = price
        values[12] = min_boost_price
        sheet.append(values)
    return _workbook_bytes(workbook)


def _ozon_ads_bytes() -> bytes:
    workbook = Workbook()
    statistics = workbook.active
    statistics.title = "Statistics"
    statistics.append(["Период"])
    statistics.append(["SKU"])
    statistics.append(["101"])
    union = workbook.create_sheet("Union")
    union.append(["Период"])
    union.append(
        [
            "SKU в продвижении",
            "",
            "",
            "",
            "",
            "SKU из объединенной карточки",
        ]
    )
    return _workbook_bytes(workbook)


def _yandex_promotion_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары и цены"
    sheet.cell(9, 3).value = "Артикул товара (SKU)"
    sheet.cell(9, 4).value = "Название товара"
    sheet.cell(9, 9).value = "Цена из каталога"
    sheet.cell(9, 12).value = "Максимальная цена для участия в Бестселлерах"
    sheet.cell(9, 14).value = "Скидка по акции, %"
    sheet.cell(9, 16).value = "Способ 2. Цена по акции"
    sheet.cell(9, 18).value = "Способ добавления товара"
    sheet.cell(10, 3).value = "Уникальный идентификатор товара"
    for row, values in enumerate(
        (
            ("A1", "Первый", 1_000, 950),
            ("A2", "Второй", 2_000, 1_800),
        ),
        start=11,
    ):
        sku, name, catalog_price, limit_price = values
        sheet.cell(row, 3).value = sku
        sheet.cell(row, 4).value = name
        sheet.cell(row, 9).value = catalog_price
        sheet.cell(row, 12).value = limit_price
        sheet.cell(row, 16).value = 123
        sheet.cell(row, 18).value = "Автоучастие"
    settings = workbook.create_sheet("Настройки")
    settings.sheet_state = "hidden"
    settings["A1"] = "settings"
    return _workbook_bytes(workbook)


def _yandex_promotion_bytes_zero_discount_needed() -> bytes:
    """One row where the max promo price already equals the catalog price -
    Yandex's own required entry discount for this SKU is exactly 0%,
    matching a seller who already lowered the price directly rather than
    via a promo markdown (mirrors the Ozon elastic-boosting case)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары и цены"
    sheet.cell(9, 3).value = "Артикул товара (SKU)"
    sheet.cell(9, 4).value = "Название товара"
    sheet.cell(9, 9).value = "Цена из каталога"
    sheet.cell(9, 12).value = "Максимальная цена для участия в Бестселлерах"
    sheet.cell(9, 14).value = "Скидка по акции, %"
    sheet.cell(9, 16).value = "Способ 2. Цена по акции"
    sheet.cell(9, 18).value = "Способ добавления товара"
    sheet.cell(10, 3).value = "Уникальный идентификатор товара"
    sheet.cell(11, 3).value = "A1"
    sheet.cell(11, 4).value = "Первый"
    sheet.cell(11, 9).value = 1_000
    sheet.cell(11, 12).value = 1_000
    settings = workbook.create_sheet("Настройки")
    settings.sheet_state = "hidden"
    settings["A1"] = "settings"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _yandex_products_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Список товаров"
    sheet.cell(1, 4).value = "Основные параметры"
    sheet.cell(1, 21).value = "Цена"
    sheet.cell(2, 4).value = "Ваш SKU *"
    sheet.cell(2, 5).value = "Название товара *"
    sheet.cell(2, 21).value = "Цена *"
    sheet.cell(3, 4).value = "Уникальный идентификатор товара"
    sheet.cell(3, 5).value = "Название товара"
    sheet.cell(3, 21).value = "Цена в валюте кабинета"
    for row, values in enumerate(
        (("A1", "Первый", 1_000), ("A2", "Второй", 2_000)),
        start=4,
    ):
        sheet.cell(row, 4).value = values[0]
        sheet.cell(row, 5).value = values[1]
        sheet.cell(row, 21).value = values[2]
    return _workbook_bytes(workbook)


def _yandex_template_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Инструкция"
    sheet = workbook.create_sheet("Товары")
    sheet["A1"] = "Товары для кампании с бустом продаж"
    sheet["A2"] = "Ваш SKU*"
    sheet["B2"] = "Ставка, %*"
    sheet["C2"] = "Название товара"
    sheet["D2"] = "Цена (валюта кабинета)"
    sheet["A3"] = "Уникальный идентификатор товара."
    sheet["B3"] = "От 0,5 до 99,9."
    workbook.create_sheet("Настройки").sheet_state = "hidden"
    return _workbook_bytes(workbook)


def test_ozon_tool_creates_reopenable_xlsx_and_csv(tmp_path: Path) -> None:
    service = MarketplaceToolsService(tmp_path)

    result = service.process_ozon_elastic_boosting(
        promo_bytes=_ozon_promo_bytes(),
        ads_bytes=_ozon_ads_bytes(),
        min_discount=2,
        max_discount=7,
        exclude_direct_ads=True,
        strict_union_exclusion=True,
    )

    assert result.stats["added_to_promo"] == 1
    assert result.stats["excluded_direct_ads"] == 1
    assert len(result.artifacts) == 2
    xlsx_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook["Товары и цены"]["K4"].value == "Да"
    assert workbook["Товары и цены"]["K5"].value is None
    assert "Не трогать" in workbook.sheetnames
    workbook.close()


def test_ozon_tool_supports_auto_add_template_variant(tmp_path: Path) -> None:
    """Ozon's "FBS auto-add" promo export uses a differently-named sheet and
    date-suffixed participation/price columns instead of the static
    "Товары и цены" layout — the tool must recognize both."""
    service = MarketplaceToolsService(tmp_path)

    result = service.process_ozon_elastic_boosting(
        promo_bytes=_ozon_promo_auto_add_bytes(),
        ads_bytes=_ozon_ads_bytes(),
        min_discount=2,
        max_discount=7,
        exclude_direct_ads=True,
        strict_union_exclusion=True,
    )

    assert result.stats["added_to_promo"] == 1
    assert result.stats["excluded_direct_ads"] == 1
    xlsx_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(xlsx_path, read_only=True)
    sheet = workbook["Участвуют с 07.08.2026 (MSK)"]
    assert sheet["N4"].value == "Да"
    assert sheet["N5"].value is None
    workbook.close()


def _ozon_promo_negative_discount_bytes() -> bytes:
    """Two rows where your_price dropped below Ozon's own min-boost
    threshold (negative entry discount) - one clean SKU, one SKU that is
    also in direct ads and must stay excluded regardless of the flag."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары и цены"
    sheet.append([""] * 15)
    headers = [""] * 15
    headers[1] = "SKU"
    headers[2] = "Артикул"
    headers[4] = "Название"
    headers[5] = "Категория"
    headers[7] = "Ваша цена, RUB"
    headers[10] = "Участие товара в акции"
    headers[11] = "Итоговая цена по акции, RUB"
    headers[14] = "Цена для минимального акционного бустинга, RUB"
    sheet.append(headers)
    sheet.append([""] * 15)
    for sku, article, your_price, min_boost_price in (
        ("100", "ART-100", 13_900, 16_344),
        ("101", "ART-101", 13_900, 16_344),
    ):
        values = [""] * 15
        values[1] = sku
        values[2] = article
        values[4] = f"Товар {sku}"
        values[5] = "Тест"
        values[7] = your_price
        values[11] = min_boost_price
        values[14] = min_boost_price
        sheet.append(values)
    return _workbook_bytes(workbook)


def test_ozon_tool_adds_negative_discount_rows_at_zero_when_enabled(tmp_path: Path) -> None:
    service = MarketplaceToolsService(tmp_path)

    result = service.process_ozon_elastic_boosting(
        promo_bytes=_ozon_promo_negative_discount_bytes(),
        ads_bytes=_ozon_ads_bytes(),
        min_discount=2,
        max_discount=7,
        exclude_direct_ads=True,
        strict_union_exclusion=True,
        zero_discount_for_negative=True,
    )

    assert result.stats["added_to_promo"] == 1
    assert result.stats["added_at_zero_discount"] == 1
    assert result.stats["excluded_direct_ads"] == 1
    xlsx_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(xlsx_path, read_only=True)
    sheet = workbook["Товары и цены"]
    assert sheet["K4"].value == "Да"
    assert sheet["L4"].value == 13_900
    assert sheet["K5"].value is None
    assert sheet["L5"].value == 16_344
    workbook.close()


def test_ozon_tool_excludes_negative_discount_rows_by_default(tmp_path: Path) -> None:
    service = MarketplaceToolsService(tmp_path)

    result = service.process_ozon_elastic_boosting(
        promo_bytes=_ozon_promo_negative_discount_bytes(),
        ads_bytes=_ozon_ads_bytes(),
        min_discount=2,
        max_discount=7,
        exclude_direct_ads=True,
        strict_union_exclusion=True,
    )

    assert result.stats["added_to_promo"] == 0
    assert result.stats["added_at_zero_discount"] == 0
    xlsx_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(xlsx_path, read_only=True)
    sheet = workbook["Товары и цены"]
    assert sheet["K4"].value is None
    assert sheet["L4"].value == 16_344
    workbook.close()


def _corrupt_workbook_style_xml(data: bytes) -> bytes:
    """Simulates the malformed XML seen in real Ozon exports: an empty
    border ``style`` attribute and an out-of-enum ``activePane`` value.
    Excel opens these files fine; openpyxl raises ValueError on them."""
    source = zipfile.ZipFile(BytesIO(data))
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = content.decode("utf-8")
                text = text.replace("<left />", '<left style=""/>', 1)
                content = text.encode("utf-8")
            elif item.filename == "xl/worksheets/sheet1.xml":
                text = content.decode("utf-8")
                text = re.sub(
                    r'(<sheetView[^>]*>)',
                    r'\1<pane activePane="top" state="frozen"/>',
                    text,
                    count=1,
                )
                content = text.encode("utf-8")
            target.writestr(item, content)
    return buffer.getvalue()


def test_load_workbook_file_recovers_from_invalid_ozon_style_xml() -> None:
    corrupted = _corrupt_workbook_style_xml(_ozon_promo_bytes())

    workbook = load_workbook_file(BytesIO(corrupted))
    try:
        assert "Товары и цены" in workbook.sheetnames
    finally:
        workbook.close()


def test_yandex_promotion_clears_rows_outside_range(tmp_path: Path) -> None:
    service = MarketplaceToolsService(tmp_path)

    result = service.prepare_yandex_promotion(
        promotions_bytes=_yandex_promotion_bytes(),
        min_discount=1,
        max_discount=6,
        target_discount=6,
    )

    assert result.stats["rows_selected"] == 1
    output_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Товары и цены"]
    assert sheet["P11"].value == 940
    assert sheet["R11"].value == "Добавлен вручную"
    assert sheet["P12"].value is None
    assert sheet["R12"].value == "Не добавлялся"
    assert workbook["Настройки"].sheet_state == "hidden"
    workbook.close()


def test_yandex_promotion_allows_zero_target_discount(tmp_path: Path) -> None:
    service = MarketplaceToolsService(tmp_path)

    result = service.prepare_yandex_promotion(
        promotions_bytes=_yandex_promotion_bytes_zero_discount_needed(),
        min_discount=0,
        max_discount=0,
        target_discount=0,
    )

    assert result.stats["rows_selected"] == 1
    output_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Товары и цены"]
    assert sheet["P11"].value == 1_000
    assert sheet["R11"].value == "Добавлен вручную"
    workbook.close()


def test_yandex_boost_creates_reopenable_template(tmp_path: Path) -> None:
    service = MarketplaceToolsService(tmp_path)

    result = service.build_yandex_boost(
        products_bytes=_yandex_products_bytes(),
        template_bytes=_yandex_template_bytes(),
        template_filename="template.xlsx",
        bid=17,
        promotions_bytes=None,
    )

    assert result.stats["rows_for_boost"] == 2
    output_path = service.resolve_artifact(
        result.artifact_id,
        result.artifacts[0].filename,
    )
    workbook = load_workbook(output_path, data_only=False)
    assert workbook["Товары"]["A4"].value == "A1"
    assert workbook["Товары"]["B4"].value == 17
    assert workbook["Товары"]["A5"].value == "A2"
    assert workbook["Настройки"].sheet_state == "hidden"
    workbook.close()
