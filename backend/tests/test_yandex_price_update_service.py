from __future__ import annotations

import zipfile
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.schemas.yandex_price_update import YandexPriceUpdateItem
from app.services.yandex_price_update_service import (
    COL_COST,
    COL_NAME,
    COL_PRICE,
    COL_PRICE_BEFORE_DISCOUNT,
    COL_SKU,
    DATA_START_ROW,
    SHEET_NAME,
    YandexPriceUpdateService,
    YandexPriceUpdateTemplateError,
)


class _FakeUnitEconomyIndexService:
    def __init__(self, products_by_sku=None):
        self._products_by_sku = products_by_sku or {}

    def list_products(self):
        return []

    def find_by_offer_id(self, offer_id):
        return self._products_by_sku.get(offer_id)

    def find_by_offer_id_for_date(self, offer_id, effective_date):
        return self._products_by_sku.get(offer_id)


class _FakeProduct:
    def __init__(self, cost_without_vat=None, expense_cost=None, price_before_discount_with_vat=None, category=None):
        self.cost_without_vat = cost_without_vat
        self.expense_cost = expense_cost
        self.price_before_discount_with_vat = price_before_discount_with_vat
        self.category = category


def _yandex_price_template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    for row in range(1, DATA_START_ROW):
        sheet.append([])
    for offset, sku in enumerate(("SKU-1", "SKU-2", "SKU-3")):
        row_number = DATA_START_ROW + offset
        sheet.cell(row=row_number, column=COL_SKU, value=sku)
        sheet.cell(row=row_number, column=COL_NAME, value=f"Товар {sku}")
        # Real Yandex exports pre-format the editable-price columns, so a
        # styled placeholder cell exists even before any value is filled in
        # - reproduce that here, since it's the exact shape
        # apply_new_prices' cell-patching regex must match.
        for column in (COL_PRICE, COL_PRICE_BEFORE_DISCOUNT, COL_COST):
            placeholder = sheet.cell(row=row_number, column=column)
            placeholder.number_format = "0.00"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _yandex_price_template_bytes_missing_optional_cells() -> bytes:
    """Real Yandex export seen with a subset of rows where the cost/
    strikethrough columns have no <c> element at all - not even a styled
    placeholder, unlike _yandex_price_template_bytes()'s rows. Reproduces
    the "Не удалось обновить ячейки" failure from a genuinely blank cell."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    for row in range(1, DATA_START_ROW):
        sheet.append([])
    row_number = DATA_START_ROW
    sheet.cell(row=row_number, column=COL_SKU, value="SKU-1")
    sheet.cell(row=row_number, column=COL_NAME, value="Товар SKU-1")
    sheet.cell(row=row_number, column=COL_PRICE, value=1000.0)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _service(products_by_sku=None) -> YandexPriceUpdateService:
    return YandexPriceUpdateService(_FakeUnitEconomyIndexService(products_by_sku))


def test_apply_new_prices_inserts_cost_cell_absent_from_source_row():
    template_bytes = _yandex_price_template_bytes_missing_optional_cells()
    service = _service({"SKU-1": _FakeProduct(expense_cost=500.0)})

    result_bytes = service.apply_new_prices(
        template_bytes,
        [YandexPriceUpdateItem(sku="SKU-1", new_price=1234.5)],
    )

    worksheet = service._load_products_sheet(result_bytes, data_only=True)
    row = next(
        r for r in worksheet.iter_rows(min_row=DATA_START_ROW)
        if r[COL_SKU - 1].value == "SKU-1"
    )
    assert row[COL_PRICE - 1].value == 1234.5
    assert row[COL_COST - 1].value == 500.0


def test_apply_new_prices_inserts_both_missing_cost_and_strikethrough_cells():
    # F104 and H104 both absent in the same row, as seen in the real file -
    # both must land correctly, in the right column order.
    template_bytes = _yandex_price_template_bytes_missing_optional_cells()
    service = _service(
        {
            "SKU-1": _FakeProduct(
                expense_cost=500.0,
                price_before_discount_with_vat=2000.0,
            )
        }
    )

    result_bytes = service.apply_new_prices(
        template_bytes,
        [YandexPriceUpdateItem(sku="SKU-1", new_price=1000.0)],
    )

    worksheet = service._load_products_sheet(result_bytes, data_only=True)
    row = next(
        r for r in worksheet.iter_rows(min_row=DATA_START_ROW)
        if r[COL_SKU - 1].value == "SKU-1"
    )
    assert row[COL_PRICE - 1].value == 1000.0
    assert row[COL_PRICE_BEFORE_DISCOUNT - 1].value == 2000.0
    assert row[COL_COST - 1].value == 500.0


def test_apply_new_prices_writes_price_and_cost():
    template_bytes = _yandex_price_template_bytes()
    service = _service({"SKU-1": _FakeProduct(expense_cost=500.0)})

    result_bytes = service.apply_new_prices(
        template_bytes,
        [YandexPriceUpdateItem(sku="SKU-1", new_price=1234.5)],
    )

    worksheet = service._load_products_sheet(result_bytes, data_only=True)
    row = next(
        r for r in worksheet.iter_rows(min_row=DATA_START_ROW)
        if r[COL_SKU - 1].value == "SKU-1"
    )
    assert row[COL_PRICE - 1].value == 1234.5
    assert row[COL_COST - 1].value == 500.0


def test_apply_new_prices_fills_strikethrough_price_within_valid_discount_range():
    template_bytes = _yandex_price_template_bytes()
    # 1000 -> 900 is a 10% discount, inside Yandex's required 5-75% range.
    service = _service({"SKU-1": _FakeProduct(price_before_discount_with_vat=1000.0)})

    result_bytes = service.apply_new_prices(
        template_bytes,
        [YandexPriceUpdateItem(sku="SKU-1", new_price=900.0)],
    )

    worksheet = service._load_products_sheet(result_bytes, data_only=True)
    row = next(
        r for r in worksheet.iter_rows(min_row=DATA_START_ROW)
        if r[COL_SKU - 1].value == "SKU-1"
    )
    assert row[COL_PRICE_BEFORE_DISCOUNT - 1].value == 1000.0


def test_apply_new_prices_skips_strikethrough_price_outside_valid_discount_range():
    template_bytes = _yandex_price_template_bytes()
    # 1000 -> 990 is only a 1% discount - below Yandex's 5% minimum, so the
    # field must be left untouched rather than uploading a value Yandex
    # would reject.
    service = _service({"SKU-1": _FakeProduct(price_before_discount_with_vat=1000.0)})

    result_bytes = service.apply_new_prices(
        template_bytes,
        [YandexPriceUpdateItem(sku="SKU-1", new_price=990.0)],
    )

    worksheet = service._load_products_sheet(result_bytes, data_only=True)
    row = next(
        r for r in worksheet.iter_rows(min_row=DATA_START_ROW)
        if r[COL_SKU - 1].value == "SKU-1"
    )
    assert row[COL_PRICE - 1].value == 990.0
    assert row[COL_PRICE_BEFORE_DISCOUNT - 1].value is None


def test_apply_new_prices_leaves_every_other_zip_part_byte_identical():
    template_bytes = _yandex_price_template_bytes()
    service = _service({"SKU-1": _FakeProduct(expense_cost=500.0)})

    result_bytes = service.apply_new_prices(
        template_bytes,
        [YandexPriceUpdateItem(sku="SKU-1", new_price=1234.5)],
    )

    original_zip = zipfile.ZipFile(BytesIO(template_bytes))
    result_zip = zipfile.ZipFile(BytesIO(result_bytes))
    assert set(original_zip.namelist()) == set(result_zip.namelist())

    changed_parts = [
        name
        for name in original_zip.namelist()
        if original_zip.read(name) != result_zip.read(name)
    ]
    assert len(changed_parts) == 1


def test_apply_new_prices_rejects_unknown_sku():
    template_bytes = _yandex_price_template_bytes()
    service = _service()

    with pytest.raises(YandexPriceUpdateTemplateError, match="SKU-999"):
        service.apply_new_prices(
            template_bytes,
            [YandexPriceUpdateItem(sku="SKU-999", new_price=1000)],
        )


def test_apply_new_prices_rejects_empty_updates():
    template_bytes = _yandex_price_template_bytes()
    service = _service()

    with pytest.raises(YandexPriceUpdateTemplateError):
        service.apply_new_prices(template_bytes, [])


def test_search_products_matches_by_unit_economy_cost():
    template_bytes = _yandex_price_template_bytes()
    worksheet_bytes_service = _service(
        {"SKU-1": _FakeProduct(cost_without_vat=1000.0, category="Столы")}
    )
    # Give SKU-1 a current price via a second write pass, matching how a
    # real export would already have prices populated.
    result = worksheet_bytes_service.search_products(template_bytes, query="")
    assert result.total_rows_in_template == 3
    sku1 = next(m for m in result.matches if m.sku == "SKU-1")
    assert sku1.unit_economy_matched is True
    assert sku1.category == "Столы"
    sku2 = next(m for m in result.matches if m.sku == "SKU-2")
    assert sku2.unit_economy_matched is False
