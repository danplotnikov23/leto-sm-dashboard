from __future__ import annotations

import zipfile
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.schemas.price_update import PriceUpdateItem
from app.services.price_update_service import (
    COL_NEW_COST,
    COL_NEW_MIN_PRICE,
    COL_NEW_PRICE_BEFORE_DISCOUNT,
    COL_NEW_PRICE_WITH_DISCOUNT,
    COL_OFFER_ID,
    DATA_START_ROW,
    DEFAULT_VAT_MULTIPLIER,
    SHEET_NAME,
    PriceUpdateService,
    PriceUpdateTemplateError,
)


class _FakeUnitEconomyIndexService:
    def __init__(self, products_by_offer_id=None):
        self._products_by_offer_id = products_by_offer_id or {}

    def list_products(self):
        return []

    def find_by_offer_id(self, offer_id):
        return self._products_by_offer_id.get(offer_id)

    def find_by_offer_id_for_date(self, offer_id, effective_date):
        return self._products_by_offer_id.get(offer_id)


class _FakeProduct:
    def __init__(self, expense_cost=None, price_before_discount_with_vat=None):
        self.expense_cost = expense_cost
        self.price_before_discount_with_vat = price_before_discount_with_vat


def _ozon_price_template_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Инструкция"
    sheet = workbook.create_sheet(SHEET_NAME)
    for row in range(1, DATA_START_ROW):
        sheet.append([])
    for offset, offer_id in enumerate(("ART-1", "ART-2", "ART-3")):
        row_number = DATA_START_ROW + offset
        sheet.cell(row=row_number, column=COL_OFFER_ID, value=offer_id)
        # Real Ozon exports pre-format the whole editable-price columns, so
        # every row carries an empty-but-styled placeholder cell there even
        # before any value is filled in - reproduce that here, since it's
        # the exact shape apply_new_prices' cell-patching regex must match.
        for column in (
            COL_NEW_PRICE_BEFORE_DISCOUNT,
            COL_NEW_PRICE_WITH_DISCOUNT,
            COL_NEW_COST,
            COL_NEW_MIN_PRICE,
        ):
            placeholder = sheet.cell(row=row_number, column=column)
            placeholder.number_format = "0.00"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _service(products_by_offer_id=None) -> PriceUpdateService:
    return PriceUpdateService(_FakeUnitEconomyIndexService(products_by_offer_id))


def test_apply_new_prices_writes_price_min_price_cost_and_price_before_discount():
    template_bytes = _ozon_price_template_bytes()
    service = _service(
        {
            "ART-1": _FakeProduct(expense_cost=500.0, price_before_discount_with_vat=2000.0),
        }
    )

    result_bytes = service.apply_new_prices(
        template_bytes,
        [
            PriceUpdateItem(offer_id="ART-1", new_price_with_discount=1234.5),
            PriceUpdateItem(offer_id="ART-3", new_price_with_discount=999),
        ],
    )

    worksheet = service._load_products_sheet(result_bytes, read_only=True, data_only=True)
    row_by_offer_id = {}
    for row in worksheet.iter_rows(min_row=DATA_START_ROW):
        offer_id = row[COL_OFFER_ID - 1].value
        if offer_id is not None:
            row_by_offer_id[offer_id] = row

    art1 = row_by_offer_id["ART-1"]
    assert art1[COL_NEW_PRICE_WITH_DISCOUNT - 1].value == 1234.5
    assert art1[COL_NEW_MIN_PRICE - 1].value == round(1234.5 / DEFAULT_VAT_MULTIPLIER, 2)
    # Below the price we just set, per Ozon's own "не может быть больше
    # текущей цены" constraint on the minimum-price field.
    assert art1[COL_NEW_MIN_PRICE - 1].value < art1[COL_NEW_PRICE_WITH_DISCOUNT - 1].value
    assert art1[COL_NEW_COST - 1].value == 500.0
    # Unlike price/min-price/cost, "цена до скидки" is copied straight from
    # unit-economy's own value, not derived from the price we just set.
    assert art1[COL_NEW_PRICE_BEFORE_DISCOUNT - 1].value == 2000.0

    art2 = row_by_offer_id["ART-2"]
    assert art2[COL_NEW_PRICE_WITH_DISCOUNT - 1].value is None
    assert art2[COL_NEW_MIN_PRICE - 1].value is None
    assert art2[COL_NEW_COST - 1].value is None
    assert art2[COL_NEW_PRICE_BEFORE_DISCOUNT - 1].value is None

    art3 = row_by_offer_id["ART-3"]
    assert art3[COL_NEW_PRICE_WITH_DISCOUNT - 1].value == 999
    assert art3[COL_NEW_MIN_PRICE - 1].value == round(999 / DEFAULT_VAT_MULTIPLIER, 2)
    # No unit-economy match for ART-3 in this test - cost and "цена до
    # скидки" are left untouched rather than guessed.
    assert art3[COL_NEW_COST - 1].value is None
    assert art3[COL_NEW_PRICE_BEFORE_DISCOUNT - 1].value is None


def test_apply_new_prices_leaves_every_other_zip_part_byte_identical():
    # A full openpyxl load-mutate-save round trip of a real Ozon export was
    # found to silently corrupt untouched cells elsewhere in the workbook
    # (see price_update_service.apply_new_prices docstring/comment) - Ozon's
    # own upload validator then rejected the result. The fix patches only
    # the target cells' XML directly, so every other part of the zip must
    # stay byte-for-byte identical to the original.
    template_bytes = _ozon_price_template_bytes()
    service = _service({"ART-1": _FakeProduct(expense_cost=500.0)})

    result_bytes = service.apply_new_prices(
        template_bytes,
        [PriceUpdateItem(offer_id="ART-1", new_price_with_discount=1234.5)],
    )

    original_zip = zipfile.ZipFile(BytesIO(template_bytes))
    result_zip = zipfile.ZipFile(BytesIO(result_bytes))
    assert set(original_zip.namelist()) == set(result_zip.namelist())

    changed_parts = [
        name
        for name in original_zip.namelist()
        if original_zip.read(name) != result_zip.read(name)
    ]
    assert changed_parts == ["xl/worksheets/sheet2.xml"]


def test_apply_new_prices_rejects_unknown_offer_id():
    template_bytes = _ozon_price_template_bytes()
    service = _service()

    with pytest.raises(PriceUpdateTemplateError, match="ART-999"):
        service.apply_new_prices(
            template_bytes,
            [PriceUpdateItem(offer_id="ART-999", new_price_with_discount=1000)],
        )


def test_apply_new_prices_rejects_empty_updates():
    template_bytes = _ozon_price_template_bytes()
    service = _service()

    with pytest.raises(PriceUpdateTemplateError):
        service.apply_new_prices(template_bytes, [])
