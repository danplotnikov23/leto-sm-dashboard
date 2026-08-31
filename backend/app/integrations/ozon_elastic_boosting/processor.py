from __future__ import annotations

import csv
import logging
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import BinaryIO, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import CategoryOverride, ProcessingConfig

logger = logging.getLogger(__name__)


REQUIRED_PROMO_HEADERS = {
    "sku": "SKU",
    "your_price": "Ваша цена, RUB",
    "participation": "Участие товара в акции",
    "promo_price": "Итоговая цена по акции, RUB",
    # Ozon uses this column both as the "search boosting" campaign's entry
    # threshold ("Цена для минимального акционного бустинга, RUB") and, in
    # plain discount campaigns, as its own suggested entry price
    # ("Рассчитанная цена для участия в акции, RUB") - same role either
    # way: the price you'd need to reach to qualify.
    "min_boost_price": (
        "Цена для минимального акционного бустинга, RUB",
        "Рассчитанная цена для участия в акции, RUB",
    ),
}

PROMO_FALLBACK_COLUMNS = {
    "sku": 2,
    "your_price": 8,
    "participation": 11,
    "promo_price": 12,
    "min_boost_price": 15,
}

REPORT_HEADERS = [
    "номер строки",
    "SKU",
    "артикул",
    "название товара",
    "категория",
    "ваша цена",
    "цена для минимального бустинга",
    "скидка для входа, %",
    "цена по акции (предложено Ozon)",
    "цена по акции в L",
    "скидка в L, %",
    "было в акции",
    "статус рекламы",
    "результат",
    "причина исключения",
]

ARTICLE_HEADER_CANDIDATES = (
    "Артикул",
    "Ваш артикул",
)

TITLE_HEADER_CANDIDATES = (
    "Название товара",
    "Название",
    "Наименование товара",
    "Наименование",
)

CATEGORY_HEADER_CANDIDATES = (
    "Категория",
    "Категория товара",
)

AUTO_ADD_PARTICIPATION_PREFIX = "участие товара в акции"
AUTO_ADD_PROMO_PRICE_PREFIX = "итоговая цена по акции"
AUTO_ADD_MIN_BOOST_HEADER = "Рассчитанная цена для участия в акции, RUB"

VALID_BORDER_STYLES = {
    "hair", "mediumDashDot", "slantDashDot", "thick", "medium", "dotted",
    "mediumDashed", "dashDot", "double", "dashDotDot", "thin",
    "mediumDashDotDot", "dashed",
}
VALID_ACTIVE_PANES = {"topLeft", "topRight", "bottomLeft", "bottomRight"}


class ProcessingError(Exception):
    """Raised when the uploaded workbook cannot be processed safely."""


@dataclass(frozen=True)
class AdvertisedSkus:
    direct: frozenset[str]
    union_promotion: frozenset[str]
    union_related: frozenset[str]
    direct_rows: int = 0
    union_promotion_rows: int = 0
    union_related_rows: int = 0

    @property
    def strict_all(self) -> frozenset[str]:
        return self.direct | self.union_promotion | self.union_related


@dataclass(frozen=True)
class ReportRow:
    row_number: int
    sku: str | None
    article: str | None
    product_name: str | None
    category: str | None
    your_price: float | None
    min_boost_price: float | None
    ozon_suggested_promo_price: float | None
    promo_price: float | None
    discount_percent: float | None
    promo_discount_percent: float | None
    original_participation: str
    ad_status: str
    result: str
    reason: str

    def as_csv_row(self) -> list[str | int | float]:
        return [
            self.row_number,
            self.sku or "",
            self.article or "",
            self.product_name or "",
            self.category or "",
            "" if self.your_price is None else self.your_price,
            "" if self.min_boost_price is None else self.min_boost_price,
            "" if self.discount_percent is None else self.discount_percent,
            "" if self.ozon_suggested_promo_price is None else self.ozon_suggested_promo_price,
            "" if self.promo_price is None else self.promo_price,
            "" if self.promo_discount_percent is None else self.promo_discount_percent,
            self.original_participation,
            self.ad_status,
            self.result,
            self.reason,
        ]


@dataclass(frozen=True)
class ProcessingSummary:
    total_product_rows: int
    valid_price_rows: int
    in_discount_range: int
    excluded_direct_ads: int
    excluded_union_cards: int
    added_to_promo: int
    invalid_price_rows: int
    unrecognized_skus: int
    initial_yes_rows: int = 0
    removed_initial_yes_discount: int = 0
    removed_initial_yes_direct_ads: int = 0
    removed_initial_yes_union_cards: int = 0
    direct_ad_rows: int = 0
    direct_ad_skus_unique: int = 0
    strict_ad_skus_unique: int = 0
    excluded_by_category: int = 0
    excluded_manually: int = 0
    added_at_zero_discount: int = 0


@dataclass(frozen=True)
class ProcessingResult:
    workbook: Workbook
    summary: ProcessingSummary
    report_rows: tuple[ReportRow, ...]
    output_filename: str
    csv_filename: str


def _strip_invalid_border_style(match: re.Match[str]) -> str:
    return match.group(0) if match.group(1) in VALID_BORDER_STYLES else ""


def _fix_invalid_active_pane(match: re.Match[str]) -> str:
    value = match.group(1)
    if value in VALID_ACTIVE_PANES:
        return match.group(0)
    return 'activePane="bottomLeft"'


def _sanitize_workbook_bytes(data: bytes) -> bytes:
    """Repair Ozon export XML quirks that openpyxl's strict schema rejects.

    Some Ozon xlsx exports contain a border ``style=""``/``style="none"``
    attribute and an ``activePane`` value outside the OOXML enum, which
    Excel tolerates but openpyxl does not. Neither affects cell data, so
    they are safe to normalize before parsing.
    """
    try:
        source = zipfile.ZipFile(BytesIO(data))
    except zipfile.BadZipFile:
        return data

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/styles.xml":
                text = content.decode("utf-8")
                text = re.sub(r' style="([^"]*)"', _strip_invalid_border_style, text)
                text = text.replace("numFmtID=", "numFmtId=")
                content = text.encode("utf-8")
            elif item.filename.startswith("xl/worksheets/") and item.filename.endswith(
                ".xml"
            ):
                text = content.decode("utf-8")
                text = re.sub(r'activePane="([^"]*)"', _fix_invalid_active_pane, text)
                content = text.encode("utf-8")
            target.writestr(item, content)
    return buffer.getvalue()


def _read_all_bytes(file: str | Path | BinaryIO | BytesIO) -> bytes | None:
    if isinstance(file, (str, Path)):
        return Path(file).read_bytes()
    if hasattr(file, "seek") and hasattr(file, "read"):
        file.seek(0)
        return file.read()
    return None


def load_workbook_file(
    file: str | Path | BinaryIO | BytesIO,
    *,
    data_only: bool = False,
    read_only: bool = False,
) -> Workbook:
    try:
        return load_workbook(file, data_only=data_only, read_only=read_only)
    except ValueError:
        raw_bytes = _read_all_bytes(file)
        if raw_bytes is None:
            raise
        try:
            sanitized = _sanitize_workbook_bytes(raw_bytes)
            return load_workbook(
                BytesIO(sanitized), data_only=data_only, read_only=read_only
            )
        except Exception as exc:  # pragma: no cover - openpyxl error details vary
            logger.exception("Failed to open workbook after sanitizing")
            raise ProcessingError(f"Не удалось открыть Excel-файл: {exc}") from exc
    except Exception as exc:  # pragma: no cover - openpyxl error details vary
        logger.exception("Failed to open workbook")
        raise ProcessingError(f"Не удалось открыть Excel-файл: {exc}") from exc


def _header_key(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _normalize_category_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).casefold()


def _build_category_overrides_map(
    overrides: Sequence[CategoryOverride],
) -> dict[str, CategoryOverride]:
    return {_normalize_category_key(item.category): item for item in overrides if item.category}


def list_promo_categories(
    promo_values_workbook: Workbook,
    config: ProcessingConfig | None = None,
) -> list[tuple[str, int]]:
    """Distinct categories in the promo template, most common first.

    Reads the same "Категория"/"Категория товара" column the main
    processing loop already extracts per row - just tallies it without
    running the full eligibility pass, so the UI can offer a category
    picker before the user commits to running the tool.
    """
    config = config or ProcessingConfig()
    sheet_name = config.promo_sheet_name
    if sheet_name not in promo_values_workbook.sheetnames:
        auto_add_name = _find_auto_add_sheet(promo_values_workbook, config.auto_add_sheet_prefix)
        if auto_add_name is None:
            return []
        sheet_name = auto_add_name

    sheet = promo_values_workbook[sheet_name]
    category_column = _find_category_column(sheet, header_row=config.header_row)
    if category_column is None:
        return []

    counts: dict[str, int] = {}
    for row in range(config.data_start_row, sheet.max_row + 1):
        value = sheet.cell(row, category_column).value
        if value in (None, ""):
            continue
        category = str(value).strip()
        if category:
            counts[category] = counts.get(category, 0) + 1

    return sorted(counts.items(), key=lambda item: -item[1])


def _iter_headers(sheet: Worksheet, header_row: int) -> Iterable[tuple[int, str]]:
    for cell in sheet[header_row]:
        value = _header_key(cell.value)
        if value:
            yield cell.column, value


def find_columns_by_headers(
    sheet: Worksheet,
    required_headers: Mapping[str, str | tuple[str, ...]],
    *,
    header_row: int = 2,
    fallback_columns: Mapping[str, int] | None = None,
) -> dict[str, int]:
    found_headers = {header: column for column, header in _iter_headers(sheet, header_row)}
    columns: dict[str, int] = {}
    missing: list[str] = []

    for key, display_name in required_headers.items():
        candidates = (display_name,) if isinstance(display_name, str) else display_name
        column = None
        for candidate in candidates:
            column = found_headers.get(_header_key(candidate))
            if column is not None:
                break
        if column is None and fallback_columns and key in fallback_columns:
            fallback_column = fallback_columns[key]
            fallback_value = _header_key(sheet.cell(header_row, fallback_column).value)
            if any(fallback_value == _header_key(candidate) for candidate in candidates):
                column = fallback_column
        if column is None:
            missing.append(" / ".join(candidates))
        else:
            columns[key] = column

    if missing:
        raise ProcessingError(
            "Не найдены обязательные столбцы на листе "
            f"«{sheet.title}»: {', '.join(missing)}"
        )

    return columns


def _find_prefixed_column(
    sheet: Worksheet,
    prefix: str,
    *,
    header_row: int,
) -> int | None:
    wanted_prefix = _header_key(prefix)
    for column, header in _iter_headers(sheet, header_row):
        if header.startswith(wanted_prefix):
            return column
    return None


def _find_auto_add_sheet(workbook: Workbook, prefix: str) -> str | None:
    wanted_prefix = prefix.strip().casefold()
    for name in workbook.sheetnames:
        if name.strip().casefold().startswith(wanted_prefix):
            return name
    return None


def resolve_promo_sheet(
    promo_workbook: Workbook,
    promo_values_workbook: Workbook,
    config: ProcessingConfig,
) -> tuple[str, dict[str, int]]:
    """Locate the promo sheet and its required columns.

    Ozon exports the same "product participation" data under two different
    layouts depending on the promo type: a static "Товары и цены" sheet, or
    (for FBS auto-add promos) a sheet named "Участвуют с {date} (MSK)" whose
    participation/price columns embed that date in the header text.
    """
    if (
        config.promo_sheet_name in promo_workbook.sheetnames
        and config.promo_sheet_name in promo_values_workbook.sheetnames
    ):
        sheet = promo_workbook[config.promo_sheet_name]
        columns = find_columns_by_headers(
            sheet,
            REQUIRED_PROMO_HEADERS,
            header_row=config.header_row,
            fallback_columns=PROMO_FALLBACK_COLUMNS,
        )
        return config.promo_sheet_name, columns

    auto_add_name = _find_auto_add_sheet(promo_workbook, config.auto_add_sheet_prefix)
    auto_add_values_name = _find_auto_add_sheet(
        promo_values_workbook, config.auto_add_sheet_prefix
    )
    if auto_add_name and auto_add_values_name:
        sheet = promo_workbook[auto_add_name]
        sku_column = _find_single_column(sheet, "SKU", header_row=config.header_row)
        your_price_column = _find_single_column(
            sheet, "Ваша цена, RUB", header_row=config.header_row
        )
        min_boost_column = _find_single_column(
            sheet, AUTO_ADD_MIN_BOOST_HEADER, header_row=config.header_row
        )
        participation_column = _find_prefixed_column(
            sheet, AUTO_ADD_PARTICIPATION_PREFIX, header_row=config.header_row
        )
        promo_price_column = _find_prefixed_column(
            sheet, AUTO_ADD_PROMO_PRICE_PREFIX, header_row=config.header_row
        )
        missing = []
        if participation_column is None:
            missing.append("Участие товара в акции с … (MSK)")
        if promo_price_column is None:
            missing.append("Итоговая цена по акции с … (MSK), RUB")
        if missing:
            raise ProcessingError(
                f"Не найдены обязательные столбцы на листе «{auto_add_name}»: "
                + ", ".join(missing)
            )
        return auto_add_name, {
            "sku": sku_column,
            "your_price": your_price_column,
            "min_boost_price": min_boost_column,
            "participation": participation_column,
            "promo_price": promo_price_column,
        }

    raise ProcessingError(
        f"В шаблоне акции не найден ни лист «{config.promo_sheet_name}», "
        f"ни лист «{config.auto_add_sheet_prefix} …» — проверьте, что загружен файл "
        "шаблона акции, а не другой отчёт Ozon"
    )


def normalize_sku(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return str(value).strip() or None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".") or None

    value_as_text = str(value).strip()
    if value_as_text.endswith(".0") and value_as_text[:-2].isdigit():
        value_as_text = value_as_text[:-2]

    return value_as_text or None


def _find_single_column(
    sheet: Worksheet,
    header_name: str,
    *,
    header_row: int,
    fallback_column: int | None = None,
) -> int:
    wanted = _header_key(header_name)
    for column, header in _iter_headers(sheet, header_row):
        if header == wanted:
            return column

    if fallback_column is not None and _header_key(sheet.cell(header_row, fallback_column).value) == wanted:
        return fallback_column

    raise ProcessingError(f"Не найден столбец «{header_name}» на листе «{sheet.title}»")


def _collect_column_sku_values(sheet: Worksheet, column: int, *, start_row: int) -> tuple[str, ...]:
    skus: list[str] = []
    for row in range(start_row, sheet.max_row + 1):
        sku = normalize_sku(sheet.cell(row, column).value)
        if sku:
            skus.append(sku)
    return tuple(skus)


def _collect_column_skus(sheet: Worksheet, column: int, *, start_row: int) -> frozenset[str]:
    return frozenset(_collect_column_sku_values(sheet, column, start_row=start_row))


def collect_advertised_skus(
    ad_workbook: Workbook,
    config: ProcessingConfig,
) -> AdvertisedSkus:
    if config.statistics_sheet_name not in ad_workbook.sheetnames:
        raise ProcessingError(f"В рекламном отчёте отсутствует лист «{config.statistics_sheet_name}»")

    statistics_sheet = ad_workbook[config.statistics_sheet_name]
    sku_column = _find_single_column(
        statistics_sheet,
        "SKU",
        header_row=config.header_row,
        fallback_column=1,
    )
    direct_sku_values = _collect_column_sku_values(
        statistics_sheet,
        sku_column,
        start_row=config.header_row + 1,
    )
    direct_skus = frozenset(direct_sku_values)

    union_promotion_skus: frozenset[str] = frozenset()
    union_related_skus: frozenset[str] = frozenset()
    union_promotion_values: tuple[str, ...] = ()
    union_related_values: tuple[str, ...] = ()
    if config.union_sheet_name in ad_workbook.sheetnames:
        union_sheet = ad_workbook[config.union_sheet_name]
        union_header_row = _detect_header_row(
            union_sheet,
            ("SKU в продвижении", "SKU из объединенной карточки"),
            preferred_row=config.header_row,
        )
        promotion_column = _find_single_column(
            union_sheet,
            "SKU в продвижении",
            header_row=union_header_row,
            fallback_column=1,
        )
        related_column = _find_single_column(
            union_sheet,
            "SKU из объединенной карточки",
            header_row=union_header_row,
            fallback_column=6,
        )
        union_promotion_values = _collect_column_sku_values(
            union_sheet,
            promotion_column,
            start_row=union_header_row + 1,
        )
        union_related_values = _collect_column_sku_values(
            union_sheet,
            related_column,
            start_row=union_header_row + 1,
        )
        union_promotion_skus = frozenset(union_promotion_values)
        union_related_skus = frozenset(union_related_values)

    return AdvertisedSkus(
        direct=frozenset(direct_skus),
        union_promotion=frozenset(union_promotion_skus),
        union_related=frozenset(union_related_skus),
        direct_rows=len(direct_sku_values),
        union_promotion_rows=len(union_promotion_values),
        union_related_rows=len(union_related_values),
    )


def _detect_header_row(
    sheet: Worksheet,
    expected_headers: Sequence[str],
    *,
    preferred_row: int,
) -> int:
    normalized_expected = {_header_key(header) for header in expected_headers}
    candidate_rows = [preferred_row, 1, 2, 3, 4, 5]
    for row in dict.fromkeys(candidate_rows):
        found = {header for _, header in _iter_headers(sheet, row)}
        if normalized_expected & found:
            return row
    return preferred_row


def _parse_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return None
        return number
    text = str(value).strip()
    if not text:
        return None
    text = (
        text.replace("\u00a0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("RUB", "")
        .replace("руб.", "")
        .replace(",", ".")
    )
    try:
        number = float(text)
    except ValueError:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def calculate_discount(
    your_price: float,
    promo_price: float,
    *,
    round_to: int = 2,
) -> float:
    return round(((your_price - promo_price) / your_price) * 100, round_to)


def _find_title_column(sheet: Worksheet, *, header_row: int) -> int | None:
    return _find_optional_column(sheet, TITLE_HEADER_CANDIDATES, header_row=header_row)


def _find_article_column(sheet: Worksheet, *, header_row: int) -> int | None:
    return _find_optional_column(sheet, ARTICLE_HEADER_CANDIDATES, header_row=header_row)


def _find_category_column(sheet: Worksheet, *, header_row: int) -> int | None:
    found = _find_optional_column(sheet, CATEGORY_HEADER_CANDIDATES, header_row=header_row)
    if found:
        return found
    fallback_column = 6
    if _header_key(sheet.cell(header_row, fallback_column).value) == _header_key("Категория"):
        return fallback_column
    return None


def _find_optional_column(
    sheet: Worksheet,
    candidates: Sequence[str],
    *,
    header_row: int,
) -> int | None:
    normalized_candidates = {_header_key(name) for name in candidates}
    for column, header in _iter_headers(sheet, header_row):
        if header in normalized_candidates:
            return column
    return None


def _last_product_row(sheet: Worksheet, columns_to_check: Sequence[int], *, data_start_row: int) -> int:
    for row in range(sheet.max_row, data_start_row - 1, -1):
        for column in columns_to_check:
            if sheet.cell(row, column).value not in (None, ""):
                return row
    return data_start_row - 1


def _snapshot_locked_values(
    sheet: Worksheet,
    *,
    writable_columns: frozenset[int],
) -> dict[tuple[int, int], object]:
    snapshot: dict[tuple[int, int], object] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column not in writable_columns:
                snapshot[(cell.row, cell.column)] = cell.value
    return snapshot


def process_promo_workbook(
    promo_workbook: Workbook,
    promo_values_workbook: Workbook,
    ad_workbook: Workbook,
    config: ProcessingConfig | None = None,
) -> ProcessingResult:
    config = config or ProcessingConfig()
    if config.min_discount < -100:
        raise ProcessingError("Минимальная скидка не может быть меньше -100%")
    if config.max_discount < config.min_discount:
        raise ProcessingError("Максимальная скидка должна быть больше или равна минимальной")
    if config.target_discount_percent is not None and not (
        0 <= config.target_discount_percent < 100
    ):
        raise ProcessingError("Скидка для акции должна быть от 0 до 99.99%")
    for override in config.category_overrides:
        if override.exclude:
            continue
        if override.min_discount is not None and override.min_discount < -100:
            raise ProcessingError(
                f"Минимальная скидка для категории «{override.category}» "
                "не может быть меньше -100%"
            )
        if (
            override.min_discount is not None
            and override.max_discount is not None
            and override.max_discount < override.min_discount
        ):
            raise ProcessingError(
                f"Максимальная скидка для категории «{override.category}» "
                "должна быть больше или равна минимальной"
            )
    overrides_map = _build_category_overrides_map(config.category_overrides)
    normalized_excluded_identifiers = {
        str(item).strip().upper() for item in config.excluded_identifiers if str(item).strip()
    }
    sheet_name, columns = resolve_promo_sheet(
        promo_workbook, promo_values_workbook, config
    )
    sheet = promo_workbook[sheet_name]
    values_sheet = promo_values_workbook[sheet_name]
    title_column = _find_title_column(sheet, header_row=config.header_row)
    article_column = _find_article_column(sheet, header_row=config.header_row)
    category_column = _find_category_column(sheet, header_row=config.header_row)
    advertised_skus = collect_advertised_skus(ad_workbook, config)

    columns_to_check = [
        columns["sku"],
        columns["your_price"],
        columns["min_boost_price"],
        columns["promo_price"],
        columns["participation"],
    ]
    if title_column:
        columns_to_check.append(title_column)
    if article_column:
        columns_to_check.append(article_column)
    if category_column:
        columns_to_check.append(category_column)
    last_product_row = _last_product_row(sheet, columns_to_check, data_start_row=config.data_start_row)

    # promo_price gets written whenever a row can be added at "no natural
    # Ozon-suggested price" (discount_percent <= 0) - which happens either
    # via the legacy zero_discount_for_negative escape hatch, or whenever
    # the (global or any category-specific) minimum discount itself dips to
    # or below 0, making that zero/negative range reachable directly - or
    # whenever a fixed target discount overrides every added row's price.
    any_effective_min_at_or_below_zero = config.min_discount <= 0 or any(
        override.min_discount is not None and override.min_discount <= 0
        for override in config.category_overrides
        if not override.exclude
    )
    promo_price_may_be_written = (
        config.zero_discount_for_negative
        or config.target_discount_percent is not None
        or any_effective_min_at_or_below_zero
    )
    writable_columns = frozenset(
        {columns["participation"], columns["promo_price"]}
        if promo_price_may_be_written
        else {columns["participation"]}
    )
    before_snapshot = _snapshot_locked_values(
        sheet,
        writable_columns=writable_columns,
    )
    before_max_row = sheet.max_row

    report_rows: list[ReportRow] = []
    valid_price_rows = 0
    in_discount_range = 0
    excluded_direct_ads = 0
    excluded_union_cards = 0
    excluded_by_category = 0
    excluded_manually = 0
    added_to_promo = 0
    added_at_zero_discount = 0
    invalid_price_rows = 0
    unrecognized_skus = 0
    initial_yes_rows = 0
    removed_initial_yes_discount = 0
    removed_initial_yes_direct_ads = 0
    removed_initial_yes_union_cards = 0

    for row in range(config.data_start_row, last_product_row + 1):
        participation_cell = sheet.cell(row, columns["participation"])
        original_participation = str(participation_cell.value or "").strip()
        was_initial_yes = original_participation == "Да"
        if was_initial_yes:
            initial_yes_rows += 1
        if config.clear_non_eligible:
            participation_cell.value = None

        sku = normalize_sku(values_sheet.cell(row, columns["sku"]).value)
        article = None
        if article_column:
            article_value = values_sheet.cell(row, article_column).value
            article = str(article_value).strip() if article_value not in (None, "") else None
        product_name = None
        if title_column:
            name_value = values_sheet.cell(row, title_column).value
            product_name = str(name_value).strip() if name_value not in (None, "") else None
        category = None
        if category_column:
            category_value = values_sheet.cell(row, category_column).value
            category = str(category_value).strip() if category_value not in (None, "") else None

        your_price = _parse_number(values_sheet.cell(row, columns["your_price"]).value)
        min_boost_price = _parse_number(values_sheet.cell(row, columns["min_boost_price"]).value)
        promo_price = _parse_number(values_sheet.cell(row, columns["promo_price"]).value)
        # Keep the price Ozon itself pre-filled in the template, separate
        # from promo_price below (which gets overwritten with our own
        # decision) - so the UI can show both side by side.
        ozon_suggested_promo_price = promo_price
        discount_percent: float | None = None
        promo_discount_percent: float | None = None
        result = ""
        ad_status = _get_ad_status(sku, advertised_skus)

        override = overrides_map.get(_normalize_category_key(category)) if category else None
        effective_min_discount = (
            override.min_discount
            if override is not None and override.min_discount is not None
            else config.min_discount
        )
        effective_max_discount = (
            override.max_discount
            if override is not None and override.max_discount is not None
            else config.max_discount
        )

        is_manually_excluded = bool(
            (sku is not None and sku.strip().upper() in normalized_excluded_identifiers)
            or (
                article is not None
                and article.strip().upper() in normalized_excluded_identifiers
            )
        )

        reason = ""
        if is_manually_excluded:
            excluded_manually += 1
            reason = "Исключено вручную"
        elif override is not None and override.exclude:
            excluded_by_category += 1
            reason = "Категория исключена вручную"
        elif sku is None:
            unrecognized_skus += 1
            reason = "Отсутствует SKU"
        elif your_price is None or your_price <= 0:
            invalid_price_rows += 1
            reason = "Некорректная ваша цена"
        elif min_boost_price is None or min_boost_price <= 0:
            invalid_price_rows += 1
            reason = "Некорректная цена для минимального бустинга"
        else:
            valid_price_rows += 1
            # min_boost_price can sit at or above your_price - that yields a
            # zero/negative discount_percent below, which is a legitimate
            # value to compare against a user-chosen range that dips below
            # 0 (e.g. -3..5), not a special case on its own.
            discount_percent = calculate_discount(
                your_price,
                min_boost_price,
                round_to=config.round_discount_to,
            )
            in_range = effective_min_discount <= discount_percent <= effective_max_discount
            # Legacy escape hatch: a zero/negative-discount row that falls
            # OUTSIDE the chosen range can still be forced in at 0% via this
            # flag, independently of the range itself.
            force_zero_discount = (
                not in_range and discount_percent <= 0 and config.zero_discount_for_negative
            )
            if not in_range and not force_zero_discount:
                reason = (
                    "Скидка для входа меньше минимальной"
                    if discount_percent < effective_min_discount
                    else "Скидка для входа больше максимальной"
                )
            else:
                if in_range:
                    in_discount_range += 1
                is_direct = config.exclude_direct_ads and sku in advertised_skus.direct
                is_union_promotion = config.strict_union_exclusion and sku in advertised_skus.union_promotion
                is_union_related = config.strict_union_exclusion and sku in advertised_skus.union_related
                if is_direct:
                    excluded_direct_ads += 1
                    reason = "SKU находится в рекламе"
                elif is_union_related:
                    excluded_union_cards += 1
                    reason = "SKU связан с рекламируемой объединённой карточкой"
                elif is_union_promotion:
                    excluded_union_cards += 1
                    reason = "SKU находится в рекламе"
                else:
                    participation_cell.value = "Да"
                    if config.target_discount_percent is not None:
                        promo_price = round(
                            your_price * (1 - config.target_discount_percent / 100),
                            config.round_discount_to,
                        )
                        sheet.cell(row, columns["promo_price"]).value = promo_price
                    elif discount_percent <= 0:
                        # No natural Ozon-suggested promo price makes sense
                        # once min_boost_price already reached your_price -
                        # participate at the regular price instead.
                        promo_price = your_price
                        sheet.cell(row, columns["promo_price"]).value = promo_price
                        added_at_zero_discount += 1
                    added_to_promo += 1
                    result = "Да"
                    if discount_percent > 0:
                        reason = "Добавлен в акцию"
                    elif config.target_discount_percent is not None:
                        reason = "Добавлен в акцию со своей скидкой (цена уже не выше минимального бустинга)"
                    else:
                        reason = "Добавлен в акцию без скидки (цена уже не выше минимального бустинга)"

        if your_price is not None and your_price > 0 and promo_price is not None and 0 < promo_price < your_price:
            promo_discount_percent = calculate_discount(
                your_price,
                promo_price,
                round_to=config.round_discount_to,
            )

        report_rows.append(
            ReportRow(
                row_number=row,
                sku=sku,
                article=article,
                product_name=product_name,
                category=category,
                your_price=your_price,
                min_boost_price=min_boost_price,
                ozon_suggested_promo_price=ozon_suggested_promo_price,
                promo_price=promo_price,
                discount_percent=discount_percent,
                promo_discount_percent=promo_discount_percent,
                original_participation=original_participation,
                ad_status=ad_status,
                result=result,
                reason=reason,
            )
        )
        if was_initial_yes and result != "Да":
            if reason == "SKU находится в рекламе":
                removed_initial_yes_direct_ads += 1
            elif reason == "SKU связан с рекламируемой объединённой карточкой":
                removed_initial_yes_union_cards += 1
            else:
                removed_initial_yes_discount += 1

    summary = ProcessingSummary(
        total_product_rows=max(0, last_product_row - config.data_start_row + 1),
        valid_price_rows=valid_price_rows,
        in_discount_range=in_discount_range,
        excluded_direct_ads=excluded_direct_ads,
        excluded_union_cards=excluded_union_cards,
        added_to_promo=added_to_promo,
        added_at_zero_discount=added_at_zero_discount,
        invalid_price_rows=invalid_price_rows,
        unrecognized_skus=unrecognized_skus,
        initial_yes_rows=initial_yes_rows,
        removed_initial_yes_discount=removed_initial_yes_discount,
        removed_initial_yes_direct_ads=removed_initial_yes_direct_ads,
        removed_initial_yes_union_cards=removed_initial_yes_union_cards,
        direct_ad_rows=advertised_skus.direct_rows,
        direct_ad_skus_unique=len(advertised_skus.direct),
        strict_ad_skus_unique=len(advertised_skus.strict_all),
        excluded_by_category=excluded_by_category,
        excluded_manually=excluded_manually,
    )
    result = ProcessingResult(
        workbook=promo_workbook,
        summary=summary,
        report_rows=tuple(report_rows),
        output_filename=_make_xlsx_filename(),
        csv_filename=_make_csv_filename(),
    )

    validate_result(
        result,
        promo_workbook=promo_workbook,
        sheet_name=sheet_name,
        before_locked_snapshot=before_snapshot,
        before_max_row=before_max_row,
        participation_column=columns["participation"],
        advertised_skus=(
            (advertised_skus.direct if config.exclude_direct_ads else frozenset())
            | (
                advertised_skus.union_promotion | advertised_skus.union_related
                if config.strict_union_exclusion
                else frozenset()
            )
        ),
        sku_column=columns["sku"],
    )
    return result


def _get_ad_status(sku: str | None, advertised_skus: AdvertisedSkus) -> str:
    if not sku:
        return ""
    if sku in advertised_skus.direct:
        return "Прямая реклама"
    if sku in advertised_skus.union_related:
        return "Объединённая карточка"
    if sku in advertised_skus.union_promotion:
        return "SKU в продвижении Union"
    return ""


def validate_result(
    result: ProcessingResult,
    *,
    promo_workbook: Workbook,
    sheet_name: str,
    before_locked_snapshot: Mapping[tuple[int, int], object],
    before_max_row: int,
    participation_column: int,
    advertised_skus: frozenset[str],
    sku_column: int,
) -> None:
    sheet = promo_workbook[sheet_name]
    if sheet.max_row != before_max_row:
        raise ProcessingError("Количество строк в готовом шаблоне отличается от исходного")

    for (row, column), old_value in before_locked_snapshot.items():
        if sheet.cell(row, column).value != old_value:
            column_letter = get_column_letter(column)
            raise ProcessingError(
                "Проверка безопасности не пройдена: изменена защищённая ячейка "
                f"{sheet.title}!{column_letter}{row}"
            )

    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, participation_column).value == "Да":
            sku = normalize_sku(sheet.cell(row, sku_column).value)
            if sku in advertised_skus:
                raise ProcessingError(f"Рекламируемый SKU {sku} ошибочно отмечен «Да»")


def create_processing_report(report_rows: Sequence[ReportRow]) -> str:
    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(REPORT_HEADERS)
    for report_row in report_rows:
        writer.writerow(report_row.as_csv_row())
    return output.getvalue()


def _make_xlsx_filename(now: datetime | None = None) -> str:
    now = now or datetime.now()
    return f"Эластичный_бустинг_готово_{now:%Y-%m-%d_%H-%M}.xlsx"


def _make_csv_filename(now: datetime | None = None) -> str:
    now = now or datetime.now()
    return f"Эластичный_бустинг_отчёт_{now:%Y-%m-%d_%H-%M}.csv"


def save_result(result: ProcessingResult, output_dir: str | Path | None = None) -> tuple[bytes, Path | None]:
    buffer = BytesIO()
    try:
        result.workbook.save(buffer)
        buffer.seek(0)
        load_workbook(buffer, read_only=True).close()
    except Exception as exc:  # pragma: no cover - openpyxl error details vary
        logger.exception("Failed to save or reopen result workbook")
        raise ProcessingError(f"Готовый XLSX-файл не удалось сохранить или открыть: {exc}") from exc

    data = buffer.getvalue()
    if output_dir is None:
        return data, None

    path = Path(output_dir) / result.output_filename
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return data, path
