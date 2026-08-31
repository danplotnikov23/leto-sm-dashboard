import csv
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook


CANONICAL_COLUMNS = [
    "sku",
    "Артикул",
    "Название товара",
    "Цена товара, ₽",
    "Показы",
    "Клики",
    "CTR (%)",
    "В корзину",
    "Средняя стоимость клика, ₽",
    "Расход, ₽, с НДС",
    "Заказы",
    "Продажи, ₽",
    "Заказы модели",
    "Продажи с заказов модели, ₽",
    "ДРР, %",
    "Заказано на сумму, ₽",
    "Общий ДРР",
    "Дата добавления",
]


@dataclass(frozen=True)
class ParsedOzonAdReportFile:
    rows: list[dict[str, str]]
    raw_report_csv: str


class OzonAdReportFileParser:
    def parse(self, filename: str, content: bytes) -> ParsedOzonAdReportFile:
        suffix = Path(filename).suffix.lower()
        if suffix in {".csv", ".txt"}:
            rows = self._parse_csv(content)
        elif suffix in {".xlsx", ".xlsm"}:
            rows = self._parse_xlsx(content)
        else:
            raise ValueError("Unsupported Ozon report format. Use CSV or XLSX.")

        if not rows:
            raise ValueError("Ozon report has no readable rows")

        return ParsedOzonAdReportFile(
            rows=rows,
            raw_report_csv=_rows_to_csv(rows),
        )

    def _parse_csv(self, content: bytes) -> list[dict[str, str]]:
        text = _decode_text(content)
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            return []

        delimiter = _detect_delimiter(lines)
        parsed_lines = [next(csv.reader([line], delimiter=delimiter)) for line in lines]
        header_index = _find_header_index(parsed_lines)
        headers = parsed_lines[header_index]

        rows: list[dict[str, str]] = []
        for raw_values in parsed_lines[header_index + 1 :]:
            row = _normalize_row(headers, raw_values)
            if any(value for value in row.values()):
                rows.append(row)

        return rows

    def _parse_xlsx(self, content: bytes) -> list[dict[str, str]]:
        with NamedTemporaryFile(suffix=".xlsx") as file:
            file.write(content)
            file.flush()
            workbook = load_workbook(file.name, read_only=True, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            raw_rows = [
                [_cell_to_string(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]

        header_index = _find_header_index(raw_rows)
        headers = raw_rows[header_index]
        rows: list[dict[str, str]] = []
        for raw_values in raw_rows[header_index + 1 :]:
            row = _normalize_row(headers, raw_values)
            if any(value for value in row.values()):
                rows.append(row)

        return rows


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue

    return content.decode("utf-8-sig", errors="replace")


def _detect_delimiter(lines: list[str]) -> str:
    candidates = [";", ",", "\t"]
    return max(candidates, key=lambda delimiter: max(line.count(delimiter) for line in lines))


def _find_header_index(rows: list[list[str]]) -> int:
    for index, row in enumerate(rows):
        if is_ozon_ad_report_header(row):
            return index

    raise ValueError("Ozon report header row was not found")


def _normalize_row(headers: list[str], values: list[str]) -> dict[str, str]:
    row = {column: "" for column in CANONICAL_COLUMNS}
    for index, header in enumerate(headers):
        canonical = _canonical_column_name(header)
        if canonical is None:
            continue

        value = values[index] if index < len(values) else ""
        row[canonical] = _cell_to_string(value)

    return row


def _canonical_column_name(header: str) -> str | None:
    normalized = _normalize_header(header)
    if normalized == "sku":
        return "sku"
    if "артикул" in normalized:
        return "Артикул"
    if "название" in normalized or "товар" == normalized:
        return "Название товара"
    if "цена" in normalized:
        return "Цена товара, ₽"
    if "показы" in normalized:
        return "Показы"
    if "клики" in normalized:
        return "Клики"
    if normalized.startswith("ctr"):
        return "CTR (%)"
    if "корзин" in normalized:
        return "В корзину"
    if "средняя стоимость клика" in normalized:
        return "Средняя стоимость клика, ₽"
    if "расход" in normalized:
        return "Расход, ₽, с НДС"
    if (
        "продажи с заказов модели" in normalized
        or "продажи в продвижении с заказов модели" in normalized
    ):
        return "Продажи с заказов модели, ₽"
    if "заказы модели" in normalized or normalized == "продано товаров модели":
        return "Заказы модели"
    if normalized in {"заказы", "продано товаров"}:
        return "Заказы"
    if normalized in {
        "продажи",
        "продажи ₽",
        "продажи в продвижении",
        "продажи в продвижении ₽",
    }:
        return "Продажи, ₽"
    if normalized in {"дрр", "дрр %", "дрр в продвижении", "дрр в продвижении %"}:
        return "ДРР, %"
    if "заказано на сумму" in normalized:
        return "Заказано на сумму, ₽"
    if "общий дрр" in normalized or ("дрр" in normalized and "общий" in normalized):
        return "Общий ДРР"
    if "дата добавления" in normalized:
        return "Дата добавления"

    return None


def canonicalize_ozon_ad_report_header(header: str) -> str:
    return _canonical_column_name(header) or header.strip()


def is_ozon_ad_report_header(headers: list[str]) -> bool:
    canonical_headers = {
        canonicalize_ozon_ad_report_header(header) for header in headers
    }
    return {
        "sku",
        "Расход, ₽, с НДС",
        "Заказы",
    }.issubset(canonical_headers)


def _normalize_header(value: str) -> str:
    normalized = (
        str(value or "")
        .replace("\n", " ")
        .replace(",", " ")
        .replace("₽", "₽")
        .strip()
        .lower()
    )
    return " ".join(normalized.split())


def _cell_to_string(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _rows_to_csv(rows: list[dict[str, str]]) -> str:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=CANONICAL_COLUMNS, delimiter=";")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()
