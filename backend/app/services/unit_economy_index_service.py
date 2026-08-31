import json
import re
import shutil
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from app.schemas.unit_economy_index import (
    UnitEconomyIndexSummary,
    UnitEconomyProduct,
    UnitEconomyVersionSummary,
)


_FILENAME_DATE_PATTERN = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4}|\d{2})")


def parse_date_from_filename(filename: str) -> date | None:
    """Best-effort ``DD.MM.YY`` / ``DD.MM.YYYY`` extraction from a workbook
    filename (e.g. "Юнит экономика 29.07.26.xlsx" -> 2026-07-29). Real
    filenames mix both year formats, so this is a suggestion for the
    upload form to prefill - never applied silently as the version's
    "valid_from"."""
    matches = list(_FILENAME_DATE_PATTERN.finditer(filename))
    if not matches:
        return None

    day_text, month_text, year_text = matches[-1].groups()
    year = int(year_text)
    if year < 100:
        year += 2000
    try:
        return date(year, int(month_text), int(day_text))
    except ValueError:
        return None


HEADER_ROW = 2
DATA_START_ROW = 4
REQUIRED_COLUMNS = (
    "артикул",
    "Себес (затраты)",
    "Себес (затраты) + Комиссия Озон",
    "Чистая прибыль, руб.",
)


@dataclass(frozen=True)
class UnitEconomyWorkbookVersion:
    path: str
    modified_at: str
    size_bytes: int
    version_id: str
    valid_from: str
    sheet_name: str


@dataclass(frozen=True)
class UnitEconomyVersionSelection:
    version: UnitEconomyWorkbookVersion
    warning: str | None


@dataclass(frozen=True)
class UnitEconomyPeriodSegment:
    date_from: str
    date_to: str
    version: UnitEconomyWorkbookVersion


@dataclass(frozen=True)
class _UnitEconomyVersionConfig:
    valid_from: date
    workbook_path: Path
    sheet_name: str | None = None


class _UnitEconomyVersionIndex:
    def __init__(self, config: _UnitEconomyVersionConfig) -> None:
        self.config = config
        self.sheet_name: str = ""
        self.products_by_offer_id: dict[str, UnitEconomyProduct] = {}
        self.products_by_sku: dict[str, UnitEconomyProduct] = {}
        self.row_count = 0
        self._load()

    def find_by_offer_id(self, offer_id: str) -> UnitEconomyProduct | None:
        return self.products_by_offer_id.get(_normalize_key(offer_id))

    def find_by_sku(self, sku: str) -> UnitEconomyProduct | None:
        return self.products_by_sku.get(_normalize_key(sku))

    def get_workbook_version(self) -> UnitEconomyWorkbookVersion:
        stat = self.config.workbook_path.stat()
        modified_at = datetime.fromtimestamp(stat.st_mtime, UTC).isoformat()
        return UnitEconomyWorkbookVersion(
            path=str(self.config.workbook_path),
            modified_at=modified_at,
            size_bytes=stat.st_size,
            version_id=(
                f"{self.config.valid_from.isoformat()}:"
                f"{self.config.workbook_path.name}:"
                f"{stat.st_mtime_ns}:{stat.st_size}"
            ),
            valid_from=self.config.valid_from.isoformat(),
            sheet_name=self.sheet_name,
        )

    def get_summary(self) -> UnitEconomyVersionSummary:
        version = self.get_workbook_version()
        return UnitEconomyVersionSummary(
            valid_from=version.valid_from,
            sheet_name=version.sheet_name,
            workbook_path=version.path,
            row_count=self.row_count,
            indexed_offer_ids=len(self.products_by_offer_id),
            indexed_skus=len(self.products_by_sku),
            version_id=version.version_id,
        )

    def _load(self) -> None:
        workbook = load_workbook(self.config.workbook_path, read_only=True, data_only=True)
        worksheet = _select_unit_economy_sheet(workbook.worksheets, self.config.sheet_name)
        self.sheet_name = worksheet.title
        columns = _build_column_map(worksheet)
        rows = worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True)

        for offset, row in enumerate(rows):
            row_number = DATA_START_ROW + offset
            offer_id = _row_string(row, columns["артикул"])
            if offer_id is None:
                continue

            product = UnitEconomyProduct(
                row_number=row_number,
                offer_id=offer_id,
                ozon_sku=_row_string(row, columns["Ozon SKU ID"]),
                title=_row_string(row, columns["Название"]),
                category=(
                    _row_string(row, columns["Тип"]) if "Тип" in columns else None
                ),
                sale_schema=_row_string(row, columns["Схема реализации"]),
                price_before_discount_with_vat=(
                    _cell_float(row, columns["Цена до скидки с НДС, руб."])
                    if "Цена до скидки с НДС, руб." in columns
                    else None
                ),
                price_with_vat=_cell_float(
                    row,
                    columns["Цена (со скидкой) с НДС, руб."],
                ),
                price_without_vat=_cell_float(
                    row,
                    columns["Цена (со скидкой) без НДС, руб."],
                ),
                cost_without_vat=_cell_float(
                    row,
                    _column(
                        columns,
                        "Закупочная цена без НДС, руб.",
                        "Себестоимость без НДС, руб.",
                    ),
                ),
                ozon_commission=_cell_float(
                    row,
                    columns["Вознаграждение OZON (от цены реализации), руб."],
                ),
                ad_cost=_cell_float(
                    row,
                    columns["Затраты на рекламу и продвижение, руб."],
                ),
                expense_cost=_cell_float(row, columns["Себес (затраты)"]),
                expense_with_ozon_commission=_cell_float(
                    row,
                    columns["Себес (затраты) + Комиссия Озон"],
                ),
                logistics_compensation=(
                    _cell_float(row, columns["Компенсация логистики"])
                    if "Компенсация логистики" in columns
                    else None
                ),
                profit_before_tax=_cell_float(
                    row,
                    columns["Прибыль до налогообложения, руб."],
                ),
                tax=_cell_float(row, columns["Налоги, руб."]),
                net_profit=_cell_float(row, columns["Чистая прибыль, руб."]),
                profitability=_cell_float(row, columns["Рентабельность, %"]),
            )

            self.products_by_offer_id[_normalize_key(product.offer_id)] = product
            if product.ozon_sku:
                self.products_by_sku[_normalize_key(product.ozon_sku)] = product

        self.row_count = len(self.products_by_offer_id)


class UnitEconomyIndexService:
    def __init__(
        self,
        workbook_path: Path,
        workbook_versions: str | None = None,
        versions_dir: Path | None = None,
    ) -> None:
        self._workbook_path = workbook_path
        self._workbook_versions = workbook_versions
        self._versions_dir = versions_dir
        self._versions = self._load_versions()

    def _manifest_path(self) -> Path | None:
        return self._versions_dir / "manifest.json" if self._versions_dir else None

    def _load_versions(self) -> list["_UnitEconomyVersionIndex"]:
        manifest_path = self._manifest_path()
        if manifest_path is not None and manifest_path.exists():
            configs = _load_manifest_configs(manifest_path)
        else:
            configs = _parse_version_configs(self._workbook_versions, self._workbook_path)
        versions = [
            _UnitEconomyVersionIndex(config)
            for config in sorted(configs, key=lambda item: item.valid_from)
        ]
        if not versions:
            raise ValueError("At least one unit-economy workbook version is required")
        return versions

    def reload(self) -> None:
        """Re-read every configured workbook version from disk.

        Each version is loaded into memory once at process startup (or the
        last reload) - editing the underlying xlsx on disk otherwise has no
        effect until this runs, which is what produced stale unit-economy
        numbers (e.g. "Себес (затраты)") after the workbook was updated and
        re-saved without the app being told about it.
        """
        self._versions = self._load_versions()

    def add_version(
        self,
        workbook_bytes: bytes,
        valid_from: date,
        original_filename: str,
    ) -> UnitEconomyIndexSummary:
        """Validate and register a new workbook version, effective from
        ``valid_from`` (until whichever later version's own valid_from
        starts, or indefinitely if this is the newest one).

        The upload is validated against a temp file before anything live is
        touched, so a bad upload can never leave the service without
        working data. Re-uploading the same ``valid_from`` replaces that
        version's file rather than creating a duplicate entry.
        """
        if self._versions_dir is None:
            raise ValueError("Unit-economy versions directory is not configured")

        self._versions_dir.mkdir(parents=True, exist_ok=True)
        temp_path = self._versions_dir / f".upload-tmp-{original_filename}"
        temp_path.write_bytes(workbook_bytes)
        try:
            _UnitEconomyVersionIndex(
                _UnitEconomyVersionConfig(valid_from=date.min, workbook_path=temp_path)
            )
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise

        stored_filename = _safe_version_filename(valid_from, original_filename)
        stored_path = self._versions_dir / stored_filename
        if stored_path.exists():
            timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
            backup_path = stored_path.with_name(f"{stored_path.stem}.backup-{timestamp}{stored_path.suffix}")
            shutil.copy2(stored_path, backup_path)
        temp_path.replace(stored_path)

        manifest_path = self._manifest_path()
        entries = _read_manifest(manifest_path) if manifest_path else []
        entries = [entry for entry in entries if entry["valid_from"] != valid_from.isoformat()]
        entries.append({"valid_from": valid_from.isoformat(), "filename": stored_filename})
        entries.sort(key=lambda entry: entry["valid_from"])
        _write_manifest(manifest_path, entries)

        self.reload()
        return self.get_summary()

    def remove_version(self, valid_from: date) -> UnitEconomyIndexSummary:
        """Drop a version from the manifest/rotation (e.g. a stale queued
        future upload superseded by a corrected one). The underlying xlsx
        file on disk is left in place - only the manifest entry goes away -
        so this can't destroy data, only stop that version from being used.
        """
        if self._versions_dir is None:
            raise ValueError("Unit-economy versions directory is not configured")

        manifest_path = self._manifest_path()
        entries = _read_manifest(manifest_path) if manifest_path else []
        remaining = [entry for entry in entries if entry["valid_from"] != valid_from.isoformat()]
        if len(remaining) == len(entries):
            raise ValueError(f"Версия от {valid_from.isoformat()} не найдена")
        if not remaining:
            raise ValueError("Нельзя удалить единственную оставшуюся версию юнитки")

        _write_manifest(manifest_path, remaining)
        self.reload()
        return self.get_summary()

    def get_summary(self) -> UnitEconomyIndexSummary:
        active = self._latest_version
        return UnitEconomyIndexSummary(
            sheet_name=active.sheet_name,
            row_count=active.row_count,
            indexed_offer_ids=len(active.products_by_offer_id),
            indexed_skus=len(active.products_by_sku),
            active_version=active.get_workbook_version().version_id,
            versions=[version.get_summary() for version in self._versions],
        )

    def find_by_offer_id(self, offer_id: str) -> UnitEconomyProduct | None:
        selected_product = self._latest_version.find_by_offer_id(offer_id)
        if _has_unit_expense(selected_product):
            return selected_product

        return (
            self._find_by_offer_id_in_fallback_versions(
                offer_id,
                self._latest_version,
            )
            or selected_product
        )

    def find_by_sku(self, sku: str) -> UnitEconomyProduct | None:
        selected_product = self._latest_version.find_by_sku(sku)
        if _has_unit_expense(selected_product):
            return selected_product

        return (
            self._find_by_sku_in_fallback_versions(
                sku,
                self._latest_version,
            )
            or selected_product
        )

    def list_products(self, limit: int | None = None) -> list[UnitEconomyProduct]:
        products = list(self._latest_version.products_by_offer_id.values())
        if limit is None:
            return products

        return products[:limit]

    def find_by_offer_id_for_date(
        self,
        offer_id: str,
        effective_date: str | date,
    ) -> UnitEconomyProduct | None:
        selected_version = self._version_for_date(effective_date)
        selected_product = selected_version.find_by_offer_id(offer_id)
        if _has_unit_expense(selected_product):
            return selected_product

        return (
            self._find_by_offer_id_in_fallback_versions(offer_id, selected_version)
            or selected_product
        )

    def find_by_sku_for_date(
        self,
        sku: str,
        effective_date: str | date,
    ) -> UnitEconomyProduct | None:
        selected_version = self._version_for_date(effective_date)
        selected_product = selected_version.find_by_sku(sku)
        if _has_unit_expense(selected_product):
            return selected_product

        return self._find_by_sku_in_fallback_versions(sku, selected_version) or selected_product

    def select_version_for_period(
        self,
        date_from: str,
        date_to: str,
    ) -> UnitEconomyVersionSelection:
        start = _parse_iso_date(date_from)
        end = _parse_iso_date(date_to)
        selected = self._version_for_date(end)
        selected_version = selected.get_workbook_version()
        crossed_versions = [
            version
            for version in self._versions
            if start < version.config.valid_from <= end
        ]
        warning = None
        if crossed_versions:
            valid_from_values = ", ".join(
                version.config.valid_from.isoformat()
                for version in crossed_versions
            )
            warning = (
                "Период пересекает смену юнитки "
                f"({valid_from_values}). Сейчас отчёт Ozon агрегирован по SKU, "
                f"поэтому применена версия юнитки на дату окончания периода: "
                f"{selected_version.valid_from}. Для точного расчёта нужен отчёт "
                "с дневной детализацией."
            )

        return UnitEconomyVersionSelection(version=selected_version, warning=warning)

    def build_period_segments(
        self,
        date_from: str,
        date_to: str,
    ) -> list[UnitEconomyPeriodSegment]:
        start = _parse_iso_date(date_from)
        end = _parse_iso_date(date_to)
        if start > end:
            raise ValueError("date_from must be earlier than or equal to date_to")

        segments: list[UnitEconomyPeriodSegment] = []
        segment_start = start
        change_dates = [
            version.config.valid_from
            for version in self._versions
            if start < version.config.valid_from <= end
        ]

        for change_date in change_dates:
            segment_end = change_date - timedelta(days=1)
            if segment_start <= segment_end:
                segment_version = self._version_for_date(segment_start).get_workbook_version()
                segments.append(
                    UnitEconomyPeriodSegment(
                        date_from=segment_start.isoformat(),
                        date_to=segment_end.isoformat(),
                        version=segment_version,
                    )
                )
            segment_start = change_date

        segment_version = self._version_for_date(segment_start).get_workbook_version()
        segments.append(
            UnitEconomyPeriodSegment(
                date_from=segment_start.isoformat(),
                date_to=end.isoformat(),
                version=segment_version,
            )
        )

        return segments

    def get_workbook_version(
        self,
        effective_date: str | date | None = None,
    ) -> UnitEconomyWorkbookVersion:
        if effective_date is None:
            return self._latest_version.get_workbook_version()

        return self._version_for_date(effective_date).get_workbook_version()

    @property
    def _latest_version(self) -> _UnitEconomyVersionIndex:
        return self._versions[-1]

    def _version_for_date(self, effective_date: str | date) -> _UnitEconomyVersionIndex:
        parsed_date = (
            effective_date
            if isinstance(effective_date, date)
            else _parse_iso_date(effective_date)
        )
        selected = self._versions[0]
        for version in self._versions:
            if version.config.valid_from <= parsed_date:
                selected = version
            else:
                break

        return selected

    def _find_by_offer_id_in_fallback_versions(
        self,
        offer_id: str,
        selected_version: _UnitEconomyVersionIndex,
    ) -> UnitEconomyProduct | None:
        fallback_product: UnitEconomyProduct | None = None
        for version in reversed(self._versions):
            if (
                version is selected_version
                or version.config.valid_from > selected_version.config.valid_from
            ):
                continue

            product = version.find_by_offer_id(offer_id)
            if _has_unit_expense(product):
                return product
            if fallback_product is None:
                fallback_product = product

        return fallback_product

    def _find_by_sku_in_fallback_versions(
        self,
        sku: str,
        selected_version: _UnitEconomyVersionIndex,
    ) -> UnitEconomyProduct | None:
        fallback_product: UnitEconomyProduct | None = None
        for version in reversed(self._versions):
            if (
                version is selected_version
                or version.config.valid_from > selected_version.config.valid_from
            ):
                continue

            product = version.find_by_sku(sku)
            if _has_unit_expense(product):
                return product
            if fallback_product is None:
                fallback_product = product

        return fallback_product


_SAFE_FILENAME_PATTERN = re.compile(r"[^0-9A-Za-zА-Яа-яЁё._() -]+")


def _safe_version_filename(valid_from: date, original_filename: str) -> str:
    clean_name = _SAFE_FILENAME_PATTERN.sub("_", Path(original_filename).name).strip(" .")
    return f"{valid_from.isoformat()}_{clean_name}" if clean_name else f"{valid_from.isoformat()}.xlsx"


def _read_manifest(manifest_path: Path) -> list[dict[str, str]]:
    if not manifest_path.exists():
        return []
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def _write_manifest(manifest_path: Path, entries: list[dict[str, str]]) -> None:
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _load_manifest_configs(manifest_path: Path) -> list[_UnitEconomyVersionConfig]:
    entries = _read_manifest(manifest_path)
    base_dir = manifest_path.parent
    return [
        _UnitEconomyVersionConfig(
            valid_from=_parse_iso_date(entry["valid_from"]),
            workbook_path=base_dir / entry["filename"],
            sheet_name=entry.get("sheet_name"),
        )
        for entry in entries
    ]


def _parse_version_configs(
    value: str | None,
    fallback_workbook_path: Path,
) -> list[_UnitEconomyVersionConfig]:
    if value is None or not value.strip():
        return [
            _UnitEconomyVersionConfig(
                valid_from=date.min,
                workbook_path=fallback_workbook_path,
                sheet_name=None,
            )
        ]

    configs: list[_UnitEconomyVersionConfig] = []
    for raw_item in value.split(","):
        item = raw_item.strip()
        if not item:
            continue

        parts = item.split("=", 2)
        if len(parts) < 2:
            raise ValueError(
                "UNIT_ECONOMY_WORKBOOK_VERSIONS entries must use "
                "valid_from=path or valid_from=path=sheet_name"
            )

        valid_from_text, path_text = parts[0], parts[1]
        sheet_name = parts[2] if len(parts) == 3 and parts[2] else None
        configs.append(
            _UnitEconomyVersionConfig(
                valid_from=_parse_iso_date(valid_from_text),
                workbook_path=Path(path_text),
                sheet_name=sheet_name,
            )
        )

    return configs


def _select_unit_economy_sheet(
    worksheets: list[ReadOnlyWorksheet],
    preferred_sheet_name: str | None,
) -> ReadOnlyWorksheet:
    if preferred_sheet_name:
        for worksheet in worksheets:
            if worksheet.title == preferred_sheet_name:
                return worksheet

        raise ValueError(f"Unit-economy sheet was not found: {preferred_sheet_name}")

    for worksheet in worksheets:
        if _worksheet_has_required_columns(worksheet):
            return worksheet

    raise ValueError("Unit-economy workbook does not contain a compatible sheet")


def _worksheet_has_required_columns(worksheet: ReadOnlyWorksheet) -> bool:
    try:
        columns = _build_column_map(worksheet)
    except StopIteration:
        return False

    return all(column in columns for column in REQUIRED_COLUMNS)


def _build_column_map(worksheet: ReadOnlyWorksheet) -> dict[str, int]:
    columns: dict[str, int] = {}
    header_values = next(
        worksheet.iter_rows(
            min_row=HEADER_ROW,
            max_row=HEADER_ROW,
            values_only=True,
        )
    )
    for col, value in enumerate(header_values, start=1):
        if value is None:
            continue

        columns[str(value).strip()] = col

    return columns


def _column(columns: dict[str, int], *names: str) -> int:
    for name in names:
        column = columns.get(name)
        if column is not None:
            return column

    raise KeyError(f"None of the unit-economy columns were found: {names}")


def _row_string(row: tuple[object, ...], col: int) -> str | None:
    value = _row_value(row, col)
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _cell_float(row: tuple[object, ...], col: int) -> float | None:
    value = _row_value(row, col)
    if value is None or value == "":
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _has_unit_expense(product: UnitEconomyProduct | None) -> bool:
    return (
        product is not None
        and product.expense_with_ozon_commission is not None
    )


def _normalize_key(value: str) -> str:
    return value.strip().lower()


def _row_value(row: tuple[object, ...], col: int) -> object | None:
    index = col - 1
    if index < 0 or index >= len(row):
        return None

    return row[index]


def _parse_iso_date(value: str) -> date:
    return date.fromisoformat(value)
