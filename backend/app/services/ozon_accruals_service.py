from collections import defaultdict
import logging
from pathlib import Path
from typing import DefaultDict
import warnings

from openpyxl import load_workbook

from app.schemas.ozon_accruals import OzonAccrualArticleSummary, OzonAccrualsSummary


SHEET_NAME = "Начисления"
PERIOD_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3
ACCOUNTING_NOTE = (
    "Accruals are accounting facts for the report period. They must not be "
    "joined one-to-one with advertising orders for the same calendar period "
    "because orders, delivery, returns, and accrual recognition can lag."
)
logger = logging.getLogger(__name__)


class OzonAccrualsService:
    def __init__(self, report_path: Path) -> None:
        self._report_path = report_path
        self._period: str | None = None
        self._row_count = 0
        self._articles: dict[str, OzonAccrualArticleSummary] = {}
        self._totals: DefaultDict[str, float] = defaultdict(float)
        self._load()

    def get_summary(self) -> OzonAccrualsSummary:
        return OzonAccrualsSummary(
            report_path=str(self._report_path),
            row_count=self._row_count,
            article_count=len(self._articles),
            period=self._period,
            totals=dict(sorted(self._totals.items())),
            accounting_note=ACCOUNTING_NOTE,
        )

    def find_by_offer_id(self, offer_id: str) -> OzonAccrualArticleSummary | None:
        return self._articles.get(_normalize_key(offer_id))

    def _load(self) -> None:
        if not self._report_path.exists():
            logger.warning("Ozon accruals report was not found: %s", self._report_path)
            return

        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            workbook = load_workbook(self._report_path, read_only=True, data_only=True)

        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        first_row = next(rows)
        self._period = str(first_row[0]) if first_row and first_row[0] else None

        headers = next(rows)
        columns = _build_column_map(headers)
        builders: dict[str, _ArticleSummaryBuilder] = {}

        for row in rows:
            if not any(value is not None and value != "" for value in row):
                continue

            self._row_count += 1
            amount = _row_float(row, columns["Сумма итого, руб."])
            group = _row_string(row, columns["Группа услуг"]) or ""
            charge_type = _row_string(row, columns["Тип начисления"]) or ""
            self._totals[group] += amount

            offer_id = _row_string(row, columns["Артикул"])
            if offer_id is None:
                continue

            key = _normalize_key(offer_id)
            builder = builders.get(key)
            if builder is None:
                builder = _ArticleSummaryBuilder(offer_id=offer_id)
                builders[key] = builder

            builder.add(
                sku=_row_string(row, columns["SKU"]),
                title=_row_string(row, columns["Название товара"]),
                quantity=_row_int(row, columns["Количество"]),
                group=group,
                charge_type=charge_type,
                amount=amount,
            )

        self._articles = {key: builder.build() for key, builder in builders.items()}


class _ArticleSummaryBuilder:
    def __init__(self, offer_id: str) -> None:
        self.offer_id = offer_id
        self.sku: str | None = None
        self.title: str | None = None
        self.sold_quantity = 0
        self.sales_total = 0.0
        self.revenue = 0.0
        self.discount_points = 0.0
        self.partner_programs = 0.0
        self.ozon_commission = 0.0
        self.acquiring = 0.0
        self.ad_charges = 0.0
        self.returns = 0.0
        self.delivery_services = 0.0
        self.other_services = 0.0
        self.compensations = 0.0
        self.total = 0.0
        self.breakdown: DefaultDict[str, float] = defaultdict(float)

    def add(
        self,
        sku: str | None,
        title: str | None,
        quantity: int,
        group: str,
        charge_type: str,
        amount: float,
    ) -> None:
        self.sku = self.sku or sku
        self.title = self.title or title
        self.total += amount
        self.breakdown[f"{group}|{charge_type}"] += amount

        if group == "Продажи":
            self.sales_total += amount
            if charge_type == "Выручка":
                self.sold_quantity += quantity
                self.revenue += amount
            elif charge_type == "Баллы за скидки":
                self.discount_points += amount
            elif charge_type == "Программы партнёров":
                self.partner_programs += amount
        elif group == "Вознаграждение Ozon":
            self.ozon_commission += amount
        elif group == "Услуги партнёров" and charge_type == "Эквайринг":
            self.acquiring += amount
        elif group == "Продвижение и реклама":
            self.ad_charges += amount
        elif group == "Возвраты":
            self.returns += amount
        elif group == "Услуги доставки":
            self.delivery_services += amount
        elif group == "Компенсации и декомпенсации":
            self.compensations += amount
        elif group in {"Другие услуги и штрафы", "Услуги FBO", "Услуги партнёров"}:
            self.other_services += amount

    def build(self) -> OzonAccrualArticleSummary:
        return OzonAccrualArticleSummary(
            offer_id=self.offer_id,
            sku=self.sku,
            title=self.title,
            sold_quantity=self.sold_quantity,
            sales_total=self.sales_total,
            revenue=self.revenue,
            discount_points=self.discount_points,
            partner_programs=self.partner_programs,
            ozon_commission=self.ozon_commission,
            acquiring=self.acquiring,
            ad_charges=self.ad_charges,
            returns=self.returns,
            delivery_services=self.delivery_services,
            other_services=self.other_services,
            compensations=self.compensations,
            total=self.total,
            breakdown=dict(sorted(self.breakdown.items())),
        )


def _build_column_map(headers: tuple[object, ...]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, value in enumerate(headers):
        if value is None:
            continue

        columns[str(value).strip()] = index

    return columns


def _row_string(row: tuple[object, ...], index: int) -> str | None:
    value = _row_value(row, index)
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _row_float(row: tuple[object, ...], index: int) -> float:
    value = _row_value(row, index)
    if value is None or value == "":
        return 0.0

    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _row_int(row: tuple[object, ...], index: int) -> int:
    value = _row_value(row, index)
    if value is None or value == "":
        return 0

    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _row_value(row: tuple[object, ...], index: int) -> object | None:
    if index < 0 or index >= len(row):
        return None

    return row[index]


def _normalize_key(value: str) -> str:
    return value.strip().lower()
