from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Final

from openpyxl import load_workbook


STATISTICS_SHEET: Final = "Statistics"
UNION_SHEET: Final = "Union"


@dataclass(frozen=True, slots=True)
class PromotionStatisticsRow:
    promoted_sku: str
    title: str | None
    campaign_id: str
    instrument: str | None
    placement: str | None
    spend_with_vat: Decimal
    direct_revenue_with_vat: Decimal
    direct_orders: int
    model_revenue_with_vat: Decimal
    model_orders: int


@dataclass(frozen=True, slots=True)
class PromotionUnionRow:
    promoted_sku: str
    promoted_title: str | None
    campaign_id: str
    instrument: str | None
    placement: str | None
    purchased_sku: str
    purchased_title: str | None
    revenue_with_vat: Decimal
    orders: int


@dataclass(frozen=True, slots=True)
class PromotionAnalyticsReport:
    statistics: list[PromotionStatisticsRow]
    union: list[PromotionUnionRow]
    period_date_from: str | None = None
    period_date_to: str | None = None


_PERIOD_PATTERN = re.compile(
    r"(\d{2})\.(\d{2})\.(\d{4})\s*-\s*(\d{2})\.(\d{2})\.(\d{4})"
)


class OzonPromotionAnalyticsParser:
    def parse(self, filename: str, content: bytes) -> PromotionAnalyticsReport:
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("Для точной атрибуции нужен XLSX из «Аналитики продвижения».")

        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        try:
            if STATISTICS_SHEET not in workbook.sheetnames:
                raise ValueError("В файле отсутствует лист Statistics.")
            if UNION_SHEET not in workbook.sheetnames:
                raise ValueError("В файле отсутствует лист Union.")

            statistics = self._parse_statistics(workbook[STATISTICS_SHEET])
            union = self._parse_union(workbook[UNION_SHEET])
            period_date_from, period_date_to = _parse_period(workbook[STATISTICS_SHEET])
        finally:
            workbook.close()

        return PromotionAnalyticsReport(
            statistics=statistics,
            union=union,
            period_date_from=period_date_from,
            period_date_to=period_date_to,
        )

    def _parse_statistics(self, sheet: object) -> list[PromotionStatisticsRow]:
        headers, header_row = _find_headers(
            sheet,
            required={"SKU", "ID кампании", "Расход ₽"},
        )
        result: list[PromotionStatisticsRow] = []
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            sku = _identifier(_value(values, headers, "SKU"))
            campaign_id = _identifier(_value(values, headers, "ID кампании"))
            if not sku or not campaign_id:
                continue

            result.append(
                PromotionStatisticsRow(
                    promoted_sku=sku,
                    title=_text(_value(values, headers, "Название товара")),
                    campaign_id=campaign_id,
                    instrument=_text(_value(values, headers, "Инструмент")),
                    placement=_text(_value(values, headers, "Место размещения")),
                    spend_with_vat=_decimal(_value(values, headers, "Расход ₽")),
                    direct_revenue_with_vat=_decimal(
                        _value(values, headers, "Продажи в продвижении ₽")
                    ),
                    direct_orders=_integer(
                        _value(values, headers, "Продано товаров шт")
                    ),
                    model_revenue_with_vat=_decimal(
                        _value(
                            values,
                            headers,
                            "Продажи в продвижении с заказов модели ₽",
                        )
                    ),
                    model_orders=_integer(
                        _value(values, headers, "Продано товаров модели шт")
                    ),
                )
            )
        return result

    def _parse_union(self, sheet: object) -> list[PromotionUnionRow]:
        headers, header_row = _find_headers(
            sheet,
            required={
                "SKU в продвижении",
                "ID кампании",
                "SKU из объединенной карточки",
            },
        )
        result: list[PromotionUnionRow] = []
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            promoted_sku = _identifier(
                _value(values, headers, "SKU в продвижении")
            )
            purchased_sku = _identifier(
                _value(values, headers, "SKU из объединенной карточки")
            )
            campaign_id = _identifier(_value(values, headers, "ID кампании"))
            if not promoted_sku or not purchased_sku or not campaign_id:
                continue

            result.append(
                PromotionUnionRow(
                    promoted_sku=promoted_sku,
                    promoted_title=_text(
                        _value(values, headers, "Название товара в продвижении")
                    ),
                    campaign_id=campaign_id,
                    instrument=_text(_value(values, headers, "Инструмент")),
                    placement=_text(_value(values, headers, "Место размещения")),
                    purchased_sku=purchased_sku,
                    purchased_title=_text(
                        _value(
                            values,
                            headers,
                            "Название товара из объединенной карточки",
                        )
                    ),
                    revenue_with_vat=_decimal(
                        _value(values, headers, "Продажи в продвижении ₽")
                    ),
                    orders=_integer(
                        _value(values, headers, "Продано товаров шт")
                    ),
                )
            )
        return result


def _parse_period(sheet: object) -> tuple[str | None, str | None]:
    for row_values in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
        for value in row_values:
            match = _PERIOD_PATTERN.search(str(value or ""))
            if match:
                d1, m1, y1, d2, m2, y2 = match.groups()
                return f"{y1}-{m1}-{d1}", f"{y2}-{m2}-{d2}"
    return None, None


def _find_headers(
    sheet: object,
    required: set[str],
) -> tuple[dict[str, int], int]:
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=10, values_only=True),
        start=1,
    ):
        headers = {
            _normalize_header(value): index
            for index, value in enumerate(values)
            if _normalize_header(value)
        }
        if all(_normalize_header(name) in headers for name in required):
            return headers, row_number
    raise ValueError(
        f"Не найдены обязательные столбцы: {', '.join(sorted(required))}."
    )


def _value(values: tuple[object, ...], headers: dict[str, int], name: str) -> object:
    index = headers.get(_normalize_header(name))
    if index is None or index >= len(values):
        return None
    return values[index]


def _normalize_header(value: object) -> str:
    text = str(value or "").replace("\n", " ").lower()
    text = re.sub(r"[,;:]+", " ", text)
    return " ".join(text.split()).strip()


def _identifier(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().removesuffix(".0")


def _text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Некорректное числовое значение: {value!r}.") from exc


def _integer(value: object) -> int:
    return int(_decimal(value))
