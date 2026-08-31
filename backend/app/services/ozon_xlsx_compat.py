"""Loads xlsx files exported by Ozon that openpyxl's strict OOXML validation rejects.

Ozon's own export tooling has repeatedly produced spreadsheets with attribute
values outside the enums openpyxl checks against: empty/"none" border styles,
"numFmtID" instead of "numFmtId", capitalized alignment ("Left" instead of
"left"), and hyphenated pane positions ("bottom-right" instead of
"bottomRight"). Excel itself is lenient about all of this; openpyxl is not.
Rather than special-case each report parser, every Ozon xlsx import goes
through this one patch-and-load step.
"""

from __future__ import annotations

import io
import re
import zipfile
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

_BORDER_STYLE_PATTERN = re.compile(r' style="(none|)"')
_ALIGNMENT_HORIZONTAL_PATTERN = re.compile(
    r'horizontal="(Left|Right|Center|General|Justify|Fill|Distributed|CenterContinuous)"'
)
_PANE_PATTERN = re.compile(r"<pane\b[^>]*>(</pane>)?")


def _fix_styles_xml(text: str) -> str:
    text = _BORDER_STYLE_PATTERN.sub("", text)
    text = text.replace("numFmtID=", "numFmtId=")
    text = _ALIGNMENT_HORIZONTAL_PATTERN.sub(
        lambda match: f'horizontal="{match.group(1).lower()}"',
        text,
    )
    return text


def _fix_worksheet_xml(text: str) -> str:
    return _PANE_PATTERN.sub("", text)


def _sanitize_ozon_xlsx(content: bytes) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(content))
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as output:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = _fix_styles_xml(data.decode("utf-8")).encode("utf-8")
            elif item.filename.startswith("xl/worksheets/") and item.filename.endswith(
                ".xml"
            ):
                data = _fix_worksheet_xml(data.decode("utf-8")).encode("utf-8")
            output.writestr(item, data)
    buffer.seek(0)
    return buffer.getvalue()


def load_ozon_workbook(
    file: bytes | BinaryIO,
    *,
    data_only: bool = False,
    read_only: bool = False,
) -> Workbook:
    """Load a workbook, transparently patching known Ozon export quirks on failure."""

    content = file if isinstance(file, bytes) else file.read()
    try:
        return load_workbook(
            io.BytesIO(content),
            data_only=data_only,
            read_only=read_only,
        )
    except ValueError:
        sanitized = _sanitize_ozon_xlsx(content)
        return load_workbook(
            io.BytesIO(sanitized),
            data_only=data_only,
            read_only=read_only,
        )
