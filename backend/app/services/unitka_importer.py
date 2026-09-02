"""Одноразовый импорт `Юнитка Лето СМ.xlsx` (лист `15.06.26`) в базу платформы.

Переносит только INPUT-столбцы (см. `app/domain/unitka.py`) — формульные не читаются
и не хранятся, они считаются заново `unitka_engine.py`. Импорт идемпотентен по
`supplier_article`: повторный запуск не создаст дублей, обновит существующие строки.
"""

from __future__ import annotations

from typing import BinaryIO
from uuid import uuid4

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.unitka import UnitkaAssumptions, UnitkaRow

SHEET_NAME = "15.06.26"

# Буква столбца Excel -> позиционный индекс (1-based), для читаемости при сверке с файлом.
_COL = {letter: openpyxl.utils.column_index_from_string(letter) for letter in [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
    "T", "V", "X", "Z", "AD", "AH", "AQ", "AR", "AS", "AT", "AU",
    "Y", "AB", "AC", "AF", "AG", "AI", "AJ", "AX", "AL",
]}


def _s(ws: Worksheet, row: int, col: str) -> str | None:
    value = ws.cell(row=row, column=_COL[col]).value
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _f(ws: Worksheet, row: int, col: str) -> float | None:
    value = ws.cell(row=row, column=_COL[col]).value
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _article(ws: Worksheet, row: int) -> str:
    value = ws.cell(row=row, column=_COL["B"]).value
    if value is None:
        msg = f"Строка {row}: пустой Артикл прайс (столбец B) — пропускаю."
        raise ValueError(msg)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_assumptions(source: str | BinaryIO) -> UnitkaAssumptions:
    """Ячейки-допущения строки 1 (см. `UnitkaAssumptions` — те же поля 1-в-1)."""
    wb = openpyxl.load_workbook(source, data_only=False)
    ws = wb[SHEET_NAME]
    defaults = UnitkaAssumptions()

    def cell1(col: str, default: float) -> float:
        value = ws.cell(row=1, column=_COL[col]).value
        return float(value) if isinstance(value, (int, float)) else default

    return UnitkaAssumptions(
        sorting_delivery_rate=cell1("Y", defaults.sorting_delivery_rate),
        designer_salary_rate=cell1("AB", defaults.designer_salary_rate),
        fast_payout_rate=cell1("AC", defaults.fast_payout_rate),
        advertising_rate=cell1("AF", defaults.advertising_rate),
        acquiring_rate=cell1("AG", defaults.acquiring_rate),
        other_costs_rate=cell1("AI", defaults.other_costs_rate),
        tax_rate=cell1("AJ", defaults.tax_rate),
        fulfillment_office_rate_per_kg=cell1("AX", defaults.fulfillment_office_rate_per_kg),
    )


def read_rows(source: str | BinaryIO) -> tuple[list[UnitkaRow], list[str]]:
    """Возвращает (строки, предупреждения). Строки без обязательных полей — пропускаются
    с предупреждением, а не молча теряются или заполняются выдумкой."""
    wb = openpyxl.load_workbook(source, data_only=False)
    ws = wb[SHEET_NAME]

    rows: list[UnitkaRow] = []
    warnings: list[str] = []

    for excel_row in range(3, ws.max_row + 1):
        title = _s(ws, excel_row, "G")
        if title is None and _s(ws, excel_row, "B") is None:
            continue  # полностью пустая строка (конец данных) — не предупреждение

        try:
            article = _article(ws, excel_row)
        except ValueError as error:
            warnings.append(str(error))
            continue

        purchase_price = _f(ws, excel_row, "X")
        if purchase_price is None:
            warnings.append(
                f"Строка {excel_row} (арт. {article}): пустая закупочная цена (X) — пропускаю."
            )
            continue

        if title is None:
            warnings.append(f"Строка {excel_row} (арт. {article}): пустое название — пропускаю.")
            continue

        rows.append(
            UnitkaRow(
                id=str(uuid4()),
                row_number=excel_row,
                supplier_article=article,
                fulfillment_scheme=_s(ws, excel_row, "C"),
                ozon_listing=_s(ws, excel_row, "D"),
                stock=_f(ws, excel_row, "E"),
                ozon_sku_id=_s(ws, excel_row, "F"),
                title=title,
                product_type=_s(ws, excel_row, "H"),
                status=_s(ws, excel_row, "I"),
                ozon_visibility=_s(ws, excel_row, "J"),
                in_stock_ozon=_f(ws, excel_row, "K"),
                in_stock_own=_f(ws, excel_row, "L"),
                volume_liters_manual=_f(ws, excel_row, "M"),
                weight_kg=_f(ws, excel_row, "N"),
                dimensions_mm=_s(ws, excel_row, "O"),
                tn_ved=_s(ws, excel_row, "P"),
                honest_mark_required=_s(ws, excel_row, "Q"),
                coinvest_percent=_f(ws, excel_row, "T") or 0.0,
                markup_multiplier=_f(ws, excel_row, "V") or 0.0,
                purchase_price_vat_included=purchase_price,
                ozon_commission_percent=_f(ws, excel_row, "AD") or 0.0,
                integration_fee=_f(ws, excel_row, "AH") or 0.0,
                competitor_price_idd=_f(ws, excel_row, "AQ"),
                competitor_price_ozon=_f(ws, excel_row, "AR"),
                url_idd=_s(ws, excel_row, "AS"),
                url_tdcsm=_s(ws, excel_row, "AT"),
                url_competitor=_s(ws, excel_row, "AU"),
            )
        )

    return rows, warnings
