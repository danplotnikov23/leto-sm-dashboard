import csv
import re
import zipfile
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from app.schemas.ozon_promotions import (
    OzonPromotionAnalysisRow,
    OzonPromotionAnalyzeResponse,
    OzonPromotionKpi,
)


@dataclass
class _InputRow:
    offer_id: str | None = None
    sku: str | None = None
    title: str | None = None
    promotion_name: str | None = None
    qty: int = 0
    revenue: float = 0
    expense: float = 0
    regular_price: float = 0
    promo_price: float = 0
    action_days: int = 0
    discount_percent: float = 0
    notes: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class _ParsedTable:
    rows: list[_InputRow]
    columns: dict[str, str]
    warnings: list[str]


@dataclass(frozen=True)
class _UnitProduct:
    offer_id: str
    title: str | None
    price_with_vat: float


class OzonPromotionsService:
    def analyze(
        self,
        unit_filename: str,
        unit_content: bytes,
        sales_filename: str,
        sales_content: bytes,
        promotions_filename: str,
        promotions_content: bytes,
    ) -> OzonPromotionAnalyzeResponse:
        unit_products, unit_warnings = _parse_unit_prices(unit_filename, unit_content)
        sales = _parse_table(sales_filename, sales_content, kind="sales")
        try:
            promotions = _parse_table(
                promotions_filename,
                promotions_content,
                kind="promotions",
            )
        except ValueError as exc:
            promotions = _ParsedTable(rows=[], columns={}, warnings=[str(exc)])

        sales_by_key: dict[str, _InputRow] = {}
        for row in sales.rows:
            key = _row_key(row)
            if key is None:
                continue
            target = sales_by_key.setdefault(
                key,
                _InputRow(offer_id=row.offer_id, sku=row.sku, title=row.title),
            )
            target.qty += row.qty
            target.revenue += row.revenue
            target.action_days = max(target.action_days, row.action_days)
            target.discount_percent = max(target.discount_percent, row.discount_percent)
            if not target.title and row.title:
                target.title = row.title

        promo_by_group: dict[tuple[str | None, str], _InputRow] = {}
        unmatched = 0
        warnings = [*unit_warnings, *sales.warnings, *promotions.warnings]
        for row in promotions.rows:
            key = _row_key(row)
            if key is None:
                unmatched += 1
                continue
            promo_name = row.promotion_name or "Акция без названия"
            target = promo_by_group.setdefault(
                (key, promo_name),
                _InputRow(
                    offer_id=row.offer_id,
                    sku=row.sku,
                    title=row.title,
                    promotion_name=promo_name,
                ),
            )
            target.qty += row.qty
            target.revenue += row.revenue
            target.expense += row.expense
            target.notes.extend(row.notes)
            if not target.title and row.title:
                target.title = row.title

        rows: list[OzonPromotionAnalysisRow] = []
        used_sales_keys: set[str] = set()
        for (key, promo_name), promo_row in promo_by_group.items():
            sales_row = sales_by_key.get(key)
            used_sales_keys.add(key)
            sales_revenue = sales_row.revenue if sales_row else 0
            sales_qty = sales_row.qty if sales_row else 0
            promo_revenue = promo_row.revenue
            promo_qty = promo_row.qty
            unit_product = unit_products.get(_normalize_offer_key(promo_row.offer_id or ""))
            unit_price = unit_product.price_with_vat if unit_product else None
            avg_promo_price = _safe_div(promo_revenue, promo_qty)
            promo_expense = 0.0
            discount_per_unit = None
            discount_percent = None
            source_notes = sorted(set(promo_row.notes))
            match_status = "matched" if sales_row else "promo_without_sales_row"
            if promo_qty <= 0 or promo_revenue <= 0:
                match_status = "zero_promo_qty"
                source_notes.append("количество или сумма по акции пустые, потери не рассчитаны")
            elif unit_price is None:
                match_status = "unit_not_found"
                source_notes.append("Артикул не найден в юнитке")
            elif avg_promo_price is not None:
                discount_per_unit = unit_price - avg_promo_price
                discount_percent = _percent(discount_per_unit, unit_price)
                if discount_per_unit < 0:
                    promo_expense = 0
                    match_status = "promo_price_higher_than_unit"
                    source_notes.append("Цена акции выше цены из юнитки — проверить цену")
                else:
                    promo_expense = discount_per_unit * promo_qty
            non_promo_revenue = max(sales_revenue - promo_revenue, 0)
            non_promo_qty = max(sales_qty - promo_qty, 0)
            display_offer_id = promo_row.offer_id or (sales_row.offer_id if sales_row else None)
            display_sku = promo_row.sku or (sales_row.sku if sales_row else None)
            display_title = (
                promo_row.title
                or (unit_product.title if unit_product else None)
                or (sales_row.title if sales_row else None)
            )
            rows.append(
                OzonPromotionAnalysisRow(
                    offer_id=display_offer_id,
                    sku=display_sku,
                    title=display_title,
                    promotion_name=promo_name,
                    unit_price_with_vat_rub=round(unit_price, 2) if unit_price is not None else None,
                    sales_total_revenue_rub=round(sales_revenue, 2),
                    sales_total_qty=sales_qty,
                    promo_revenue_rub=round(promo_revenue, 2),
                    promo_qty=promo_qty,
                    promo_expense_rub=round(promo_expense, 2),
                    discount_per_unit_rub=round(discount_per_unit, 2) if discount_per_unit is not None else None,
                    discount_percent=discount_percent,
                    non_promo_revenue_rub=round(non_promo_revenue, 2),
                    non_promo_qty=non_promo_qty,
                    promo_drr_percent=_percent(promo_expense, promo_revenue),
                    promo_share_percent=_percent(promo_revenue, sales_revenue),
                    avg_total_price_rub=_safe_div(sales_revenue, sales_qty),
                    avg_promo_price_rub=_safe_div(promo_revenue, promo_qty),
                    match_status=match_status,
                    source_notes=source_notes,
                )
            )

        for key, sales_row in sales_by_key.items():
            if key in used_sales_keys:
                continue
            has_action_days = sales_row.action_days > 0
            source_notes = []
            if has_action_days:
                source_notes.append(
                    f"в отчёте продаж Ozon указано {sales_row.action_days} дней в акциях; "
                    "точные промо-продажи нужны из расширенной аналитики акций"
                )
            if sales_row.discount_percent:
                source_notes.append(
                    f"средняя скидка от вашей цены: {round(sales_row.discount_percent * 100, 1)}%"
                )
            rows.append(
                OzonPromotionAnalysisRow(
                    offer_id=sales_row.offer_id,
                    sku=sales_row.sku,
                    title=sales_row.title,
                    promotion_name="Был в акциях" if has_action_days else "Без акции",
                    sales_total_revenue_rub=round(sales_row.revenue, 2),
                    sales_total_qty=sales_row.qty,
                    non_promo_revenue_rub=round(sales_row.revenue, 2),
                    non_promo_qty=sales_row.qty,
                    avg_total_price_rub=_safe_div(sales_row.revenue, sales_row.qty),
                    match_status="sales_with_action_days_only" if has_action_days else "sales_without_promo",
                    source_notes=source_notes,
                )
            )

        rows.sort(
            key=lambda row: (
                row.match_status != "matched",
                -(row.promo_expense_rub or 0),
                -(row.promo_revenue_rub or 0),
                row.offer_id or row.sku or "",
            )
        )

        kpi = _build_kpi(rows, unmatched)
        if any(row.match_status == "unit_not_found" for row in rows):
            warnings.append("Часть акционных артикулов не найдена в юнитке, их потери не включены в итог.")
        if not any(row.promo_expense_rub for row in rows):
            warnings.append(
                "Потери на акциях не рассчитались: проверь юнитку, цены и количество продаж по акции."
            )

        return OzonPromotionAnalyzeResponse(
            kpi=kpi,
            rows=rows,
            warnings=_unique(warnings),
            sales_columns=sales.columns,
            promotion_columns=promotions.columns,
        )


def _parse_table(filename: str, content: bytes, kind: str) -> _ParsedTable:
    raw_rows = _read_rows(filename, content, kind=kind)
    header_index = _find_header(raw_rows, kind)
    headers = _combined_headers(raw_rows, header_index)
    columns = _map_columns(headers, kind)
    warnings: list[str] = []
    parsed: list[_InputRow] = []

    for values in raw_rows[header_index + 1 :]:
        row = _build_row(values, columns)
        if row is not None:
            parsed.append(row)

    if not parsed:
        raise ValueError(f"В файле {filename} не найдено строк с товарами")

    if kind == "promotions" and "expense" not in columns:
        estimated = 0
        for row in parsed:
            if row.expense == 0 and row.regular_price > 0 and row.promo_price > 0 and row.qty > 0:
                row.expense = max(row.regular_price - row.promo_price, 0) * row.qty
                row.notes.append("затраты рассчитаны как разница цены до акции и акционной цены")
                estimated += 1
        if estimated:
            warnings.append(f"Затраты на акции рассчитаны по ценам до/после акции: {estimated} строк.")

    readable_columns = {
        key: headers[index]
        for key, index in columns.items()
        if index is not None and index < len(headers)
    }
    return _ParsedTable(rows=parsed, columns=readable_columns, warnings=warnings)


def _parse_unit_prices(filename: str, content: bytes) -> tuple[dict[str, _UnitProduct], list[str]]:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Юнит-экономика должна быть XLSX/XLSM файлом")

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = _select_unit_sheet_for_promotions(workbook.worksheets)
    header_row = 2
    header = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        )
    )
    columns = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
    offer_index = _unit_column(columns, "артикул", fallback=0)
    title_index = _unit_column(columns, "Название", fallback=3)
    price_index = _unit_column(columns, "Цена (со скидкой) с НДС, руб.", fallback=13)
    products: dict[str, _UnitProduct] = {}

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        offer_id = _optional_text(_unit_value(row, offer_index))
        if not offer_id or _normalize_header(offer_id) == "артикул":
            continue

        price = _parse_decimal(_unit_value(row, price_index))
        if price <= 0:
            continue

        products[_normalize_offer_key(offer_id)] = _UnitProduct(
            offer_id=offer_id,
            title=_optional_text(_unit_value(row, title_index)),
            price_with_vat=price,
        )

    warnings: list[str] = []
    if not products:
        warnings.append("В юнитке не найдено артикулов с ценой в колонке N.")
    return products, warnings


def _read_rows(filename: str, content: bytes, kind: str | None = None) -> list[list[str]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            worksheet = _select_ozon_sheet(workbook.worksheets, kind)
            return [
                [_cell_to_string(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
                if any(_cell_to_string(value) for value in row)
            ]
        except ValueError:
            return _read_xlsx_rows_without_styles(content)

    if suffix in {".csv", ".txt"}:
        text = _decode_text(content)
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            return []
        delimiter = max([";", ",", "\t"], key=lambda item: max(line.count(item) for line in lines))
        return [next(csv.reader(StringIO(line), delimiter=delimiter)) for line in lines]

    raise ValueError("Поддерживаются выгрузки Ozon в XLSX, XLSM, CSV или TXT")


def _read_xlsx_rows_without_styles(content: bytes) -> list[list[str]]:
    with zipfile.ZipFile(BytesIO(content)) as archive:
        sheet_name = _first_worksheet_name(archive)
        shared_strings = _read_shared_strings(archive)
        sheet_root = ET.fromstring(archive.read(sheet_name))

    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: list[list[str]] = []
    for row_node in sheet_root.findall(".//x:sheetData/x:row", namespace):
        values: list[str] = []
        for cell_node in row_node.findall("x:c", namespace):
            column_index = _cell_column_index(cell_node.attrib.get("r", ""))
            while len(values) < column_index:
                values.append("")
            values.append(_xlsx_cell_value(cell_node, shared_strings, namespace))
        if any(value.strip() for value in values):
            rows.append(values)

    return rows


def _first_worksheet_name(archive: zipfile.ZipFile) -> str:
    worksheet_names = [
        name
        for name in archive.namelist()
        if name.startswith("xl/worksheets/") and name.endswith(".xml")
    ]
    if not worksheet_names:
        raise ValueError("В XLSX не найден лист с данными")

    return sorted(worksheet_names)[0]


def _select_ozon_sheet(worksheets, kind: str | None):
    return worksheets[0]


def _select_unit_sheet_for_promotions(worksheets):
    for worksheet in worksheets:
        try:
            header = next(
                worksheet.iter_rows(
                    min_row=2,
                    max_row=2,
                    values_only=True,
                )
            )
        except StopIteration:
            continue

        normalized = [_normalize_header(value) for value in header]
        has_offer = any(value == "артикул" for value in normalized)
        has_price = any(
            "цена" in value and "со скид" in value and "ндс" in value
            for value in normalized
        )
        if has_offer and has_price:
            return worksheet

    return worksheets[0]


def _unit_column(columns: dict[str, int], header: str, fallback: int) -> int:
    if header in columns:
        return columns[header]

    normalized_header = _normalize_header(header)
    for name, index in columns.items():
        if _normalize_header(name) == normalized_header:
            return index

    return fallback


def _unit_value(row: tuple[object, ...], index: int) -> object:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("x:si", namespace):
        parts = [node.text or "" for node in item.findall(".//x:t", namespace)]
        strings.append("".join(parts))

    return strings


def _xlsx_cell_value(
    cell_node: ET.Element,
    shared_strings: list[str],
    namespace: dict[str, str],
) -> str:
    cell_type = cell_node.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell_node.findall(".//x:t", namespace)).strip()

    value_node = cell_node.find("x:v", namespace)
    if value_node is None or value_node.text is None:
        return ""

    if cell_type == "s":
        index = _parse_int(value_node.text)
        return shared_strings[index].strip() if 0 <= index < len(shared_strings) else ""

    return value_node.text.strip()


def _cell_column_index(reference: str) -> int:
    letters = "".join(char for char in reference.upper() if "A" <= char <= "Z")
    if not letters:
        return 0

    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _find_header(rows: list[list[str]], kind: str) -> int:
    for index, row in enumerate(rows[:40]):
        normalized = [_normalize_header(value) for value in _combined_headers(rows, index)]
        has_product = any(_looks_like_product_header(header) for header in normalized)
        has_qty = any(_looks_like_qty_header(header) for header in normalized)
        has_money = any(_looks_like_revenue_header(header) for header in normalized)
        has_promo = any("акц" in header or "promotion" in header for header in normalized)
        if has_product and has_qty and has_money and (kind == "sales" or has_promo):
            return index
        if has_product and has_money and kind == "sales":
            return index

    sample = " | ".join(rows[0][:10]) if rows else "пустой файл"
    if kind == "promotions":
        raise ValueError(
            "Не нашёл таблицу расширенной аналитики акций. "
            "Похоже, выгрузка пустая или скачана не та вкладка Ozon. "
            f"Первые колонки: {sample}"
        )
    raise ValueError(f"Не нашёл строку заголовков в выгрузке {kind}. Первые колонки: {sample}")


def _combined_headers(rows: list[list[str]], index: int) -> list[str]:
    current = rows[index] if 0 <= index < len(rows) else []
    previous = rows[index - 1] if index > 0 else []
    width = max(len(current), len(previous))
    headers: list[str] = []
    for column_index in range(width):
        current_value = _value(current, column_index)
        previous_value = _value(previous, column_index)
        if current_value and previous_value and current_value != previous_value:
            headers.append(f"{previous_value} {current_value}")
        else:
            headers.append(current_value or previous_value)
    return headers


def _map_columns(headers: list[str], kind: str) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if _looks_like_offer_header(normalized):
            columns.setdefault("offer_id", index)
        elif normalized == "sku" or "ozon sku" in normalized or "sku ozon" in normalized:
            columns.setdefault("sku", index)
        elif _looks_like_title_header(normalized):
            columns.setdefault("title", index)
        elif _looks_like_qty_header(normalized):
            columns.setdefault("qty", index)
        elif _looks_like_revenue_header(normalized):
            columns.setdefault("revenue", index)
        elif kind == "sales" and _looks_like_action_days_header(normalized):
            columns.setdefault("action_days", index)
        elif kind == "sales" and _looks_like_sales_discount_header(normalized):
            columns.setdefault("discount_percent", index)
        elif kind == "promotions" and _looks_like_expense_header(normalized):
            columns.setdefault("expense", index)
        elif _looks_like_promotion_header(normalized):
            columns.setdefault("promotion_name", index)
        elif _looks_like_regular_price_header(normalized):
            columns.setdefault("regular_price", index)
        elif _looks_like_promo_price_header(normalized):
            columns.setdefault("promo_price", index)

    if kind == "promotions":
        promo_qty_index = _find_column(headers, _looks_like_promo_qty_header)
        promo_revenue_index = _find_column(headers, _looks_like_promo_revenue_header)
        if promo_qty_index is not None:
            columns["qty"] = promo_qty_index
        if promo_revenue_index is not None:
            columns["revenue"] = promo_revenue_index

    required = ["qty", "revenue"] if kind == "sales" else ["qty"]
    missing = [name for name in required if name not in columns]
    if missing:
        raise ValueError(
            "Не хватает колонок в выгрузке Ozon: "
            f"{', '.join(missing)}. Заголовки: {', '.join(headers[:20])}"
        )

    return columns


def _build_row(values: list[str], columns: dict[str, int]) -> _InputRow | None:
    offer_id = _extract_offer_id(_value(values, columns.get("offer_id"))) or _extract_offer_id(
        _value(values, columns.get("title"))
    )
    sku = _optional_text(_value(values, columns.get("sku")))
    title = _optional_text(_value(values, columns.get("title")))
    qty = _parse_int(_value(values, columns.get("qty")))
    revenue = _parse_decimal(_value(values, columns.get("revenue")))
    expense = _parse_decimal(_value(values, columns.get("expense")))
    regular_price = _parse_decimal(_value(values, columns.get("regular_price")))
    promo_price = _parse_decimal(_value(values, columns.get("promo_price")))
    action_days = _parse_action_days(_value(values, columns.get("action_days")))
    discount_percent = _parse_decimal(_value(values, columns.get("discount_percent")))
    promotion_name = _optional_text(_value(values, columns.get("promotion_name")))

    if not offer_id and not sku and not title:
        return None
    if title and _normalize_header(title).startswith("итого"):
        return None
    if qty == 0 and revenue == 0 and expense == 0:
        return None

    return _InputRow(
        offer_id=offer_id,
        sku=sku,
        title=title,
        promotion_name=promotion_name,
        qty=qty,
        revenue=revenue,
        expense=expense,
        regular_price=regular_price,
        promo_price=promo_price,
        action_days=action_days,
        discount_percent=discount_percent,
    )


def _build_kpi(rows: list[OzonPromotionAnalysisRow], unmatched: int) -> OzonPromotionKpi:
    sales_by_key: dict[str, tuple[float, int]] = {}
    promo_revenue = 0.0
    promo_qty = 0
    promo_expense = 0.0
    promo_articles: set[str] = set()
    missing_unit_articles: set[str] = set()
    for row in rows:
        key = row.offer_id or row.sku or row.title or ""
        if key and key not in sales_by_key:
            sales_by_key[key] = (row.sales_total_revenue_rub, row.sales_total_qty)
        promo_revenue += row.promo_revenue_rub
        promo_qty += row.promo_qty
        promo_expense += row.promo_expense_rub
        if row.promo_qty > 0 and row.offer_id:
            promo_articles.add(_normalize_offer_key(row.offer_id))
        if row.match_status == "unit_not_found" and row.offer_id:
            missing_unit_articles.add(_normalize_offer_key(row.offer_id))

    sales_total = sum(value[0] for value in sales_by_key.values())
    sales_qty = sum(value[1] for value in sales_by_key.values())
    non_promo_revenue = max(sales_total - promo_revenue, 0)
    non_promo_qty = max(sales_qty - promo_qty, 0)
    return OzonPromotionKpi(
        sales_total_revenue_rub=round(sales_total, 2),
        sales_total_qty=sales_qty,
        promo_revenue_rub=round(promo_revenue, 2),
        promo_qty=promo_qty,
        promo_expense_rub=round(promo_expense, 2),
        non_promo_revenue_rub=round(non_promo_revenue, 2),
        non_promo_qty=non_promo_qty,
        promo_drr_percent=_percent(promo_expense, promo_revenue),
        promo_share_percent=_percent(promo_revenue, sales_total),
        discount_share_percent=_percent(promo_expense, sales_total),
        average_promo_discount_percent=_percent(promo_expense, promo_revenue),
        promo_articles_count=len(promo_articles),
        missing_unit_articles_count=len(missing_unit_articles),
        rows_count=len(rows),
        unmatched_promo_rows=unmatched,
    )


def _row_key(row: _InputRow) -> str | None:
    if row.offer_id:
        return f"offer:{_normalize_offer_key(row.offer_id)}"
    if row.sku:
        return f"sku:{row.sku}"
    if row.title:
        return f"title:{_normalize_header(row.title)}"
    return None


def _normalize_offer_key(value: str) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def _looks_like_product_header(value: str) -> bool:
    return _looks_like_offer_header(value) or value == "sku" or _looks_like_title_header(value)


def _looks_like_offer_header(value: str) -> bool:
    return (
        "артикул" in value
        or "offer id" in value
        or "offer_id" in value
        or "seller article" in value
    )


def _looks_like_title_header(value: str) -> bool:
    return value in {"товар", "товары", "название"} or "название товара" in value or "наименование" in value


def _looks_like_qty_header(value: str) -> bool:
    return (
        ("заказ" in value and ("товар" in value or "шт" in value))
        or "кол во" in value
        or "количество" in value
        or value in {"шт", "qty", "quantity"}
    )


def _looks_like_promo_qty_header(value: str) -> bool:
    return "заказ" in value and "товар" in value and "по акц" in value


def _looks_like_revenue_header(value: str) -> bool:
    if (
        "abc" in value
        or "средн" in value
        or "цена" in value
        or "доля" in value
        or "динамика" in value
        or "%" in value
    ):
        return False
    return (
        ("заказ" in value and "сумм" in value)
        or "выруч" in value
        or "оборот" in value
        or "продаж" in value and ("руб" in value or "сум" in value)
        or value in {"revenue", "sales"}
    )


def _looks_like_promo_revenue_header(value: str) -> bool:
    return "заказ" in value and "сумм" in value and "по акц" in value


def _looks_like_expense_header(value: str) -> bool:
    if "процент" in value or "%" in value:
        return False
    return (
        "затрат" in value
        or "расход" in value
        or "скидк" in value
        or "компенсац" in value
        or "discount amount" in value
    )


def _looks_like_sales_discount_header(value: str) -> bool:
    return "скидк" in value and ("вашей цен" in value or "ваша цен" in value)


def _looks_like_action_days_header(value: str) -> bool:
    return "дней" in value and "акц" in value


def _looks_like_promotion_header(value: str) -> bool:
    return "название акции" in value or value in {"акция", "акции"} or "promotion" in value


def _looks_like_regular_price_header(value: str) -> bool:
    return "цена до" in value or "старая цена" in value or "regular price" in value


def _looks_like_promo_price_header(value: str) -> bool:
    return "акционная цена" in value or "цена акции" in value or "action price" in value


def _normalize_header(value: object) -> str:
    normalized = str(value or "").replace("ё", "е").replace("\n", " ").lower()
    normalized = re.sub(r"[_/.,;:()]+", " ", normalized)
    return " ".join(normalized.split())


def _find_column(headers: list[str], predicate) -> int | None:
    for index, header in enumerate(headers):
        if predicate(_normalize_header(header)):
            return index
    return None


def _value(values: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(values):
        return ""
    return _cell_to_string(values[index])


def _cell_to_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8-sig", errors="replace")


def _optional_text(value: str) -> str | None:
    text = str(value or "").strip()
    if text in {"-", "–", "—"}:
        return None
    return text or None


def _extract_offer_id(value: str) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"(?:арт\.?\s*)?([А-ЯA-Z]{1,4}-\d{6,})", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return text if "-" in text and any(char.isdigit() for char in text) else None


def _parse_decimal(value: object) -> float:
    normalized = (
        str(value or "")
        .replace("\u00a0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("%", "")
        .replace(",", ".")
        .strip()
    )
    if normalized in {"", "-"}:
        return 0
    try:
        return float(normalized)
    except ValueError:
        return 0


def _parse_int(value: object) -> int:
    return int(round(_parse_decimal(value)))


def _parse_action_days(value: object) -> int:
    text = str(value or "").strip()
    match = re.search(r"\d+", text)
    if match:
        return int(match.group(0))
    return _parse_int(text)


def _safe_div(numerator: float, denominator: float | int) -> float | None:
    if not denominator:
        return None
    return round(numerator / denominator, 2)


def _percent(numerator: float, denominator: float) -> float | None:
    if not denominator:
        return None
    return round(numerator / denominator * 100, 2)


def _unique(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result
