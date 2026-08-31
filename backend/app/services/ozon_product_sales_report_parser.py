"""Parses Ozon's "Аналитика по товарам" export (analytics_report / analytics_report_daily).

Both the daily and period variants share the same column set; the daily one
additionally carries a per-row "День" (day) column, which is what lets us
store data at day granularity and later answer for any sub-period without a
re-upload. When that column is absent (the period variant), every row is
tagged with the report's own declared period instead.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl.worksheet.worksheet import Worksheet

from app.services.ozon_xlsx_compat import load_ozon_workbook

DATA_SHEET_NAME = "По товарам"
HEADER_SCAN_ROWS = 15
_PERIOD_PATTERN = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})\s*[–-]\s*(\d{2})\.(\d{2})\.(\d{4})")

_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "sku": ("sku",),
    "offer_id": ("артикул",),
    "title": ("товары",),
    "day": ("день",),
    "ordered_units": ("заказано товаров",),
    "redeemed_units": ("выкуплено товаров",),
    "cancelled_units": ("отменено товаров (на дату заказа)",),
    "avg_price": ("средняя цена",),
    "discount_fraction": ("скидка от вашей цены",),
}


@dataclass(frozen=True, slots=True)
class ProductSalesRow:
    sku: str
    offer_id: str | None
    title: str | None
    date_from: str
    date_to: str
    ordered_units: int
    redeemed_units: int
    cancelled_units: int
    avg_price: float
    discount_fraction: float


@dataclass(frozen=True, slots=True)
class ProductSalesReport:
    rows: tuple[ProductSalesRow, ...]
    period_date_from: str | None
    period_date_to: str | None
    has_daily_breakdown: bool


class ProductSalesReportError(Exception):
    """Raised when the uploaded workbook cannot be parsed as an Ozon sales report."""


class OzonProductSalesReportParser:
    def parse(self, filename: str, content: bytes) -> ProductSalesReport:
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise ProductSalesReportError(
                "Нужен XLSX-отчёт «Аналитика по товарам» из Ozon."
            )

        workbook = load_ozon_workbook(content, read_only=False, data_only=True)
        try:
            if DATA_SHEET_NAME not in workbook.sheetnames:
                raise ProductSalesReportError(
                    f"В файле отсутствует лист «{DATA_SHEET_NAME}»."
                )
            sheet = workbook[DATA_SHEET_NAME]
            period_date_from, period_date_to = _parse_declared_period(sheet)
            columns, header_row = _find_columns(sheet)
            has_daily_breakdown = "day" in columns
            rows = _parse_rows(
                sheet,
                columns,
                header_row=header_row,
                has_daily_breakdown=has_daily_breakdown,
                fallback_date_from=period_date_from,
                fallback_date_to=period_date_to,
            )
        finally:
            workbook.close()

        if not rows:
            raise ProductSalesReportError("В отчёте не найдено ни одной товарной строки.")

        return ProductSalesReport(
            rows=tuple(rows),
            period_date_from=period_date_from,
            period_date_to=period_date_to,
            has_daily_breakdown=has_daily_breakdown,
        )


def _parse_declared_period(sheet: Worksheet) -> tuple[str | None, str | None]:
    for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
        for value in row:
            match = _PERIOD_PATTERN.search(str(value or ""))
            if match:
                d1, m1, y1, d2, m2, y2 = match.groups()
                return f"{y1}-{m1}-{d1}", f"{y2}-{m2}-{d2}"
    return None, None


def _normalize_header(value: object) -> str:
    text = str(value or "").replace("\n", " ").strip().lower()
    return " ".join(text.split())


def _find_columns(sheet: Worksheet) -> tuple[dict[str, int], int]:
    for header_row in range(1, HEADER_SCAN_ROWS):
        row_values = [
            _normalize_header(cell.value) for cell in sheet[header_row]
        ]
        if "sku" not in row_values:
            continue

        next_row_values = [
            _normalize_header(cell.value) for cell in sheet[header_row + 1]
        ]
        merged = [
            next_value or row_values[index] if index < len(row_values) else next_value
            for index, next_value in enumerate(next_row_values)
        ]

        columns: dict[str, int] = {}
        for key, aliases in _COLUMN_ALIASES.items():
            for column_index, header in enumerate(merged, start=1):
                if header in aliases:
                    columns[key] = column_index
                    break

        required = {"sku", "redeemed_units", "avg_price"}
        if required.issubset(columns):
            return columns, header_row + 1

    raise ProductSalesReportError(
        "Не найдены обязательные столбцы SKU / Выкуплено товаров / Средняя цена."
    )


def _parse_rows(
    sheet: Worksheet,
    columns: dict[str, int],
    *,
    header_row: int,
    has_daily_breakdown: bool,
    fallback_date_from: str | None,
    fallback_date_to: str | None,
) -> list[ProductSalesRow]:
    rows: list[ProductSalesRow] = []
    for row_cells in sheet.iter_rows(min_row=header_row + 1):
        sku = _to_sku(_cell(row_cells, columns.get("sku")))
        if sku is None:
            continue

        row_date_from = fallback_date_from
        row_date_to = fallback_date_to
        if has_daily_breakdown:
            day_value = _cell(row_cells, columns.get("day"))
            day_iso = _to_iso_date(day_value)
            if day_iso is None:
                continue
            row_date_from = day_iso
            row_date_to = day_iso

        if row_date_from is None or row_date_to is None:
            continue

        rows.append(
            ProductSalesRow(
                sku=sku,
                offer_id=_to_text(_cell(row_cells, columns.get("offer_id"))),
                title=_to_text(_cell(row_cells, columns.get("title"))),
                date_from=row_date_from,
                date_to=row_date_to,
                ordered_units=_to_int(_cell(row_cells, columns.get("ordered_units"))),
                redeemed_units=_to_int(_cell(row_cells, columns.get("redeemed_units"))),
                cancelled_units=_to_int(_cell(row_cells, columns.get("cancelled_units"))),
                avg_price=_to_float(_cell(row_cells, columns.get("avg_price"))),
                discount_fraction=_to_float(
                    _cell(row_cells, columns.get("discount_fraction"))
                ),
            )
        )
    return rows


def _cell(row_cells: tuple, column_index: int | None) -> object:
    if column_index is None or column_index > len(row_cells):
        return None
    return row_cells[column_index - 1].value


def _to_sku(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if not float(value).is_integer():
            return None
        return str(int(value))
    text = str(value).strip()
    return text if text.isdigit() else None


def _to_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_int(value: object) -> int:
    if value is None or isinstance(value, bool):
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _to_float(value: object) -> float:
    if value is None or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(" ", "").replace(" ", "").replace(",", ".").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_iso_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date().isoformat()
    except ValueError:
        return None
