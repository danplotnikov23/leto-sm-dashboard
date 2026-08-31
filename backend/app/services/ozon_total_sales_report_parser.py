import csv
import re
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from app.schemas.ozon import OzonPromotionInfo


@dataclass(frozen=True)
class OzonTotalSalesReportRow:
    offer_id: str | None
    sku: str | None
    title: str | None
    ordered_amount_with_vat: float
    orders: int
    seller_product_id: str | None = None
    promotions: list[OzonPromotionInfo] | None = None


@dataclass(frozen=True)
class ParsedOzonTotalSalesReport:
    rows: list[OzonTotalSalesReportRow]


class OzonTotalSalesReportParser:
    def parse(self, filename: str, content: bytes) -> ParsedOzonTotalSalesReport:
        suffix = Path(filename).suffix.lower()
        if suffix in {".csv", ".txt"}:
            raw_rows = _parse_csv(content)
        elif suffix in {".xlsx", ".xlsm"}:
            raw_rows = _parse_xlsx(content)
        else:
            raise ValueError("Unsupported Ozon sales report format. Use CSV or XLSX.")

        rows = _build_sales_rows(raw_rows)
        if not rows:
            raise ValueError("Ozon sales report has no readable product rows")

        return ParsedOzonTotalSalesReport(rows=rows)


def _parse_csv(content: bytes) -> list[list[str]]:
    text = _decode_text(content)
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []

    delimiter = _detect_delimiter(lines)
    return [next(csv.reader([line], delimiter=delimiter)) for line in lines]


def _parse_xlsx(content: bytes) -> list[list[str]]:
    with NamedTemporaryFile(suffix=".xlsx") as file:
        file.write(content)
        file.flush()
        workbook = load_workbook(file.name, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        return [
            [_cell_to_string(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]


def _build_sales_rows(raw_rows: list[list[str]]) -> list[OzonTotalSalesReportRow]:
    header_index = _find_sales_header_index(raw_rows)
    headers = raw_rows[header_index]
    columns = _build_column_map(headers)
    rows: list[OzonTotalSalesReportRow] = []

    for raw_values in raw_rows[header_index + 1 :]:
        row = _normalize_sales_row(raw_values, columns)
        if row is not None:
            rows.append(row)

    return rows


def _find_sales_header_index(rows: list[list[str]]) -> int:
    for index, row in enumerate(rows):
        normalized_headers = [_normalize_header(value) for value in row]
        has_amount = any(
            "заказано" in header and "сумм" in header
            for header in normalized_headers
        )
        has_orders = any(
            "заказано" in header
            and "сумм" not in header
            and ("товар" in header or "шт" in header)
            for header in normalized_headers
        )
        has_product = any(
            "артикул" in header or "название" in header or header == "sku"
            for header in normalized_headers
        )
        if has_amount and has_orders and has_product:
            return index

    raise ValueError("Ozon total sales report header row was not found")


def _build_column_map(headers: list[str]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if normalized == "sku" or "ozon sku" in normalized:
            columns["sku"] = index
        elif "артикул" in normalized:
            columns["offer_id"] = index
        elif "название" in normalized or normalized == "товар":
            columns.setdefault("title", index)
        elif "заказано" in normalized and "сумм" in normalized:
            columns["ordered_amount"] = index
        elif (
            "заказано" in normalized
            and "сумм" not in normalized
            and ("товар" in normalized or "шт" in normalized)
        ):
            columns["orders"] = index

    if "ordered_amount" not in columns or "orders" not in columns:
        raise ValueError("Ozon total sales report misses amount or orders columns")

    return columns


def _normalize_sales_row(
    values: list[str],
    columns: dict[str, int],
) -> OzonTotalSalesReportRow | None:
    ordered_amount = _parse_decimal(_value(values, columns["ordered_amount"]))
    orders = _parse_int(_value(values, columns["orders"]))
    if ordered_amount == 0 and orders == 0:
        return None

    raw_offer = _value(values, columns.get("offer_id"))
    raw_title = _value(values, columns.get("title"))
    offer_id = _extract_offer_id(raw_offer) or _extract_offer_id(raw_title)
    sku = _optional_text(_value(values, columns.get("sku")))
    if offer_id is None and sku is None:
        return None

    title = _clean_title(raw_title)

    return OzonTotalSalesReportRow(
        offer_id=offer_id,
        sku=sku,
        title=title,
        ordered_amount_with_vat=ordered_amount,
        orders=orders,
    )


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue

    return content.decode("utf-8-sig", errors="replace")


def _detect_delimiter(lines: list[str]) -> str:
    candidates = [";", ",", "\t"]
    return max(candidates, key=lambda delimiter: max(line.count(delimiter) for line in lines))


def _normalize_header(value: str) -> str:
    normalized = (
        str(value or "")
        .replace("\n", " ")
        .replace(",", " ")
        .strip()
        .lower()
    )
    return " ".join(normalized.split())


def _value(values: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(values):
        return ""

    return _cell_to_string(values[index])


def _cell_to_string(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _optional_text(value: str) -> str | None:
    text = value.strip()
    return text or None


def _clean_title(value: str) -> str | None:
    text = value.strip()
    if not text:
        return None

    return re.sub(r"\s+", " ", text)


def _extract_offer_id(value: str) -> str | None:
    text = value.strip()
    if not text:
        return None

    match = re.search(r"(?:арт\.?\s*)?([А-ЯA-Z]{1,4}-\d{6,})", text, re.IGNORECASE)
    if match is not None:
        return match.group(1).upper()

    return text if "-" in text and any(char.isdigit() for char in text) else None


def _parse_decimal(value: object) -> float:
    if value is None:
        return 0

    normalized = (
        str(value)
        .replace("\u00a0", "")
        .replace(" ", "")
        .replace("₽", "")
        .replace("%", "")
        .replace(",", ".")
        .strip()
    )
    try:
        return float(normalized)
    except ValueError:
        return 0


def _parse_int(value: object) -> int:
    return int(round(_parse_decimal(value)))
