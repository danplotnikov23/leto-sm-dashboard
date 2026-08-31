from __future__ import annotations

from openpyxl.utils import get_column_letter

from app.schemas.price_update import (
    PriceUpdateCategoriesResponse,
    PriceUpdateCategory,
    PriceUpdateItem,
    PriceUpdateProductMatch,
    PriceUpdateSearchResponse,
)
from app.services.ozon_period_validation import get_moscow_today
from app.services.ozon_xlsx_compat import load_ozon_workbook
from app.services.unit_economy_index_service import UnitEconomyIndexService
from app.services.xlsx_cell_patch import (
    XlsxCellPatchError,
    patch_cells,
    resolve_sheet_xml_path,
)

DEFAULT_VAT_MULTIPLIER = 1.22

SHEET_NAME = "Товары и цены"
DATA_START_ROW = 5

COL_OFFER_ID = 1
COL_SKU = 2
COL_NAME = 3
COL_PRICE_BEFORE_DISCOUNT = 15
COL_PRICE_WITH_DISCOUNT = 16
COL_VAT_PERCENT = 22
COL_MIN_PRICE = 23
COL_TEMPLATE_COST = 27
COL_NEW_PRICE_BEFORE_DISCOUNT = 54
COL_NEW_PRICE_WITH_DISCOUNT = 55
COL_NEW_COST = 58
COL_NEW_MIN_PRICE = 60

NEW_PRICE_BEFORE_DISCOUNT_COLUMN_LETTER = get_column_letter(COL_NEW_PRICE_BEFORE_DISCOUNT)
NEW_PRICE_COLUMN_LETTER = get_column_letter(COL_NEW_PRICE_WITH_DISCOUNT)
NEW_COST_COLUMN_LETTER = get_column_letter(COL_NEW_COST)
NEW_MIN_PRICE_COLUMN_LETTER = get_column_letter(COL_NEW_MIN_PRICE)


class PriceUpdateTemplateError(ValueError):
    pass


class PriceUpdateService:
    def __init__(self, unit_economy_index_service: UnitEconomyIndexService) -> None:
        self._unit_economy_index_service = unit_economy_index_service

    def list_categories(self) -> PriceUpdateCategoriesResponse:
        counts: dict[str, int] = {}
        for product in self._unit_economy_index_service.list_products():
            if not product.category:
                continue
            counts[product.category] = counts.get(product.category, 0) + 1

        categories = [
            PriceUpdateCategory(category=category, product_count=count)
            for category, count in sorted(counts.items(), key=lambda item: -item[1])
        ]
        return PriceUpdateCategoriesResponse(categories=categories)

    def search_products(
        self,
        template_bytes: bytes,
        query: str,
        category: str | None = None,
    ) -> PriceUpdateSearchResponse:
        worksheet = self._load_products_sheet(template_bytes, read_only=True, data_only=True)

        today = get_moscow_today()
        normalized_query = query.strip().lower()
        matches: list[PriceUpdateProductMatch] = []
        total_rows = 0

        for row in worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
            offer_id = row[COL_OFFER_ID - 1]
            if offer_id is None or str(offer_id).strip() == "":
                continue
            total_rows += 1

            offer_id_text = str(offer_id).strip()
            name = row[COL_NAME - 1]
            if normalized_query and (
                normalized_query not in str(name or "").lower()
                and normalized_query not in offer_id_text.lower()
            ):
                continue

            unit_product = self._unit_economy_index_service.find_by_offer_id_for_date(
                offer_id_text, today
            )

            if category and (
                unit_product is None or unit_product.category != category
            ):
                continue

            # Base the markup on "Закупочная цена без НДС" (raw purchase
            # price), matching the "Торговая наценка" reference column
            # already in the unit-economy workbook exactly - confirmed
            # against real rows (e.g. ЦБ-00203546: workbook says 359.97%,
            # (price_without_vat - purchase_price) / purchase_price gives
            # the same number). "Себес (затраты)" (full internal cost incl.
            # logistics/warehouse/etc) is a different, much higher base and
            # produces a much lower - and unfamiliar - percentage, which is
            # why this was switched from that field.
            unit_economy_cost = (
                unit_product.cost_without_vat if unit_product is not None else None
            )

            current_price_with_discount = _optional_float(row[COL_PRICE_WITH_DISCOUNT - 1])
            vat_percent = _optional_float(row[COL_VAT_PERCENT - 1])
            vat_multiplier = (
                1 + vat_percent / 100 if vat_percent is not None else DEFAULT_VAT_MULTIPLIER
            )
            current_markup_percent = None
            if (
                unit_economy_cost is not None
                and unit_economy_cost > 0
                and current_price_with_discount is not None
            ):
                price_without_vat = current_price_with_discount / vat_multiplier
                current_markup_percent = (
                    (price_without_vat - unit_economy_cost) / unit_economy_cost * 100
                )

            matches.append(
                PriceUpdateProductMatch(
                    offer_id=offer_id_text,
                    sku=_optional_text(row[COL_SKU - 1]),
                    name=str(name or offer_id_text),
                    category=unit_product.category if unit_product is not None else None,
                    current_price_before_discount=_optional_float(
                        row[COL_PRICE_BEFORE_DISCOUNT - 1]
                    ),
                    current_price_with_discount=current_price_with_discount,
                    current_min_price=_optional_float(row[COL_MIN_PRICE - 1]),
                    current_vat_percent=vat_percent,
                    current_markup_percent=current_markup_percent,
                    template_cost=_optional_float(row[COL_TEMPLATE_COST - 1]),
                    unit_economy_cost=unit_economy_cost,
                    unit_economy_price_before_discount=(
                        unit_product.price_before_discount_with_vat
                        if unit_product is not None
                        else None
                    ),
                    unit_economy_expense_cost=(
                        unit_product.expense_cost if unit_product is not None else None
                    ),
                    unit_economy_matched=unit_economy_cost is not None,
                )
            )

        return PriceUpdateSearchResponse(
            query=query,
            category=category,
            total_rows_in_template=total_rows,
            matches=matches,
        )

    def apply_new_prices(
        self,
        template_bytes: bytes,
        updates: list[PriceUpdateItem],
    ) -> bytes:
        # This deliberately does NOT load the workbook with openpyxl and
        # re-save it - a full round-trip of a real Ozon export (2500+ rows,
        # thousands of per-cell data validations, inline strings) was found
        # to silently rewrite untouched cells into schema-invalid XML (e.g.
        # a blank cell serialized as `t="inlineStr"` with no `<is>` child),
        # which Ozon's own upload validator then rejected with a generic
        # "Ошибка загрузки файла" - even though Excel opened the file fine.
        # Instead, only the exact target cells are patched directly in the
        # worksheet XML inside the zip; every other byte of the original
        # Ozon export - which Ozon itself produced and accepts - is passed
        # through untouched.
        if not updates:
            raise PriceUpdateTemplateError("Список изменений пуст")

        worksheet = self._load_products_sheet(template_bytes, read_only=True, data_only=True)
        today = get_moscow_today()
        row_by_offer_id: dict[str, int] = {}
        for row in worksheet.iter_rows(
            min_row=DATA_START_ROW, min_col=COL_OFFER_ID, max_col=COL_OFFER_ID
        ):
            cell = row[0]
            if cell.value is None:
                continue
            row_by_offer_id[str(cell.value).strip()] = cell.row

        missing_offer_ids = [
            update.offer_id for update in updates if update.offer_id not in row_by_offer_id
        ]
        if missing_offer_ids:
            raise PriceUpdateTemplateError(
                "Артикулы не найдены в шаблоне: " + ", ".join(missing_offer_ids[:20])
            )

        # Ozon rejects the row ("Некорректная цена") unless "Новая
        # минимальная цена" is also set once the price changes - its own
        # hint says the value "не может быть больше текущей цены", so it's
        # derived from the price we're setting (without VAT) rather than
        # read from unit-economy's own O column, which guarantees it never
        # exceeds the new price we just wrote, regardless of markup chosen.
        # "Новая себестоимость" is filled from "Себес (затраты)" so Ozon's
        # own margin estimate isn't left blank. "Новая цена до скидки" is
        # filled straight from unit-economy's own "Цена до скидки с НДС" -
        # unlike the other three fields it isn't derived from the price we
        # just computed, since it reflects the strikethrough/"was" price a
        # shopper sees, not anything tied to our markup override.
        cell_updates: list[tuple[int, str, float]] = []
        for update in updates:
            row_number = row_by_offer_id[update.offer_id]
            new_price = round(update.new_price_with_discount, 2)
            cell_updates.append((row_number, NEW_PRICE_COLUMN_LETTER, new_price))
            cell_updates.append(
                (row_number, NEW_MIN_PRICE_COLUMN_LETTER, round(new_price / DEFAULT_VAT_MULTIPLIER, 2))
            )
            unit_product = self._unit_economy_index_service.find_by_offer_id_for_date(
                update.offer_id, today
            )
            if unit_product is not None and unit_product.expense_cost is not None:
                cell_updates.append(
                    (row_number, NEW_COST_COLUMN_LETTER, round(unit_product.expense_cost, 2))
                )
            if (
                unit_product is not None
                and unit_product.price_before_discount_with_vat is not None
            ):
                cell_updates.append(
                    (
                        row_number,
                        NEW_PRICE_BEFORE_DISCOUNT_COLUMN_LETTER,
                        round(unit_product.price_before_discount_with_vat, 2),
                    )
                )

        try:
            sheet_xml_path = resolve_sheet_xml_path(template_bytes, SHEET_NAME)
            return patch_cells(template_bytes, sheet_xml_path, cell_updates)
        except XlsxCellPatchError as exc:
            raise PriceUpdateTemplateError(str(exc)) from exc

    def _load_products_sheet(
        self,
        template_bytes: bytes,
        *,
        read_only: bool,
        data_only: bool,
    ):
        workbook = load_ozon_workbook(
            template_bytes,
            data_only=data_only,
            read_only=read_only,
        )
        if SHEET_NAME not in workbook.sheetnames:
            raise PriceUpdateTemplateError(
                f"В файле нет листа «{SHEET_NAME}» - это не шаблон Ozon для цен"
            )
        return workbook[SHEET_NAME]


def _optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _optional_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
