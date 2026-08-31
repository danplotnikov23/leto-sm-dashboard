import asyncio
import csv
import hashlib
import logging
from datetime import UTC, date, datetime, timedelta
from io import BytesIO, StringIO
from zoneinfo import ZoneInfo

from fastapi import UploadFile

from app.core.config import Settings
from app.schemas.ozon import (
    OzonCampaign,
    OzonCampaignEfficiencyResponse,
    OzonCampaignPerformanceMetrics,
    OzonCampaignProduct,
    OzonPromotionInfo,
    OzonSalesAnalyticsResponse,
    OzonSalesAnalyticsRow,
    OzonSalesForecastRequest,
    OzonSalesForecastResponse,
    OzonSalesForecastRow,
    OzonSkuEfficiencyRow,
    OzonSkuEfficiencyResponse,
    OzonSkuEfficiencySegment,
    OzonSkuDailyProfitResponse,
    OzonSkuDailyProfitRow,
    OzonSkuDailyProfitTotal,
    OzonStatisticsReportStatus,
    OzonStoredReportSummary,
)
from app.schemas.ozon_ad_attribution import (
    OzonAdAttributionResponse,
    OzonPromotionAnalyticsImportSummary,
)
from app.services.ozon_ad_economics_calculator import (
    OzonAdEconomicsCalculator,
    PROFIT_TAX_RATE,
    VAT_MULTIPLIER,
    can_include_model_attribution_in_sku,
    parse_decimal,
    parse_int,
    parse_optional_decimal,
)
from app.services.ozon_ad_report_file_parser import (
    OzonAdReportFileParser,
    canonicalize_ozon_ad_report_header,
    is_ozon_ad_report_header,
)
from app.services.ozon_ad_report_repository import OzonAdReportRepository
from app.services.ozon_ad_attribution_calculator import (
    OzonAdAttributionCalculator,
    ResolvedAttributionProduct,
)
from app.services.ozon_errors import OzonApiError, OzonReportNotReadyError
from app.services.ozon_performance_client import OzonPerformanceClient
from app.services.ozon_seller_client import OzonSellerClient
from app.services.ozon_total_sales_report_parser import OzonTotalSalesReportParser
from app.services.ozon_total_sales_report_parser import OzonTotalSalesReportRow
from app.services.ozon_promotion_analytics_parser import (
    OzonPromotionAnalyticsParser,
    PromotionAnalyticsReport,
)
from app.schemas.ozon_daily_profit import OzonDailyProfitSnapshot
from app.schemas.ozon_order_lookup import (
    OzonOrderItem,
    OzonOrderLookupResponse,
    OzonTodayOrdersSummary,
)
from app.schemas.ozon_product_sales import (
    OzonProductSalesImportSummary,
    OzonProductSalesResponse,
    OzonProductSalesRow,
    OzonProductSalesTotal,
)
from app.services.ozon_product_sales_report_parser import (
    OzonProductSalesReportParser,
    ProductSalesReportError,
    ProductSalesRow,
)
from app.services.financial_calculator import net_after_profit_tax, without_vat
from decimal import Decimal
from app.schemas.unit_economy_index import UnitEconomyProduct
from app.services.unit_economy_index_service import (
    UnitEconomyIndexService,
    UnitEconomyPeriodSegment,
    UnitEconomyVersionSelection,
    UnitEconomyWorkbookVersion,
)


logger = logging.getLogger(__name__)
MOSCOW_TZ = ZoneInfo("Europe/Moscow")


class OzonAdsService:
    # Class-level (not per-instance) so it's shared across every request-scoped
    # OzonAdsService AND the scheduler's own long-lived instance. Ozon allows
    # only one statistics report in flight per account - without this, the
    # hourly scheduler tick and a manual "Обновить сейчас" click can race to
    # create a report at the same time, both fall into Ozon's "active report"
    # retry-wait loop, and each pointlessly waits on the other (confirmed
    # live: a manual refresh sat for 60s+ just trying to create the report).
    _promotion_report_lock = asyncio.Lock()

    def __init__(
        self,
        settings: Settings,
        unit_economy_index_service: UnitEconomyIndexService | None = None,
        ozon_ad_report_repository: OzonAdReportRepository | None = None,
    ) -> None:
        self._settings = settings
        self._seller_client = OzonSellerClient(settings)
        self._performance_client = OzonPerformanceClient(settings)
        self._unit_economy_index_service = unit_economy_index_service
        self._ozon_ad_report_repository = ozon_ad_report_repository
        self._economics_calculator = OzonAdEconomicsCalculator()
        self._ad_attribution_calculator = OzonAdAttributionCalculator()
        self._report_file_parser = OzonAdReportFileParser()
        self._promotion_analytics_parser = OzonPromotionAnalyticsParser()
        self._total_sales_report_parser = OzonTotalSalesReportParser()
        self._product_sales_report_parser = OzonProductSalesReportParser()

    async def check_seller_connection(self) -> None:
        await self._seller_client.check_connection()

    async def check_performance_connection(self) -> None:
        await self._performance_client.check_connection()

    async def get_campaigns(self) -> list[OzonCampaign]:
        data = await self._performance_request_with_active_limit_wait(
            "GET",
            "/api/client/campaign",
        )
        raw_campaigns = data.get("list") or data.get("campaigns") or []

        if not isinstance(raw_campaigns, list):
            return []

        campaigns: list[OzonCampaign] = []
        for raw in raw_campaigns:
            if not isinstance(raw, dict):
                continue

            campaign_id = raw.get("id") or raw.get("campaign_id")
            title = raw.get("title") or raw.get("name")
            if campaign_id is None or title is None:
                continue

            state = raw.get("state") or raw.get("status")
            campaigns.append(
                OzonCampaign(
                    id=str(campaign_id),
                    title=str(title),
                    state=str(state) if state is not None else None,
                    raw=raw,
                )
            )

        return campaigns

    async def get_campaigns_for_period(
        self,
        date_from: str,
        date_to: str,
    ) -> list[OzonCampaign]:
        campaigns = await self.get_campaigns()
        return [
            campaign
            for campaign in campaigns
            if _campaign_overlaps_period(campaign, date_from, date_to)
        ]

    async def get_campaign_products(self, campaign_id: str) -> list[OzonCampaignProduct]:
        return await self._get_campaign_products(campaign_id, effective_date=None)

    async def _get_campaign_products(
        self,
        campaign_id: str,
        effective_date: str | None,
    ) -> list[OzonCampaignProduct]:
        data = await self._performance_client.request(
            "GET",
            f"/api/client/campaign/{campaign_id}/objects",
        )
        raw_products = data.get("list") or data.get("products") or data.get("objects") or []

        if not isinstance(raw_products, list):
            return []

        performance_objects: list[dict[str, object]] = []
        skus: list[str] = []

        for raw in raw_products:
            if not isinstance(raw, dict):
                continue

            object_id = raw.get("id") or raw.get("sku") or raw.get("product_id")
            if object_id is None:
                continue

            sku = str(object_id)
            performance_objects.append(raw)
            skus.append(sku)

        seller_products_by_sku = await self._seller_client.get_products_by_sku(skus)

        products: list[OzonCampaignProduct] = []
        for raw in performance_objects:
            object_id = raw.get("id") or raw.get("sku") or raw.get("product_id")
            if object_id is None:
                continue

            sku = str(object_id)
            seller_product = seller_products_by_sku.get(sku, {})
            offer_id = _optional_str(seller_product.get("offer_id"))
            unit_product = self._find_unit_product(offer_id, effective_date)
            products.append(
                OzonCampaignProduct(
                    campaign_id=campaign_id,
                    performance_object_id=sku,
                    seller_product_id=_optional_str(seller_product.get("id")),
                    sku=sku,
                    offer_id=offer_id,
                    title=_optional_str(seller_product.get("name")),
                    price=_optional_str(seller_product.get("price")),
                    old_price=_optional_str(seller_product.get("old_price")),
                    primary_image=_first_string(seller_product.get("primary_image")),
                    vat=_optional_str(seller_product.get("vat")),
                    unit_economy_match_key=offer_id,
                    unit_economy=unit_product,
                    raw=raw,
                )
            )

        return products

    async def get_campaign_efficiency(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonCampaignEfficiencyResponse:
        metrics = await self._get_campaign_performance_metrics(
            campaign_id=campaign_id,
            date_from=date_from,
            date_to=date_to,
        )
        products = await self.get_campaign_products(campaign_id)
        matched_products = sum(1 for product in products if product.unit_economy is not None)
        modeled_economics = self._economics_calculator.build_modeled_economics(
            metrics,
            products,
        )

        can_calculate_sku_profit = False
        profit_calculation_status = (
            "Campaign-level Ozon statistics loaded. SKU profit requires SKU-level "
            "statistics report; allocation by price or product count is forbidden."
        )

        return OzonCampaignEfficiencyResponse(
            date_from=date_from,
            date_to=date_to,
            metrics=metrics,
            modeled_economics=modeled_economics,
            products=products,
            matched_products=matched_products,
            can_calculate_sku_profit=can_calculate_sku_profit,
            profit_calculation_status=profit_calculation_status,
        )

    async def get_campaign_sku_efficiency(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
        report_uuid: str | None = None,
    ) -> OzonSkuEfficiencyResponse:
        campaign_ids = await self._resolve_campaign_ids(campaign_id, date_from, date_to)
        if len(campaign_ids) > 10:
            return await self._get_campaign_batch_sku_efficiency(
                campaign_id,
                campaign_ids,
                date_from,
                date_to,
            )

        return await self._get_campaign_ids_sku_efficiency(
            campaign_id,
            campaign_ids,
            date_from,
            date_to,
            report_uuid,
        )

    async def _get_campaign_ids_sku_efficiency(
        self,
        campaign_key: str,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
        report_uuid: str | None = None,
    ) -> OzonSkuEfficiencyResponse:
        unit_version_selection = self._select_unit_economy_version(date_from, date_to)
        effective_date = (
            unit_version_selection.version.valid_from
            if unit_version_selection is not None
            else date_to
        )
        should_recreate_missing_report = True
        cached_raw_report_csv: str | None = None
        if report_uuid is None:
            report_uuid = await self._create_statistics_report_with_limit_wait(
                campaign_ids,
                date_from,
                date_to,
            )
            await self._save_report_status(
                OzonStatisticsReportStatus(
                    campaign_id=campaign_key,
                    report_uuid=report_uuid,
                    state="CREATED",
                ),
                date_from,
                date_to,
            )
            report_state = await self._wait_statistics_report(report_uuid)
        else:
            saved_result = await self.get_saved_sku_efficiency_report_result(report_uuid)
            if saved_result is not None and _result_matches_unit_version(
                saved_result,
                unit_version_selection,
            ) and _result_has_campaign_source_metadata(
                saved_result
            ) and _result_uses_direct_sku_profit(
                saved_result
            ) and _result_uses_model_attribution_rule(
                saved_result
            ) and _result_uses_non_negative_profit_tax(
                saved_result
            ) and _result_has_product_identity(
                saved_result
            ) and _result_has_price_diagnostics(
                saved_result
            ) and _result_has_promotion_diagnostics(
                saved_result
            ) and self._result_has_current_unit_costs(saved_result, effective_date):
                return saved_result

            cached_raw_report_csv = await self._get_saved_raw_report_csv(report_uuid)
            if cached_raw_report_csv is not None:
                report_state = "OK"
                should_recreate_missing_report = False
            else:
                report_status = await self.get_statistics_report_status(
                    campaign_key,
                    report_uuid,
                )
                report_state = report_status.state
                if report_state in {"CREATED", "NOT_STARTED", "IN_PROGRESS"}:
                    report_state = await self._wait_statistics_report(report_uuid)
                elif report_state != "OK":
                    raise OzonReportNotReadyError(report_uuid, report_state)

        if cached_raw_report_csv is None:
            report_csv, report_uuid, report_state = await self._download_statistics_report_or_recreate(
                campaign_key=campaign_key,
                campaign_ids=campaign_ids,
                date_from=date_from,
                date_to=date_to,
                report_uuid=report_uuid,
                allow_recreate=should_recreate_missing_report,
            )
        else:
            report_csv = cached_raw_report_csv
        report_rows = _parse_sku_report_csv(report_csv)
        products_by_sku = await self._build_products_by_sku_from_report_rows(
            campaign_key,
            report_rows,
            effective_date,
            date_from,
            date_to,
        )
        calculation = self._economics_calculator.calculate_sku_efficiency(
            campaign_key,
            report_rows,
            products_by_sku,
        )

        response = OzonSkuEfficiencyResponse(
            campaign_id=campaign_key,
            date_from=date_from,
            date_to=date_to,
            rows=calculation.rows,
            total=calculation.total,
            adjustment_ad_spend_with_vat=calculation.adjustment_ad_spend_with_vat,
            adjustment_ad_spend_without_vat=calculation.adjustment_ad_spend_without_vat,
            report_uuid=report_uuid,
            report_state=report_state,
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            unit_economy_version_valid_from=(
                unit_version_selection.version.valid_from
                if unit_version_selection is not None
                else None
            ),
            unit_economy_workbook_path=(
                unit_version_selection.version.path
                if unit_version_selection is not None
                else None
            ),
            unit_economy_warning=(
                unit_version_selection.warning
                if unit_version_selection is not None
                else None
            ),
            promotion_report_state="API",
        )
        if self._ozon_ad_report_repository is not None:
            await self._ozon_ad_report_repository.save_efficiency_result(
                response=response,
                raw_report_csv=report_csv,
                unit_economy_version=(
                    unit_version_selection.version
                    if unit_version_selection is not None
                    else None
                ),
            )

        return response

    async def _get_campaign_batch_sku_efficiency(
        self,
        campaign_id: str,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
    ) -> OzonSkuEfficiencyResponse:
        responses: list[OzonSkuEfficiencyResponse] = []
        skipped_campaign_ids: list[str] = []
        for chunk in _chunked(campaign_ids, 10):
            chunk_responses, chunk_skipped_campaign_ids = (
                await self._get_campaign_batch_chunk_sku_efficiency(
                    chunk,
                    date_from,
                    date_to,
                )
            )
            responses.extend(chunk_responses)
            skipped_campaign_ids.extend(chunk_skipped_campaign_ids)

        calculation = self._economics_calculator.combine_sku_efficiency_responses(
            campaign_id,
            responses,
        )
        unit_version_selection = self._select_unit_economy_version(date_from, date_to)
        batch_warning = (
            f"Все кампании рассчитаны батчами по 10 из-за лимита Ozon. "
            f"Кампаний: {len(campaign_ids)}."
        )
        if skipped_campaign_ids:
            skipped_preview = ", ".join(skipped_campaign_ids[:20])
            skipped_tail = (
                f" и ещё {len(skipped_campaign_ids) - 20}"
                if len(skipped_campaign_ids) > 20
                else ""
            )
            batch_warning = (
                f"{batch_warning} Ozon запретил статистику по "
                f"{len(skipped_campaign_ids)} камп.: {skipped_preview}{skipped_tail}."
            )

        return OzonSkuEfficiencyResponse(
            campaign_id=campaign_id,
            date_from=date_from,
            date_to=date_to,
            rows=calculation.rows,
            total=calculation.total,
            adjustment_ad_spend_with_vat=calculation.adjustment_ad_spend_with_vat,
            adjustment_ad_spend_without_vat=calculation.adjustment_ad_spend_without_vat,
            report_uuid=_build_campaign_group_key(campaign_ids),
            report_state="OK",
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            unit_economy_version_valid_from=(
                unit_version_selection.version.valid_from
                if unit_version_selection is not None
                else None
            ),
            unit_economy_workbook_path=(
                unit_version_selection.version.path
                if unit_version_selection is not None
                else None
            ),
            unit_economy_warning=_append_warning(
                unit_version_selection.warning if unit_version_selection is not None else None,
                batch_warning,
            ),
            promotion_report_state="API",
        )

    async def _get_campaign_batch_chunk_sku_efficiency(
        self,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
    ) -> tuple[list[OzonSkuEfficiencyResponse], list[str]]:
        if not campaign_ids:
            return [], []

        chunk_key = _build_campaign_group_key(campaign_ids)
        saved_status = await self._find_saved_report_status(
            chunk_key,
            date_from,
            date_to,
        )
        try:
            if saved_status is None:
                response = await self._get_campaign_ids_sku_efficiency(
                    chunk_key,
                    campaign_ids,
                    date_from,
                    date_to,
                )
            else:
                response = await self._get_campaign_ids_sku_efficiency(
                    chunk_key,
                    campaign_ids,
                    date_from,
                    date_to,
                    report_uuid=saved_status.report_uuid,
                )
            return [response], []
        except OzonApiError as exc:
            if not _is_forbidden_campaign_list_error(str(exc)):
                raise

            if len(campaign_ids) == 1:
                return [], campaign_ids

            middle = len(campaign_ids) // 2
            left_responses, left_skipped = await self._get_campaign_batch_chunk_sku_efficiency(
                campaign_ids[:middle],
                date_from,
                date_to,
            )
            right_responses, right_skipped = await self._get_campaign_batch_chunk_sku_efficiency(
                campaign_ids[middle:],
                date_from,
                date_to,
            )
            return left_responses + right_responses, left_skipped + right_skipped

    async def get_campaign_sku_efficiency_composite(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonSkuEfficiencyResponse:
        if self._unit_economy_index_service is None:
            return await self.get_campaign_sku_efficiency(campaign_id, date_from, date_to)

        segments = self._unit_economy_index_service.build_period_segments(
            date_from,
            date_to,
        )
        if campaign_id == "ALL":
            campaign_ids_by_segment = [
                (
                    segment,
                    await self._resolve_campaign_ids(
                        campaign_id,
                        segment.date_from,
                        segment.date_to,
                    ),
                )
                for segment in segments
            ]
            segment_responses = [
                await self._get_or_build_campaign_ids_sku_efficiency_segment(
                    campaign_id,
                    segment_campaign_ids,
                    segment.date_from,
                    segment.date_to,
                )
                for segment, segment_campaign_ids in campaign_ids_by_segment
            ]
        else:
            segment_responses = [
                await self._get_or_build_sku_efficiency_segment(
                    campaign_id,
                    segment.date_from,
                    segment.date_to,
                )
                for segment in segments
            ]
        if len(segment_responses) == 1:
            single_response = segment_responses[0]
            return single_response.model_copy(
                update={
                    "is_composite": False,
                    "segments": [
                        _build_response_segment(single_response)
                    ],
                }
            )

        calculation = self._economics_calculator.combine_sku_efficiency_responses(
            campaign_id,
            segment_responses,
        )
        response = OzonSkuEfficiencyResponse(
            campaign_id=campaign_id,
            date_from=date_from,
            date_to=date_to,
            rows=calculation.rows,
            total=calculation.total,
            adjustment_ad_spend_with_vat=calculation.adjustment_ad_spend_with_vat,
            adjustment_ad_spend_without_vat=calculation.adjustment_ad_spend_without_vat,
            report_uuid=_build_composite_report_uuid(campaign_id, date_from, date_to),
            report_state="OK",
            unit_economy_warning=_build_composite_warning(segments),
            is_composite=True,
            promotion_report_state="API",
            segments=[
                _build_response_segment(segment_response)
                for segment_response in segment_responses
            ],
        )

        return response

    async def import_sku_efficiency_report(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
        file: UploadFile,
    ) -> OzonSkuEfficiencyResponse:
        content = await file.read()
        filename = file.filename or "ozon-report.csv"
        parsed_report = self._report_file_parser.parse(filename, content)
        unit_version_selection = self._select_unit_economy_version(date_from, date_to)
        effective_date = (
            unit_version_selection.version.valid_from
            if unit_version_selection is not None
            else date_to
        )
        products_by_sku = await self._build_import_products_by_sku(
            campaign_id,
            parsed_report.rows,
            effective_date,
            date_from,
            date_to,
        )
        calculation = self._economics_calculator.calculate_sku_efficiency(
            campaign_id,
            parsed_report.rows,
            products_by_sku,
        )
        report_uuid = _build_import_report_uuid(
            campaign_id,
            date_from,
            date_to,
            filename,
            content,
        )
        response = OzonSkuEfficiencyResponse(
            campaign_id=campaign_id,
            date_from=date_from,
            date_to=date_to,
            rows=calculation.rows,
            total=calculation.total,
            adjustment_ad_spend_with_vat=calculation.adjustment_ad_spend_with_vat,
            adjustment_ad_spend_without_vat=calculation.adjustment_ad_spend_without_vat,
            report_uuid=report_uuid,
            report_state="IMPORTED",
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            unit_economy_version_valid_from=(
                unit_version_selection.version.valid_from
                if unit_version_selection is not None
                else None
            ),
            unit_economy_workbook_path=(
                unit_version_selection.version.path
                if unit_version_selection is not None
                else None
            ),
            unit_economy_warning=(
                unit_version_selection.warning
                if unit_version_selection is not None
                else None
            ),
            promotion_report_state="API",
        )
        if self._ozon_ad_report_repository is not None:
            await self._ozon_ad_report_repository.save_efficiency_result(
                response=response,
                raw_report_csv=parsed_report.raw_report_csv,
                unit_economy_version=(
                    unit_version_selection.version
                    if unit_version_selection is not None
                    else None
                ),
            )

        return response

    async def get_ad_attribution_from_api(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonAdAttributionResponse:
        response = await self.get_campaign_sku_efficiency_composite(
            campaign_id,
            date_from,
            date_to,
        )
        return self._ad_attribution_calculator.calculate_api(response)

    async def import_ad_attribution_report(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
        file: UploadFile,
    ) -> OzonAdAttributionResponse:
        content = await file.read()
        filename = file.filename or "ozon-promotion-analytics.xlsx"
        parsed_report = self._promotion_analytics_parser.parse(filename, content)
        selected_report = _filter_promotion_report_by_campaign(
            parsed_report,
            campaign_id,
        )
        date_from = selected_report.period_date_from or date_from
        date_to = selected_report.period_date_to or date_to
        unit_version_selection = self._select_unit_economy_version(date_from, date_to)
        effective_date = (
            unit_version_selection.version.valid_from
            if unit_version_selection is not None
            else date_to
        )
        report_rows = [
            {
                "sku": sku,
                "Название товара": title or "",
            }
            for sku, title in {
                **{
                    row.promoted_sku: row.title
                    for row in selected_report.statistics
                },
                **{
                    row.purchased_sku: row.purchased_title
                    for row in selected_report.union
                },
            }.items()
        ]
        campaign_for_lookup = (
            campaign_id
            if campaign_id != "ALL"
            else (
                selected_report.statistics[0].campaign_id
                if selected_report.statistics
                else "ALL"
            )
        )
        campaign_products = await self._build_import_products_by_sku(
            campaign_for_lookup,
            report_rows,
            effective_date,
            date_from,
            date_to,
        )
        products_by_sku = {
            sku: _resolve_attribution_product(sku, product)
            for sku, product in campaign_products.items()
        }
        response = self._ad_attribution_calculator.calculate_import(
            campaign_id=campaign_id,
            date_from=date_from,
            date_to=date_to,
            report=selected_report,
            products_by_sku=products_by_sku,
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            unit_economy_warning=(
                unit_version_selection.warning
                if unit_version_selection is not None
                else None
            ),
        )
        if self._ozon_ad_report_repository is not None:
            await self._ozon_ad_report_repository.save_promotion_analytics_import(
                campaign_id=campaign_id,
                date_from=date_from,
                date_to=date_to,
                source_filename=filename,
                response=response,
            )
        return response

    async def list_promotion_analytics_imports(
        self,
        limit: int = 100,
    ) -> list[OzonPromotionAnalyticsImportSummary]:
        if self._ozon_ad_report_repository is None:
            return []

        return await self._ozon_ad_report_repository.list_promotion_analytics_imports(
            limit
        )

    async def get_promotion_analytics_import(
        self,
        import_id: int,
    ) -> OzonAdAttributionResponse | None:
        if self._ozon_ad_report_repository is None:
            return None

        return await self._ozon_ad_report_repository.get_promotion_analytics_import(
            import_id
        )

    async def import_product_sales_report(
        self,
        file: UploadFile,
    ) -> OzonProductSalesResponse:
        if self._ozon_ad_report_repository is None:
            raise ValueError("Хранилище отчётов не настроено")

        content = await file.read()
        filename = file.filename or "ozon-product-sales.xlsx"
        try:
            report = self._product_sales_report_parser.parse(filename, content)
        except ProductSalesReportError as exc:
            raise ValueError(str(exc)) from exc

        await self._ozon_ad_report_repository.save_product_sales_rows(report.rows)
        date_from = min(row.date_from for row in report.rows)
        date_to = max(row.date_to for row in report.rows)
        sku_count = len({row.sku for row in report.rows})
        await self._ozon_ad_report_repository.save_product_sales_import(
            date_from=date_from,
            date_to=date_to,
            source_filename=filename,
            sku_count=sku_count,
            row_count=len(report.rows),
            has_daily_breakdown=report.has_daily_breakdown,
        )
        return await self.get_product_sales_profit(date_from, date_to)

    async def list_product_sales_imports(
        self,
        limit: int = 100,
    ) -> list[OzonProductSalesImportSummary]:
        if self._ozon_ad_report_repository is None:
            return []

        return await self._ozon_ad_report_repository.list_product_sales_imports(limit)

    async def get_product_sales_profit(
        self,
        date_from: str,
        date_to: str,
    ) -> OzonProductSalesResponse:
        if self._ozon_ad_report_repository is None:
            raise ValueError("Хранилище отчётов не настроено")

        raw_rows = await self._ozon_ad_report_repository.get_product_sales_rows(
            date_from,
            date_to,
        )
        if not raw_rows:
            raise ValueError(
                "За этот период нет загруженных данных о продажах. "
                "Сначала загрузите отчёт «Аналитика по товарам»."
            )

        unit_version_selection = self._select_unit_economy_version(date_from, date_to)
        effective_date = (
            unit_version_selection.version.valid_from
            if unit_version_selection is not None
            else date_to
        )

        rows_by_sku = _group_product_sales_rows_by_sku(raw_rows)
        result_rows = [
            self._build_product_sales_row(sku, rows, effective_date)
            for sku, rows in rows_by_sku.items()
        ]
        result_rows.sort(
            key=lambda row: (
                row.profit_before_ads is None,
                -(row.profit_before_ads or 0),
            )
        )
        total = _build_product_sales_total(result_rows)

        ad_import = await self._ozon_ad_report_repository.get_promotion_analytics_import_by_period(
            "ALL",
            date_from,
            date_to,
        )
        warning = (
            unit_version_selection.warning
            if unit_version_selection is not None
            else None
        )
        if ad_import is not None:
            ad_spend = ad_import.direct_total.spend_without_vat
            total = total.model_copy(
                update={
                    "ad_spend_without_vat": ad_spend,
                    "net_profit": (
                        total.profit_before_ads - ad_spend
                        if total.profit_before_ads is not None
                        else None
                    ),
                    "drr_percent": (
                        ad_spend / total.revenue_without_vat * 100
                        if total.revenue_without_vat > 0
                        else None
                    ),
                }
            )
            ad_spend_matched = True
        else:
            ad_spend_matched = False
            warning = _append_warning(
                warning,
                "За этот период нет сохранённой загрузки «Реклама Ozon» (кампания "
                "«Все кампании») — расход на рекламу не вычтен из чистой прибыли.",
            )

        return OzonProductSalesResponse(
            date_from=date_from,
            date_to=date_to,
            rows=result_rows,
            total=total,
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            unit_economy_warning=warning,
            ad_spend_matched=ad_spend_matched,
            warning=warning,
        )

    def _build_product_sales_row(
        self,
        sku: str,
        rows: list[ProductSalesRow],
        effective_date: str,
    ) -> OzonProductSalesRow:
        offer_id = next((row.offer_id for row in rows if row.offer_id), None)
        title = next((row.title for row in rows if row.title), None)
        ordered_units = sum(row.ordered_units for row in rows)
        redeemed_units = sum(row.redeemed_units for row in rows)
        cancelled_units = sum(row.cancelled_units for row in rows)
        # "Скидка от вашей цены" is Ozon's own "Баллы за скидку" promotion: the
        # buyer sees a lower price, but Ozon pays the seller their own set price
        # (avg_price) minus commission and reimburses the gap as points. Applying
        # the discount here would understate revenue by the reimbursed amount.
        revenue_with_vat = sum(row.redeemed_units * row.avg_price for row in rows)
        revenue_without_vat = float(without_vat(Decimal(str(revenue_with_vat))))

        unit_product = self._find_unit_product(
            offer_id,
            effective_date,
        ) or self._find_unit_product_by_sku(sku, effective_date)
        unit_expense = (
            unit_product.expense_with_ozon_commission
            if unit_product is not None
            else None
        )
        profit_before_ads: float | None = None
        if unit_expense is not None:
            profit_before_tax = revenue_without_vat - unit_expense * redeemed_units
            profit_before_ads = float(
                net_after_profit_tax(Decimal(str(profit_before_tax)))
            )

        return OzonProductSalesRow(
            sku=sku,
            offer_id=offer_id,
            title=title,
            ordered_units=ordered_units,
            redeemed_units=redeemed_units,
            cancelled_units=cancelled_units,
            revenue_with_vat=revenue_with_vat,
            revenue_without_vat=revenue_without_vat,
            unit_expense_with_ozon_commission=unit_expense,
            profit_before_ads=profit_before_ads,
            matched_unit_economy=unit_expense is not None,
        )

    async def list_daily_profit_snapshots(
        self,
        limit: int = 120,
    ) -> list[OzonDailyProfitSnapshot]:
        if self._ozon_ad_report_repository is None:
            return []

        return await self._ozon_ad_report_repository.list_daily_profit_snapshots(
            limit
        )

    async def delete_daily_profit_snapshot(self, date_str: str, run_type: str) -> None:
        if self._ozon_ad_report_repository is None:
            return

        await self._ozon_ad_report_repository.delete_daily_profit_snapshot(
            date_str,
            run_type,
        )

    async def export_daily_profit_xlsx(
        self,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> bytes:
        snapshots = await self.list_daily_profit_snapshots(limit=1000)
        if date_from is not None:
            snapshots = [snapshot for snapshot in snapshots if snapshot.date >= date_from]
        if date_to is not None:
            snapshots = [snapshot for snapshot in snapshots if snapshot.date <= date_to]
        snapshots = sorted(snapshots, key=lambda snapshot: (snapshot.date, snapshot.run_type))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise ValueError("openpyxl недоступен: экспорт Excel невозможен") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Реестр дневной прибыли"
        headers = [
            "Дата",
            "Тип",
            "Заказано, шт",
            "Отменено, шт",
            "Нетто, шт",
            "Выручка с НДС",
            "Выручка без НДС",
            "Прибыль до рекламы",
            "Реклама без НДС",
            "Чистая прибыль",
            "Версия юнитки",
            "Посчитано",
            "Предупреждение",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for snapshot in snapshots:
            sheet.append([
                snapshot.date,
                RUN_TYPE_LABELS.get(snapshot.run_type, snapshot.run_type),
                snapshot.ordered_units,
                snapshot.cancelled_units,
                snapshot.net_units,
                round(snapshot.revenue_with_vat, 2),
                round(snapshot.revenue_without_vat, 2),
                round(snapshot.profit_before_ads, 2) if snapshot.profit_before_ads is not None else None,
                round(snapshot.ad_spend_without_vat, 2) if snapshot.ad_spend_without_vat is not None else None,
                round(snapshot.net_profit, 2) if snapshot.net_profit is not None else None,
                snapshot.unit_economy_version,
                snapshot.computed_at,
                snapshot.warning,
            ])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(length + 2, 10), 60
            )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def export_orders_xlsx(self, date_from: str, date_to: str) -> bytes:
        all_days = []
        current = date.fromisoformat(date_from)
        end = date.fromisoformat(date_to)
        while current <= end:
            all_days.append(current.isoformat())
            current += timedelta(days=1)

        orders: list[OzonOrderLookupResponse] = []
        for day in all_days:
            orders.extend(await self.list_orders_profit_for_date(day, auto_compute=True))

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise ValueError("openpyxl недоступен: экспорт Excel невозможен") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Заказы"
        headers = [
            "Дата",
            "Время",
            "Номер отправления",
            "Номер заказа",
            "Статус",
            "Артикул",
            "SKU",
            "Товар",
            "Кол-во",
            "Цена",
            "Выручка без НДС",
            "Себестоимость + комиссия",
            "Чистая прибыль по товару",
            "Чистая прибыль по заказу",
            "Предупреждение",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for order in orders:
            order_date = order.in_process_at[:10] if order.in_process_at else ""
            order_time = order.in_process_at[11:16] if order.in_process_at else ""
            if not order.items:
                sheet.append([
                    order_date, order_time, order.posting_number, order.order_number,
                    order.status_label, None, None, None, None, None,
                    round(order.revenue_without_vat_total, 2), None, None,
                    round(order.net_profit_total, 2) if order.net_profit_total is not None else None,
                    order.warning,
                ])
                continue
            for item in order.items:
                sheet.append([
                    order_date,
                    order_time,
                    order.posting_number,
                    order.order_number,
                    order.status_label,
                    item.offer_id,
                    item.sku,
                    item.name,
                    item.quantity,
                    round(item.price_with_vat, 2),
                    round(item.revenue_without_vat, 2),
                    round(item.cost_with_commission, 2) if item.cost_with_commission is not None else None,
                    round(item.net_profit, 2) if item.net_profit is not None else None,
                    round(order.net_profit_total, 2) if order.net_profit_total is not None else None,
                    order.warning,
                ])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(length + 2, 10), 60
            )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def export_orders_for_date_xlsx(self, date_str: str) -> bytes:
        """One row per ORDER (not per line item, unlike export_orders_xlsx)
        with the delivery detail - address, lift, the "Перевели за доставку"
        breakdown, actual cost, result - that the "Заказы" page's table and
        order card show but the period export above doesn't carry.
        """
        orders = await self.list_orders_profit_for_date(date_str, auto_compute=True)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise ValueError("openpyxl недоступен: экспорт Excel невозможен") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Заказы"
        headers = [
            "Дата",
            "Время",
            "Номер отправления",
            "Номер заказа",
            "Статус",
            "Выручка без НДС",
            "Чистая прибыль",
            "Адрес доставки",
            "Способ доставки",
            "Подъём",
            "Перевели: доставка",
            "Перевели: подъём",
            "Перевели: компенсация логистики",
            "Перевели: итого",
            "Доставка факт",
            "Доставка итог за заказ",
            "Чистая прибыль и доставка",
            "Предупреждение",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for order in orders:
            order_date = order.in_process_at[:10] if order.in_process_at else ""
            order_time = order.in_process_at[11:16] if order.in_process_at else ""
            lift_text = order.lift_option_label or "—"
            if order.lift_floor:
                lift_text += f" ({order.lift_floor} этаж)"
            net_profit_and_delivery = (
                order.net_profit_with_delivery_actual
                if order.net_profit_with_delivery_actual is not None
                else None
            )
            sheet.append([
                order_date,
                order_time,
                order.posting_number,
                order.order_number,
                order.status_label,
                round(order.revenue_without_vat_total, 2),
                round(order.net_profit_total, 2) if order.net_profit_total is not None else None,
                order.delivery_address,
                order.delivery_method_name,
                lift_text,
                round(order.delivery_price_transferred, 2)
                if order.delivery_price_transferred is not None else None,
                round(order.lift_price, 2) if order.lift_price is not None else None,
                round(order.logistics_compensation_total, 2)
                if order.logistics_compensation_total is not None else None,
                round(order.delivery_total_transferred, 2)
                if order.delivery_total_transferred is not None else None,
                round(order.delivery_cost_actual, 2)
                if order.delivery_cost_actual is not None else None,
                round(order.delivery_result, 2) if order.delivery_result is not None else None,
                round(net_profit_and_delivery, 2) if net_profit_and_delivery is not None else None,
                order.warning,
            ])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(length + 2, 10), 60
            )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def get_ad_spend_for_date(self, date_str: str) -> tuple[float | None, str | None]:
        """Public wrapper around the same ad-spend lookup the daily-profit
        snapshot uses, for callers (like the "Заказы" page for a past date)
        that just need the number - no sales matching, no persistence.
        """
        return await self._get_period_ad_spend_without_vat(date_str, date_str)

    async def compute_daily_profit_snapshot(
        self,
        date_str: str,
        run_type: str,
    ) -> OzonDailyProfitSnapshot:
        """Builds the registry's daily snapshot from the SAME per-posting
        pipeline as the "Заказы" page (list_orders_profit_for_date), not
        Ozon's aggregate /v1/analytics/data SKU report used before this.

        Two concrete problems with the old aggregate-report approach drove
        this switch (both confirmed live against real orders on 27.08 -
        Реестр showed 909 342 ₽ revenue / 134 765 ₽ net profit for a day
        where "Заказы" showed 830 618 ₽ / 77 936 ₽ for the same date):
        1. Revenue there summed ALL ordered units' revenue unconditionally,
           including cancelled orders - "Заказы"'s per-date tiles exclude
           cancelled orders from revenue entirely. Now both agree (revenue
           below only sums active/non-cancelled orders).
        2. It had no way to know about manually-entered "фактическая
           стоимость доставки" per order (that only exists on individual FBS
           postings) - so it could never reflect real logistics deviations,
           only юнитка's estimate. net_profit_and_delivery below fixes that.
        """
        if self._ozon_ad_report_repository is None:
            raise ValueError("Хранилище отчётов не настроено")

        unit_version_selection = self._select_unit_economy_version(date_str, date_str)
        warning = (
            unit_version_selection.warning
            if unit_version_selection is not None
            else None
        )

        orders = await self.list_orders_profit_for_date(date_str, auto_compute=True)
        active_orders = [order for order in orders if not order.is_cancelled]
        cancelled_orders = [order for order in orders if order.is_cancelled]

        # "units" here means item quantity (matches the old /v1/analytics/data
        # "ordered_units" metric this replaces), not order/posting count.
        ordered_units_total = sum(
            item.quantity for order in orders for item in order.items
        )
        cancelled_units_total = sum(
            item.quantity for order in cancelled_orders for item in order.items
        )
        net_units_total = ordered_units_total - cancelled_units_total

        revenue_with_vat_total = sum(
            item.price_with_vat * item.quantity
            for order in active_orders
            for item in order.items
        )
        revenue_without_vat_total = sum(
            order.revenue_without_vat_total for order in active_orders
        )

        matched_orders = [
            order for order in active_orders if order.net_profit_total is not None
        ]
        profit_before_ads_total = sum(
            order.net_profit_total for order in matched_orders
        )
        matched_any = bool(matched_orders)
        if matched_orders and len(matched_orders) < len(active_orders):
            warning = _append_warning(
                warning,
                "У части заказов есть товары, не найденные в юнит-экономике - "
                "они не учтены в итоговой прибыли.",
            )

        ad_spend_without_vat, ad_spend_warning = await self._get_period_ad_spend_without_vat(
            date_str,
            date_str,
        )
        warning = _append_warning(warning, ad_spend_warning) if ad_spend_warning else warning

        net_profit = (
            profit_before_ads_total - ad_spend_without_vat
            if matched_any and ad_spend_without_vat is not None
            else None
        )
        # Only orders with a manually-entered "фактическая доставка" have a
        # known delivery_result - orders without one contribute 0, same
        # convention as get_today_orders_summary's net_profit_and_delivery_total.
        delivery_result_total = sum(
            order.delivery_result or 0 for order in matched_orders
        )
        net_profit_and_delivery = (
            net_profit + delivery_result_total if net_profit is not None else None
        )

        snapshot = OzonDailyProfitSnapshot(
            date=date_str,
            run_type=run_type,
            ordered_units=ordered_units_total,
            cancelled_units=cancelled_units_total,
            net_units=net_units_total,
            revenue_with_vat=revenue_with_vat_total,
            revenue_without_vat=revenue_without_vat_total,
            profit_before_ads=profit_before_ads_total if matched_any else None,
            ad_spend_without_vat=ad_spend_without_vat,
            net_profit=net_profit,
            net_profit_and_delivery=net_profit_and_delivery,
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            warning=warning,
            computed_at=datetime.now(UTC).isoformat(),
        )
        await self._ozon_ad_report_repository.save_daily_profit_snapshot(snapshot)
        return snapshot

    async def compute_order_profit(self, posting_number: str) -> OzonOrderLookupResponse:
        posting = await self._seller_client.get_fbs_posting(posting_number)
        order = self._build_order_lookup_response(posting, posting_number)
        if self._ozon_ad_report_repository is not None:
            actual_cost = await self._ozon_ad_report_repository.get_delivery_actual_cost(
                order.posting_number
            )
            order = _apply_delivery_actual_cost(order, actual_cost)
            await self._ozon_ad_report_repository.save_order_profit_snapshot(order)
        return order

    async def save_delivery_actual_cost(
        self,
        posting_number: str,
        actual_cost: float,
    ) -> OzonOrderLookupResponse:
        if self._ozon_ad_report_repository is None:
            raise ValueError("Хранилище отчётов не настроено")
        await self._ozon_ad_report_repository.save_delivery_actual_cost(
            posting_number,
            actual_cost,
        )
        return await self.compute_order_profit(posting_number)

    def _moscow_day_bounds_iso(self, date_str: str) -> tuple[str, str]:
        day = date.fromisoformat(date_str)
        start_moscow = datetime(day.year, day.month, day.day, tzinfo=MOSCOW_TZ)
        now_moscow = datetime.now(MOSCOW_TZ)
        end_moscow = now_moscow if day == now_moscow.date() else start_moscow + timedelta(days=1)
        return (
            start_moscow.astimezone(UTC).isoformat().replace("+00:00", "Z"),
            end_moscow.astimezone(UTC).isoformat().replace("+00:00", "Z"),
        )

    async def compute_orders_profit_for_date(
        self,
        date_str: str,
    ) -> list[OzonOrderLookupResponse]:
        if self._ozon_ad_report_repository is None:
            raise ValueError("Хранилище отчётов не настроено")

        since_iso, to_iso = self._moscow_day_bounds_iso(date_str)
        postings = await self._seller_client.list_fbs_postings(since_iso, to_iso)
        posting_numbers = [
            str(posting.get("posting_number") or "")
            for posting in postings
            if posting.get("posting_number")
        ]
        actual_costs = await self._ozon_ad_report_repository.get_delivery_actual_costs(
            posting_numbers
        )

        orders: list[OzonOrderLookupResponse] = []
        for posting_number in posting_numbers:
            try:
                # /v3/posting/fbs/list doesn't carry delivery_price or the full
                # prr_option (price/floor) - only /v3/posting/fbs/get does, so
                # each posting needs its own detail call to get that data.
                full_posting = await self._seller_client.get_fbs_posting(posting_number)
                order = self._build_order_lookup_response(full_posting, posting_number)
            except Exception:  # noqa: BLE001 - one bad posting must not break the batch
                logger.exception("Failed to compute profit for posting %s", posting_number)
                continue
            order = _apply_delivery_actual_cost(order, actual_costs.get(posting_number))
            await self._ozon_ad_report_repository.save_order_profit_snapshot(order)
            orders.append(order)

        return orders

    async def compute_today_orders_profit(self) -> list[OzonOrderLookupResponse]:
        today_iso = datetime.now(MOSCOW_TZ).date().isoformat()
        return await self.compute_orders_profit_for_date(today_iso)

    async def refresh_today_orders_and_ad_spend(self) -> list[OzonOrderLookupResponse]:
        """Recomputes today's orders now, kicks off an ad spend refresh in the background.

        Orders are fast (one Ozon list call, no ads involved) so those are
        awaited and returned immediately. Ad spend is NOT awaited here: its
        promotion-report fallback generates one Ozon report per 10 running
        campaigns, sequentially, and each one can queue behind Ozon's "one
        report at a time" limit for up to ~15 minutes - across a few batches
        that's easily 30-90 minutes in the worst case. Blocking the refresh
        button on that turned it into an indefinite hang in practice
        (confirmed live: a request sat for 200s+ with nothing back). Running
        it in the background instead means the button responds in seconds,
        and the ad spend tile picks up the fresh number on its own next
        poll once the background task finishes.
        """
        today_iso = datetime.now(MOSCOW_TZ).date().isoformat()
        orders = await self.compute_today_orders_profit()
        asyncio.create_task(self._refresh_today_ad_spend_in_background(today_iso))
        return orders

    def trigger_today_ad_spend_refresh(self) -> None:
        """Kicks off just the ad-spend recompute (no orders refetch) in the
        background, for a dedicated "Обновить рекламу" action separate from
        the full "Обновить сейчас" - see refresh_today_orders_and_ad_spend's
        docstring for why this can't be awaited directly (can take up to
        30-90 minutes in the worst case). The frontend polls
        get_today_orders_summary afterwards to notice when it lands.
        """
        today_iso = datetime.now(MOSCOW_TZ).date().isoformat()
        asyncio.create_task(self._refresh_today_ad_spend_in_background(today_iso))

    async def _refresh_today_ad_spend_in_background(self, date_str: str) -> None:
        try:
            await self.compute_daily_profit_snapshot(date_str, "preliminary")
        except Exception:  # noqa: BLE001 - background task, must not go unnoticed but must not crash anything
            logger.exception("Background ad spend refresh failed for %s", date_str)

    async def list_orders_profit_for_date(
        self,
        date_str: str,
        auto_compute: bool = True,
    ) -> list[OzonOrderLookupResponse]:
        if self._ozon_ad_report_repository is None:
            return []

        since_iso, to_iso = self._moscow_day_bounds_iso(date_str)
        cached = await self._ozon_ad_report_repository.list_order_profit_snapshots(
            since_iso=since_iso,
            to_iso=to_iso,
        )
        if cached or not auto_compute:
            return cached

        return await self.compute_orders_profit_for_date(date_str)

    async def list_today_orders_profit(self) -> list[OzonOrderLookupResponse]:
        today_iso = datetime.now(MOSCOW_TZ).date().isoformat()
        # Today relies on the 30-min background scheduler rather than
        # computing on every page load - past days (see get_orders_for_date)
        # have no such scheduler, so those auto-compute on first request.
        return await self.list_orders_profit_for_date(today_iso, auto_compute=False)

    async def get_today_orders_summary(self) -> OzonTodayOrdersSummary:
        orders = await self.list_today_orders_profit()
        active_orders = [order for order in orders if not order.is_cancelled]

        revenue_without_vat_total = sum(
            order.revenue_without_vat_total for order in active_orders
        )
        matched_orders = [
            order for order in active_orders if order.net_profit_total is not None
        ]
        net_profit_before_ads_total = sum(
            order.net_profit_total for order in matched_orders
        )

        warning: str | None = None
        if len(matched_orders) < len(active_orders):
            warning = _append_warning(
                warning,
                "У части заказов есть товары, не найденные в юнит-экономике - "
                "они не учтены в итоговой прибыли.",
            )

        # The live-fetch path can take a couple of minutes for today (Ozon's
        # fast /expense export has nothing yet, so it falls back to the
        # per-campaign promotion report) - too slow for a page load. Instead,
        # read the value the hourly daily-profit scheduler already computed
        # for today's "preliminary" snapshot, so this stays fast and the two
        # pages agree on the same number.
        today_iso = datetime.now(MOSCOW_TZ).date().isoformat()
        ad_spend_without_vat: float | None = None
        ad_spend_warning: str | None = None
        if self._ozon_ad_report_repository is not None:
            daily_snapshots = await self._ozon_ad_report_repository.list_daily_profit_snapshots(
                limit=50
            )
            today_snapshot = next(
                (
                    snapshot
                    for snapshot in daily_snapshots
                    if snapshot.date == today_iso and snapshot.run_type == "preliminary"
                ),
                None,
            )
            if today_snapshot is not None:
                ad_spend_without_vat = today_snapshot.ad_spend_without_vat
                if ad_spend_without_vat is None and today_snapshot.warning:
                    ad_spend_warning = today_snapshot.warning
            else:
                ad_spend_warning = (
                    "Расход на рекламу за сегодня ещё не подтянулся фоновым "
                    "обновлением - появится в течение часа."
                )
        if ad_spend_warning:
            warning = _append_warning(warning, ad_spend_warning)

        net_profit_total = (
            net_profit_before_ads_total - ad_spend_without_vat
            if ad_spend_without_vat is not None
            else None
        )
        # Only orders with a manually-entered "фактическая доставка" have a
        # known delivery_result - orders without one contribute 0 (i.e. the
        # tile falls back to plain net profit for them, rather than guessing).
        delivery_result_total = sum(
            order.delivery_result or 0 for order in matched_orders
        )
        net_profit_and_delivery_total = (
            net_profit_total + delivery_result_total
            if net_profit_total is not None
            else None
        )

        return OzonTodayOrdersSummary(
            orders_count=len(active_orders),
            matched_orders_count=len(matched_orders),
            revenue_without_vat_total=revenue_without_vat_total,
            net_profit_before_ads_total=(
                net_profit_before_ads_total if active_orders else None
            ),
            ad_spend_without_vat=ad_spend_without_vat,
            net_profit_total=net_profit_total if active_orders else None,
            net_profit_and_delivery_total=(
                net_profit_and_delivery_total if active_orders else None
            ),
            warning=warning,
            computed_at=datetime.now(UTC).isoformat(),
        )

    def _build_order_lookup_response(
        self,
        posting: dict[str, object],
        posting_number: str,
    ) -> OzonOrderLookupResponse:
        status = str(posting.get("status") or "")
        in_process_at = posting.get("in_process_at")
        in_process_at = str(in_process_at) if in_process_at else None
        effective_date = (
            str(in_process_at)[:10] if in_process_at else datetime.now(UTC).date().isoformat()
        )

        unit_version_selection = self._select_unit_economy_version(
            effective_date,
            effective_date,
        )
        warning = (
            unit_version_selection.warning if unit_version_selection is not None else None
        )

        raw_products = posting.get("products")
        raw_products = raw_products if isinstance(raw_products, list) else []

        items: list[OzonOrderItem] = []
        revenue_without_vat_total = 0.0
        net_profit_total = 0.0
        all_matched = True

        for raw_product in raw_products:
            if not isinstance(raw_product, dict):
                continue

            offer_id = str(raw_product.get("offer_id") or "")
            if not offer_id:
                continue

            name = str(raw_product.get("name") or offer_id)
            sku = raw_product.get("sku")
            sku = str(sku) if sku else None
            quantity = int(raw_product.get("quantity") or 0)
            price_with_vat = float(raw_product.get("price") or 0)

            row_revenue_with_vat = price_with_vat * quantity
            row_revenue_without_vat = float(
                without_vat(Decimal(str(row_revenue_with_vat)))
            )
            revenue_without_vat_total += row_revenue_without_vat

            unit_product = self._find_unit_product(offer_id, effective_date)
            if unit_product is None or unit_product.expense_with_ozon_commission is None:
                all_matched = False
                items.append(
                    OzonOrderItem(
                        offer_id=offer_id,
                        sku=sku,
                        name=name,
                        quantity=quantity,
                        price_with_vat=price_with_vat,
                        revenue_without_vat=row_revenue_without_vat,
                        unit_economy_matched=False,
                    )
                )
                continue

            cost_with_commission = unit_product.expense_with_ozon_commission * quantity
            profit_before_tax = row_revenue_without_vat - cost_with_commission
            net_profit = float(net_after_profit_tax(Decimal(str(profit_before_tax))))
            net_profit_total += net_profit

            # Markup/commission % are rates (per unit, not scaled by quantity);
            # cost/commission/logistics rubles mirror cost_with_commission's
            # convention of being the row's total across quantity.
            markup_percent = None
            if (
                unit_product.price_without_vat is not None
                and unit_product.cost_without_vat is not None
                and unit_product.cost_without_vat > 0
            ):
                markup_percent = (
                    (unit_product.price_without_vat - unit_product.cost_without_vat)
                    / unit_product.cost_without_vat
                    * 100
                )
            ozon_commission_percent = None
            if (
                unit_product.ozon_commission is not None
                and unit_product.price_without_vat is not None
                and unit_product.price_without_vat > 0
            ):
                ozon_commission_percent = (
                    unit_product.ozon_commission / unit_product.price_without_vat * 100
                )

            items.append(
                OzonOrderItem(
                    offer_id=offer_id,
                    sku=sku,
                    name=name,
                    quantity=quantity,
                    price_with_vat=price_with_vat,
                    revenue_without_vat=row_revenue_without_vat,
                    unit_economy_matched=True,
                    cost_with_commission=cost_with_commission,
                    profit_before_tax=profit_before_tax,
                    net_profit=net_profit,
                    unit_economy_price=unit_product.price_with_vat,
                    markup_percent=markup_percent,
                    ozon_commission_rub=(
                        unit_product.ozon_commission * quantity
                        if unit_product.ozon_commission is not None
                        else None
                    ),
                    ozon_commission_percent=ozon_commission_percent,
                    unit_economy_cost=(
                        unit_product.expense_cost * quantity
                        if unit_product.expense_cost is not None
                        else None
                    ),
                    logistics_compensation=(
                        unit_product.logistics_compensation * quantity
                        if unit_product.logistics_compensation is not None
                        else None
                    ),
                )
            )

        if not all_matched:
            warning = _append_warning(
                warning,
                "Часть товаров заказа не найдена в юнит-экономике по артикулу - "
                "прибыль по ним не учтена в итоге.",
            )

        is_cancelled = status in {"cancelled", "not_accepted"}
        if is_cancelled:
            warning = _append_warning(
                warning,
                "Заказ отменён - расчёт прибыли не имеет смысла, показан для справки.",
            )
        elif status not in {"delivered"}:
            warning = _append_warning(
                warning,
                "Заказ ещё не доставлен/не выкуплен покупателем - это оценка на основе "
                "юнит-экономики, а не итоговый расчёт Ozon.",
            )

        customer = posting.get("customer")
        customer = customer if isinstance(customer, dict) else {}
        address = customer.get("address")
        address = address if isinstance(address, dict) else {}
        delivery_method = posting.get("delivery_method")
        delivery_method = delivery_method if isinstance(delivery_method, dict) else {}
        lift_code, lift_label, lift_price, lift_floor = _parse_prr_option(
            posting.get("prr_option")
        )
        delivery_price_transferred = _optional_float(posting.get("delivery_price"))
        # logistics_compensation is юнитка's per-item realFBS delivery
        # reimbursement - net_profit_total only nets cost + Ozon commission
        # (expense_with_ozon_commission), so this isn't factored in there;
        # summed here alongside delivery_price/lift_price into one delivery
        # total, per the user's request.
        logistics_compensation_total = (
            sum(item.logistics_compensation or 0 for item in items)
            if items and all_matched
            else None
        )
        delivery_total_transferred = (
            delivery_price_transferred + (lift_price or 0) + (logistics_compensation_total or 0)
            if delivery_price_transferred is not None
            else None
        )

        net_profit = net_profit_total if items and all_matched else None
        net_profit_with_transferred = (
            net_profit + delivery_total_transferred
            if net_profit is not None and delivery_total_transferred is not None
            else None
        )
        net_profit_with_actual = None
        # delivery_cost_actual is user-entered, merged in by the async caller
        # after this call returns (this method has no DB access) - left None
        # here on purpose; see compute_order_profit / compute_orders_profit_for_date.

        return OzonOrderLookupResponse(
            posting_number=str(posting.get("posting_number") or posting_number),
            order_number=str(posting.get("order_number") or ""),
            status=status,
            status_label=_ORDER_STATUS_LABELS.get(status, status),
            is_cancelled=is_cancelled,
            in_process_at=in_process_at,
            items=items,
            revenue_without_vat_total=revenue_without_vat_total,
            net_profit_total=net_profit,
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            warning=warning,
            computed_at=datetime.now(UTC).isoformat(),
            delivery_address=_optional_str(address.get("address_tail")) or None,
            delivery_comment=_optional_str(address.get("comment")) or None,
            delivery_method_name=_optional_str(delivery_method.get("name")) or None,
            delivery_price_transferred=delivery_price_transferred,
            lift_option_code=lift_code,
            lift_option_label=lift_label,
            lift_price=lift_price,
            lift_floor=lift_floor,
            logistics_compensation_total=logistics_compensation_total,
            delivery_total_transferred=delivery_total_transferred,
            net_profit_with_delivery_transferred=net_profit_with_transferred,
            net_profit_with_delivery_actual=net_profit_with_actual,
        )

    async def import_total_sales_report(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
        file: UploadFile,
    ) -> OzonSkuEfficiencyResponse:
        content = await file.read()
        filename = file.filename or "ozon-total-sales-report.csv"
        parsed_report = self._total_sales_report_parser.parse(filename, content)
        ad_response = await self.get_campaign_sku_efficiency_composite(
            campaign_id,
            date_from,
            date_to,
        )

        return self._economics_calculator.enrich_with_total_sales(
            ad_response,
            parsed_report.rows,
        )

    async def get_total_sales_report_from_api(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonSkuEfficiencyResponse:
        ad_response = await self.get_campaign_sku_efficiency_composite(
            campaign_id,
            date_from,
            date_to,
        )
        total_sales_rows = await self._get_total_sales_rows_from_api_cached(
            date_from,
            date_to,
        )

        enriched_response = self._economics_calculator.enrich_with_total_sales(
            ad_response,
            total_sales_rows,
        )
        return enriched_response.model_copy(
            update={
                "total_sales_report_state": "API",
                "promotion_report_state": enriched_response.promotion_report_state or "API",
                "total_sales_warning": _append_warning(
                    enriched_response.total_sales_warning,
                    "Общие продажи загружены из Seller API /v1/analytics/data.",
                ),
            }
        )

    async def get_sku_daily_profit(
        self,
        date_from: str,
        date_to: str,
        sku: str | None = None,
        offer_id: str | None = None,
        ad_spend_without_vat: float = 0,
        profit_tax_rate: float = 0.22,
    ) -> OzonSkuDailyProfitResponse:
        if sku is None and offer_id is None:
            raise ValueError("SKU or offer_id is required")

        resolved_sku = sku
        resolved_offer_id = offer_id
        seller_product: dict[str, object] = {}
        if resolved_sku is not None:
            seller_products_by_sku = await self._seller_client.get_products_by_sku(
                [resolved_sku]
            )
            seller_product = seller_products_by_sku.get(resolved_sku, {})
            resolved_offer_id = (
                resolved_offer_id
                or _optional_str(seller_product.get("offer_id"))
            )
        else:
            products_by_offer = await self._seller_client.get_products_by_offer_id(
                [resolved_offer_id or ""]
            )
            seller_product = products_by_offer.get(resolved_offer_id or "", {})
            product_sku = seller_product.get("sku") or seller_product.get("fbo_sku")
            resolved_sku = _optional_str(product_sku)

        if resolved_sku is None:
            raise ValueError("SKU was not resolved from Ozon Seller API")

        api_rows = await self._seller_client.get_analytics_sales_by_sku_day(
            date_from,
            date_to,
            resolved_sku,
        )
        daily_sales_rows = [
            parsed
            for row in api_rows
            if (parsed := _parse_daily_sales_api_row(row)) is not None
        ]
        daily_sales_rows.sort(key=lambda row: str(row["date"]))

        total_revenue_with_vat = sum(
            parse_decimal(row["ordered_amount_with_vat"])
            for row in daily_sales_rows
        )
        total_revenue_without_vat = total_revenue_with_vat / VAT_MULTIPLIER
        total_units = sum(parse_int(row["ordered_units"]) for row in daily_sales_rows)
        title = _optional_str(seller_product.get("name")) or _first_daily_title(
            daily_sales_rows
        )

        rows: list[OzonSkuDailyProfitRow] = []
        net_profit_before_ads_total = 0.0
        has_missing_unit_economy = False

        for sales_row in daily_sales_rows:
            row_date = str(sales_row["date"])
            ordered_amount_with_vat = parse_decimal(
                sales_row["ordered_amount_with_vat"]
            )
            ordered_units = parse_int(sales_row["ordered_units"])
            ordered_amount_without_vat = ordered_amount_with_vat / VAT_MULTIPLIER
            allocated_ad_spend_without_vat = (
                ad_spend_without_vat
                * ordered_amount_without_vat
                / total_revenue_without_vat
                if total_revenue_without_vat
                else 0
            )

            unit_product = self._find_unit_product(
                resolved_offer_id,
                row_date,
            ) or self._find_unit_product_by_sku(resolved_sku, row_date)
            unit_version = (
                self._unit_economy_index_service.get_workbook_version(row_date)
                if self._unit_economy_index_service is not None
                else None
            )

            average_unit_price_with_vat: float | None = None
            average_unit_price_without_vat: float | None = None
            profit_before_tax_per_unit: float | None = None
            profit_tax_per_unit: float | None = None
            net_profit_per_unit: float | None = None
            net_profit_before_ads: float | None = None
            net_profit_after_ads: float | None = None
            unit_expense = (
                unit_product.expense_with_ozon_commission
                if unit_product is not None
                else None
            )

            if ordered_units > 0:
                average_unit_price_with_vat = (
                    ordered_amount_with_vat / ordered_units
                )
                average_unit_price_without_vat = (
                    ordered_amount_without_vat / ordered_units
                )

            if (
                ordered_units > 0
                and average_unit_price_without_vat is not None
                and unit_expense is not None
            ):
                profit_before_tax_per_unit = (
                    average_unit_price_without_vat - unit_expense
                )
                profit_tax_per_unit = (
                    max(profit_before_tax_per_unit, 0) * profit_tax_rate
                )
                net_profit_per_unit = (
                    profit_before_tax_per_unit - profit_tax_per_unit
                )
                net_profit_before_ads = net_profit_per_unit * ordered_units
                net_profit_after_ads = (
                    net_profit_before_ads - allocated_ad_spend_without_vat
                )
                net_profit_before_ads_total += net_profit_before_ads
            elif ordered_units > 0 or ordered_amount_with_vat > 0:
                has_missing_unit_economy = True

            rows.append(
                OzonSkuDailyProfitRow(
                    date=row_date,
                    sku=resolved_sku,
                    offer_id=resolved_offer_id,
                    title=title,
                    ordered_amount_with_vat=ordered_amount_with_vat,
                    ordered_units=ordered_units,
                    average_unit_price_with_vat=average_unit_price_with_vat,
                    average_unit_price_without_vat=average_unit_price_without_vat,
                    unit_economy_version_valid_from=(
                        unit_version.valid_from if unit_version is not None else None
                    ),
                    unit_economy_sheet_name=(
                        unit_version.sheet_name if unit_version is not None else None
                    ),
                    unit_expense_with_ozon_commission=unit_expense,
                    profit_before_tax_per_unit=profit_before_tax_per_unit,
                    profit_tax_per_unit=profit_tax_per_unit,
                    net_profit_per_unit=net_profit_per_unit,
                    net_profit_before_ads=net_profit_before_ads,
                    allocated_ad_spend_without_vat=allocated_ad_spend_without_vat,
                    net_profit_after_ads=net_profit_after_ads,
                    matched_unit_economy=unit_expense is not None,
                )
            )

        net_profit_before_ads_result: float | None = (
            None if has_missing_unit_economy else net_profit_before_ads_total
        )
        net_profit_after_ads_result: float | None = (
            None
            if net_profit_before_ads_result is None
            else net_profit_before_ads_result - ad_spend_without_vat
        )

        return OzonSkuDailyProfitResponse(
            date_from=date_from,
            date_to=date_to,
            sku=resolved_sku,
            offer_id=resolved_offer_id,
            title=title,
            vat_multiplier=VAT_MULTIPLIER,
            profit_tax_rate=profit_tax_rate,
            rows=rows,
            total=OzonSkuDailyProfitTotal(
                ordered_amount_with_vat=total_revenue_with_vat,
                ordered_amount_without_vat=total_revenue_without_vat,
                ordered_units=total_units,
                net_profit_before_ads=net_profit_before_ads_result,
                ad_spend_without_vat=ad_spend_without_vat,
                net_profit_after_ads=net_profit_after_ads_result,
                matched_unit_economy=not has_missing_unit_economy,
            ),
        )

    async def get_sales_analytics(
        self,
        date_from: str,
        date_to: str,
    ) -> OzonSalesAnalyticsResponse:
        unit_version_selection = self._select_unit_economy_version(date_from, date_to)
        effective_date = (
            unit_version_selection.version.valid_from
            if unit_version_selection is not None
            else date_to
        )
        total_sales_rows = await self._get_total_sales_rows_from_api_cached(
            date_from,
            date_to,
        )
        ad_response = await self.get_campaign_sku_efficiency_composite(
            "ALL",
            date_from,
            date_to,
        )
        ad_rows_by_key = _index_ad_rows(ad_response.rows)
        sales_rows_by_key = _index_total_sales_rows(total_sales_rows)
        result_rows: list[OzonSalesAnalyticsRow] = []
        used_ad_keys: set[str] = set()

        for total_sales_row in total_sales_rows:
            match_keys = _row_match_keys(total_sales_row.offer_id, total_sales_row.sku)
            ad_row = _first_matching_row(ad_rows_by_key, match_keys)
            used_ad_keys.update(match_keys)

            result_rows.append(
                self._build_sales_analytics_row(
                    total_sales_row=total_sales_row,
                    ad_row=ad_row,
                    effective_date=effective_date,
                )
            )

        for match_key, ad_row in ad_rows_by_key.items():
            if match_key in used_ad_keys or match_key in sales_rows_by_key:
                continue

            result_rows.append(
                self._build_sales_analytics_row(
                    total_sales_row=None,
                    ad_row=ad_row,
                    effective_date=effective_date,
                )
            )

        result_rows.sort(
            key=lambda row: (
                row.net_profit is None,
                -(row.net_profit or 0),
            )
        )
        total = self._build_sales_analytics_total(result_rows)

        return OzonSalesAnalyticsResponse(
            date_from=date_from,
            date_to=date_to,
            rows=result_rows,
            total=total,
            unit_economy_version=(
                unit_version_selection.version.version_id
                if unit_version_selection is not None
                else None
            ),
            unit_economy_version_valid_from=(
                unit_version_selection.version.valid_from
                if unit_version_selection is not None
                else None
            ),
            unit_economy_workbook_path=(
                unit_version_selection.version.path
                if unit_version_selection is not None
                else None
            ),
            warning=_append_warning(
                unit_version_selection.warning
                if unit_version_selection is not None
                else None,
                _append_warning(
                    ad_response.unit_economy_warning,
                    ad_response.total_sales_warning,
                ),
            ),
        )

    async def get_sales_forecast(
        self,
        payload: OzonSalesForecastRequest,
    ) -> OzonSalesForecastResponse:
        sales_response = await self.get_sales_analytics(payload.date_from, payload.date_to)
        sales_rows_by_key = {
            key: row
            for row in sales_response.rows
            for key in _row_match_keys(row.offer_id, row.sku)
        }

        if payload.items:
            rows: list[OzonSalesForecastRow] = []
            for item in payload.items:
                source_row = _first_matching_sales_row(
                    sales_rows_by_key,
                    _row_match_keys(item.offer_id, item.sku),
                )
                if source_row is None:
                    continue

                rows.append(
                    self._build_sales_forecast_row(
                        row=source_row,
                        target_ad_spend_with_vat=item.target_ad_spend_with_vat,
                        target_drr_percent=item.target_drr_percent,
                    )
                )

            if not rows:
                raise ValueError("Selected forecast items were not found in Ozon sales data")

            forecast_note = (
                "Прогноз: рассчитаны только выбранные товары. "
                "Для каждого артикула использованы индивидуальные бюджет и DRR."
            )
        else:
            if (
                payload.target_ad_spend_with_vat is None
                or payload.target_drr_percent is None
            ):
                raise ValueError(
                    "Forecast requires either selected items or common budget and DRR"
                )

            target_ad_spend_without_vat = payload.target_ad_spend_with_vat / VAT_MULTIPLIER
            allocation_rows = [
                row for row in sales_response.rows if row.ad_spend_without_vat > 0
            ]
            allocation_uses_ad_spend = True
            if not allocation_rows:
                allocation_rows = [
                    row for row in sales_response.rows if row.total_revenue_without_vat > 0
                ]
                allocation_uses_ad_spend = False

            allocation_total = (
                sum(row.ad_spend_without_vat for row in allocation_rows)
                if allocation_uses_ad_spend
                else sum(row.total_revenue_without_vat for row in allocation_rows)
            )
            allocation_keys = {_sales_row_key(row) for row in allocation_rows}
            rows = [
                self._build_sales_forecast_row(
                    row=row,
                    target_ad_spend_with_vat=(
                        target_ad_spend_without_vat
                        * (
                            _forecast_allocation_share(
                                row,
                                allocation_total,
                                allocation_uses_ad_spend,
                            )
                            if _sales_row_key(row) in allocation_keys
                            else 0
                        )
                        * VAT_MULTIPLIER
                    ),
                    target_drr_percent=payload.target_drr_percent,
                )
                for row in sales_response.rows
            ]
            forecast_note = (
                "Прогноз: органика сохранена как факт выбранного периода. "
                "Рекламный бюджет распределён по исторической доле рекламного расхода."
                if allocation_uses_ad_spend
                else "Прогноз: органика сохранена как факт выбранного периода. "
                "В периоде не найден рекламный расход, поэтому бюджет распределён по доле продаж."
            )

        target_ad_spend_with_vat = sum(row.target_ad_spend_with_vat for row in rows)
        target_ad_spend_without_vat = sum(
            row.forecast_ad_spend_without_vat for row in rows
        )
        target_ad_revenue_without_vat = sum(
            row.forecast_ad_revenue_without_vat for row in rows
        )
        target_drr_percent = (
            target_ad_spend_without_vat / target_ad_revenue_without_vat * 100
            if target_ad_revenue_without_vat
            else 0
        )
        rows = _with_forecast_budget_shares(rows)
        rows.sort(
            key=lambda row: (
                row.forecast_total_net_profit is None,
                -(row.forecast_total_net_profit or 0),
            )
        )
        total = self._build_sales_forecast_total(rows)

        return OzonSalesForecastResponse(
            date_from=payload.date_from,
            date_to=payload.date_to,
            target_ad_spend_with_vat=target_ad_spend_with_vat,
            target_ad_spend_without_vat=target_ad_spend_without_vat,
            target_drr_percent=target_drr_percent,
            target_ad_revenue_without_vat=target_ad_revenue_without_vat,
            rows=rows,
            total=total,
            unit_economy_version=sales_response.unit_economy_version,
            unit_economy_version_valid_from=sales_response.unit_economy_version_valid_from,
            unit_economy_workbook_path=sales_response.unit_economy_workbook_path,
            warning=_append_warning(sales_response.warning, forecast_note),
        )

    async def create_sku_efficiency_report(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonStatisticsReportStatus:
        saved_status = await self._find_saved_report_status(campaign_id, date_from, date_to)
        if saved_status is not None:
            if saved_status.state in {"CREATED", "NOT_STARTED", "IN_PROGRESS", "OK"}:
                return saved_status

        try:
            report_uuid = await self._create_statistics_report_with_limit_wait(
                await self._resolve_campaign_ids(campaign_id, date_from, date_to),
                date_from,
                date_to,
            )
        except OzonApiError as exc:
            if _is_active_report_limit_error(str(exc)):
                active_status = await self._find_active_report_status()
                if active_status is not None:
                    return active_status

            raise

        status = await self.get_statistics_report_status(campaign_id, report_uuid)
        await self._save_report_status(status, date_from, date_to)
        return status

    async def get_statistics_report_status(
        self,
        campaign_id: str,
        report_uuid: str,
    ) -> OzonStatisticsReportStatus:
        data = await self._performance_client.request(
            "GET",
            f"/api/client/statistics/{report_uuid}",
        )
        state = data.get("state")
        if not isinstance(state, str):
            raise ValueError("Ozon statistics report state is missing")

        status = OzonStatisticsReportStatus(
            campaign_id=campaign_id,
            report_uuid=report_uuid,
            state=state,
            link=_optional_str(data.get("link")),
            created_at=_optional_str(data.get("createdAt")),
            updated_at=_optional_str(data.get("updatedAt")),
        )
        if self._ozon_ad_report_repository is not None:
            await self._ozon_ad_report_repository.update_report_status(status)

        return status

    async def _get_or_build_sku_efficiency_segment(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonSkuEfficiencyResponse:
        saved_status = await self._find_saved_report_status(
            campaign_id,
            date_from,
            date_to,
        )
        if saved_status is None:
            return await self.get_campaign_sku_efficiency(
                campaign_id,
                date_from,
                date_to,
            )

        saved_result = await self.get_saved_sku_efficiency_report_result(
            saved_status.report_uuid
        )
        if (
            saved_result is not None
            and _result_matches_unit_version(
                saved_result,
                self._select_unit_economy_version(date_from, date_to),
            )
            and _result_has_campaign_source_metadata(saved_result)
            and _result_uses_direct_sku_profit(saved_result)
            and _result_uses_model_attribution_rule(saved_result)
            and _result_uses_non_negative_profit_tax(saved_result)
            and _result_has_product_identity(saved_result)
            and _result_has_price_diagnostics(saved_result)
            and _result_has_promotion_diagnostics(saved_result)
            and self._result_has_current_unit_costs(saved_result, date_to)
        ):
            return saved_result

        if saved_status.state in {"CREATED", "NOT_STARTED", "IN_PROGRESS"}:
            await self._wait_statistics_report(saved_status.report_uuid)

        return await self.get_campaign_sku_efficiency(
            campaign_id,
            date_from,
            date_to,
            report_uuid=saved_status.report_uuid,
        )

    async def _get_or_build_campaign_ids_sku_efficiency_segment(
        self,
        campaign_id: str,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
    ) -> OzonSkuEfficiencyResponse:
        saved_status = await self._find_saved_report_status(
            campaign_id,
            date_from,
            date_to,
        )
        if saved_status is None:
            if len(campaign_ids) > 10:
                return await self._get_campaign_batch_sku_efficiency(
                    campaign_id,
                    campaign_ids,
                    date_from,
                    date_to,
                )

            return await self._get_campaign_ids_sku_efficiency(
                campaign_id,
                campaign_ids,
                date_from,
                date_to,
            )

        saved_result = await self.get_saved_sku_efficiency_report_result(
            saved_status.report_uuid
        )
        if (
            saved_result is not None
            and _result_matches_unit_version(
                saved_result,
                self._select_unit_economy_version(date_from, date_to),
            )
            and _result_has_campaign_source_metadata(saved_result)
            and _result_uses_direct_sku_profit(saved_result)
            and _result_uses_model_attribution_rule(saved_result)
            and _result_uses_non_negative_profit_tax(saved_result)
            and _result_has_product_identity(saved_result)
            and _result_has_price_diagnostics(saved_result)
            and _result_has_promotion_diagnostics(saved_result)
            and self._result_has_current_unit_costs(saved_result, date_to)
        ):
            return saved_result

        if saved_status.state in {"CREATED", "NOT_STARTED", "IN_PROGRESS"}:
            await self._wait_statistics_report(saved_status.report_uuid)

        if len(campaign_ids) > 10:
            return await self._get_campaign_batch_sku_efficiency(
                campaign_id,
                campaign_ids,
                date_from,
                date_to,
            )

        return await self._get_campaign_ids_sku_efficiency(
            campaign_id,
            campaign_ids,
            date_from,
            date_to,
            report_uuid=saved_status.report_uuid,
        )

    async def get_saved_sku_efficiency_report_result(
        self,
        report_uuid: str,
    ) -> OzonSkuEfficiencyResponse | None:
        if self._ozon_ad_report_repository is None:
            return None

        response = await self._ozon_ad_report_repository.get_efficiency_result(report_uuid)
        if response is None:
            return None

        return self._economics_calculator.normalize_profit_tax(response)

    async def _get_saved_raw_report_csv(self, report_uuid: str) -> str | None:
        if self._ozon_ad_report_repository is None:
            return None

        return await self._ozon_ad_report_repository.get_raw_report_csv(report_uuid)

    async def list_saved_sku_efficiency_reports(
        self,
        limit: int = 50,
    ) -> list[OzonStoredReportSummary]:
        if self._ozon_ad_report_repository is None:
            return []

        return await self._ozon_ad_report_repository.list_reports(limit)

    async def _save_report_status(
        self,
        status: OzonStatisticsReportStatus,
        date_from: str,
        date_to: str,
    ) -> None:
        if self._ozon_ad_report_repository is None:
            return

        await self._ozon_ad_report_repository.upsert_report_status(
            status=status,
            date_from=date_from,
            date_to=date_to,
            unit_economy_version=(
                selection.version
                if (selection := self._select_unit_economy_version(date_from, date_to))
                is not None
                else None
            ),
        )

    async def _find_saved_report_status(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonStatisticsReportStatus | None:
        if self._ozon_ad_report_repository is None:
            return None

        return await self._ozon_ad_report_repository.find_report_status(
            campaign_id,
            date_from,
            date_to,
        )

    async def _find_active_report_status(self) -> OzonStatisticsReportStatus | None:
        if self._ozon_ad_report_repository is None:
            return None

        return await self._ozon_ad_report_repository.find_active_report_status()

    def _get_unit_economy_version(self) -> UnitEconomyWorkbookVersion | None:
        if self._unit_economy_index_service is None:
            return None

        return self._unit_economy_index_service.get_workbook_version()

    def _select_unit_economy_version(
        self,
        date_from: str,
        date_to: str,
    ) -> UnitEconomyVersionSelection | None:
        if self._unit_economy_index_service is None:
            return None

        return self._unit_economy_index_service.select_version_for_period(
            date_from,
            date_to,
        )

    async def _get_period_ad_spend_without_vat(
        self,
        date_from: str,
        date_to: str,
    ) -> tuple[float | None, str | None]:
        try:
            expense_csv = await self._performance_client.get_campaign_expense_csv(
                date_from,
                date_to,
            )
            # Ozon publishes per-campaign expense with a lag - for the current,
            # still-in-progress day the CSV often comes back with only the
            # header row (no campaigns yet), even though Ozon's own ad
            # dashboard already shows accruing spend from a different,
            # near-real-time internal source. Treat "no rows" as "not
            # published yet" rather than a real zero, so we don't understate
            # net profit by silently omitting today's ad spend.
            expense_lines = [line for line in expense_csv.splitlines() if line.strip()]
            if len(expense_lines) > 1:
                ad_spend_with_vat = _sum_campaign_expense_csv(expense_csv)
                if ad_spend_with_vat < 0:
                    # A real total shouldn't come out negative - this means
                    # Ozon's export contained a refund/credit line for the
                    # period that outweighs actual spend. Surfacing "-N ₽"
                    # as ad spend is misleading, so flag it instead of
                    # showing a number that looks like a bug.
                    return None, (
                        f"Расход на рекламу за {date_from}-{date_to} получился "
                        f"отрицательным ({ad_spend_with_vat:.2f} ₽ с НДС) - похоже, "
                        "в выгрузке Ozon есть строка возврата/компенсации. "
                        "Нужна ручная проверка в личном кабинете."
                    )
                return float(without_vat(Decimal(str(ad_spend_with_vat)))), None
        except Exception as exc:  # noqa: BLE001 - background job, must not crash
            return None, f"Не удалось получить расход на рекламу за {date_from}-{date_to}: {exc}"

        # /statistics/expense has nothing yet for this period - Ozon usually
        # publishes it only the next day. If the period is exactly today,
        # fall back to the per-campaign promotion analytics report, which
        # does have same-day numbers (confirmed live: /expense returns just
        # a header for today while this report already shows real spend).
        # It's slower (~1 report generation per 10 running campaigns,
        # sequential - a few minutes for ~30 campaigns), so it's only used
        # from background schedulers, never from a live request handler.
        today_iso = datetime.now(MOSCOW_TZ).date().isoformat()
        if date_from == date_to == today_iso:
            try:
                ad_spend_with_vat = await self._get_today_ad_spend_via_promotion_report(
                    today_iso
                )
                if ad_spend_with_vat < 0:
                    return None, (
                        f"Расход на рекламу за {today_iso} получился отрицательным "
                        f"({ad_spend_with_vat:.2f} ₽ с НДС) - похоже, в отчёте Ozon "
                        "есть строка возврата/компенсации. Нужна ручная проверка "
                        "в личном кабинете."
                    )
                return float(without_vat(Decimal(str(ad_spend_with_vat)))), None
            except Exception as exc:  # noqa: BLE001 - background job, must not crash
                return None, (
                    f"Не удалось получить расход на рекламу за {today_iso} ни из "
                    f"экспорта, ни из аналитики продвижения: {exc}"
                )

        period = date_from if date_from == date_to else f"{date_from} - {date_to}"
        return None, (
            f"Ozon ещё не опубликовал расход на рекламу за {period} - "
            "обычно данные появляются на следующий день. Чистая прибыль "
            "пока не учитывает рекламу за этот период."
        )

    async def _get_today_ad_spend_via_promotion_report(self, date_str: str) -> float:
        async with self._promotion_report_lock:
            campaigns = await self.get_campaigns()
            # This report format is only valid for SKU (product promotion)
            # campaigns. Other types - confirmed live with "Оплата за заказ:
            # выбранные товары" (advObjectType SEARCH_PROMO) - make Ozon
            # reject the WHOLE batch they're mixed into with "generation of
            # this type of report is forbidden for the transferred list of
            # campaigns". That campaign also shows zero spend in the
            # per-campaign expense export on days that have data, so
            # excluding non-SKU types here doesn't meaningfully undercount.
            running_ids = [
                campaign.id for campaign in campaigns
                if campaign.state == "CAMPAIGN_STATE_RUNNING"
                and campaign.raw.get("advObjectType") == "SKU"
            ]

            total_with_vat = 0.0
            for chunk in _chunked(running_ids, 10):
                report_uuid = await self._create_statistics_report_with_limit_wait(
                    chunk,
                    date_str,
                    date_str,
                )
                await self._wait_statistics_report(report_uuid)
                csv_text = await self._performance_client.download_report(report_uuid)
                total_with_vat += _sum_promotion_report_spend(csv_text)

            return total_with_vat

    def _find_unit_product(
        self,
        offer_id: str | None,
        effective_date: str | None,
    ) -> UnitEconomyProduct | None:
        if self._unit_economy_index_service is None or offer_id is None:
            return None

        if effective_date is None:
            return self._unit_economy_index_service.find_by_offer_id(offer_id)

        return self._unit_economy_index_service.find_by_offer_id_for_date(
            offer_id,
            effective_date,
        )

    def _find_unit_product_by_sku(
        self,
        sku: str | None,
        effective_date: str | None,
    ) -> UnitEconomyProduct | None:
        if self._unit_economy_index_service is None or sku is None:
            return None

        if effective_date is None:
            return self._unit_economy_index_service.find_by_sku(sku)

        return self._unit_economy_index_service.find_by_sku_for_date(
            sku,
            effective_date,
        )

    def _result_has_current_unit_costs(
        self,
        response: OzonSkuEfficiencyResponse,
        effective_date: str | None,
    ) -> bool:
        for row in response.rows:
            if row.sku == "TOTAL":
                continue

            unit_product = self._find_unit_product(
                row.offer_id,
                effective_date,
            ) or self._find_unit_product_by_sku(row.sku, effective_date)
            if (
                unit_product is None
                or unit_product.expense_with_ozon_commission is None
            ):
                continue

            if row.unit_expense_with_ozon_commission is None:
                return False

            if not _money_values_close(
                row.unit_expense_with_ozon_commission,
                unit_product.expense_with_ozon_commission,
            ):
                return False

        return True

    async def _build_import_products_by_sku(
        self,
        campaign_id: str,
        report_rows: list[dict[str, str]],
        effective_date: str | None,
        date_from: str | None,
        date_to: str | None,
    ) -> dict[str, OzonCampaignProduct]:
        return await self._build_products_by_sku_from_report_rows(
            campaign_id,
            report_rows,
            effective_date,
            date_from,
            date_to,
        )

    async def _build_products_by_sku_from_report_rows(
        self,
        campaign_id: str,
        report_rows: list[dict[str, str]],
        effective_date: str | None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict[str, OzonCampaignProduct]:
        products_by_sku: dict[str, OzonCampaignProduct] = {}
        for raw_row in report_rows:
            sku = raw_row.get("sku", "")
            if not sku.isdigit():
                continue

            offer_id = _optional_str(raw_row.get("Артикул"))
            unit_product = self._find_unit_product(
                offer_id,
                effective_date,
            ) or self._find_unit_product_by_sku(sku, effective_date)
            products_by_sku[sku] = OzonCampaignProduct(
                campaign_id=campaign_id,
                performance_object_id=sku,
                sku=sku,
                offer_id=offer_id or (unit_product.offer_id if unit_product is not None else None),
                title=_optional_str(raw_row.get("Название товара")),
                unit_economy_match_key=offer_id or sku,
                unit_economy=unit_product,
                raw=raw_row,
            )

        sku_values = list(products_by_sku)
        if not sku_values:
            return products_by_sku

        try:
            seller_products_by_sku = await self._seller_client.get_products_by_sku(sku_values)
        except OzonApiError:
            return products_by_sku

        products_by_sku = await self._attach_promotions_to_products(
            products_by_sku,
            seller_products_by_sku,
            date_from,
            date_to,
        )

        for sku in sku_values:
            seller_product = seller_products_by_sku.get(sku, {})
            offer_id = _optional_str(seller_product.get("offer_id"))
            unit_product = self._find_unit_product(
                offer_id,
                effective_date,
            ) or self._find_unit_product_by_sku(sku, effective_date)
            existing_product = products_by_sku[sku]
            products_by_sku[sku] = existing_product.model_copy(
                update={
                    "seller_product_id": _optional_str(seller_product.get("id")),
                    "offer_id": offer_id or existing_product.offer_id,
                    "title": _optional_str(seller_product.get("name")) or existing_product.title,
                    "price": _optional_str(seller_product.get("price")),
                    "old_price": _optional_str(seller_product.get("old_price")),
                    "primary_image": _first_string(seller_product.get("primary_image")),
                    "vat": _optional_str(seller_product.get("vat")),
                    "unit_economy_match_key": offer_id or existing_product.unit_economy_match_key,
                    "unit_economy": unit_product or existing_product.unit_economy,
                    "promotions": existing_product.promotions,
                }
            )

        return products_by_sku

    async def _attach_promotions_to_products(
        self,
        products_by_sku: dict[str, OzonCampaignProduct],
        seller_products_by_sku: dict[str, dict[str, object]],
        date_from: str | None,
        date_to: str | None,
    ) -> dict[str, OzonCampaignProduct]:
        if date_from is None or date_to is None:
            return products_by_sku

        product_ids_by_sku = {
            sku: product_id
            for sku, seller_product in seller_products_by_sku.items()
            if (product_id := _optional_str(seller_product.get("id"))) is not None
        }
        if not product_ids_by_sku:
            return products_by_sku

        try:
            promotions_by_product_id = await self._seller_client.get_promotions_by_product_id(
                date_from,
                date_to,
                list(product_ids_by_sku.values()),
            )
        except OzonApiError:
            return products_by_sku

        return {
            sku: product.model_copy(
                update={
                    "promotions": promotions_by_product_id.get(
                        product_ids_by_sku.get(sku, ""),
                        [],
                    )
                }
            )
            for sku, product in products_by_sku.items()
        }

    async def _build_total_sales_rows_from_api(
        self,
        api_rows: list[dict[str, object]],
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[OzonTotalSalesReportRow]:
        skus = [_analytics_row_sku(row) for row in api_rows]
        sku_values = [sku for sku in skus if sku is not None]
        seller_products_by_sku = await self._seller_client.get_products_by_sku(sku_values)
        product_ids = [
            product_id
            for seller_product in seller_products_by_sku.values()
            if (product_id := _optional_str(seller_product.get("id"))) is not None
        ]
        promotions_by_product_id: dict[str, list[object]] = {}
        if date_from is not None and date_to is not None and product_ids:
            try:
                promotions_by_product_id = await self._seller_client.get_promotions_by_product_id(
                    date_from,
                    date_to,
                    product_ids,
                )
            except OzonApiError:
                promotions_by_product_id = {}
        rows: list[OzonTotalSalesReportRow] = []

        for row in api_rows:
            sku = _analytics_row_sku(row)
            if sku is None:
                continue

            metrics = row.get("metrics")
            if not isinstance(metrics, list) or len(metrics) < 2:
                continue

            seller_product = seller_products_by_sku.get(sku, {})
            seller_product_id = _optional_str(seller_product.get("id"))
            rows.append(
                OzonTotalSalesReportRow(
                    offer_id=_optional_str(seller_product.get("offer_id")),
                    sku=sku,
                    title=_analytics_row_name(row)
                    or _optional_str(seller_product.get("name")),
                    ordered_amount_with_vat=parse_decimal(metrics[1]),
                    orders=parse_int(metrics[0]),
                    seller_product_id=seller_product_id,
                    promotions=promotions_by_product_id.get(seller_product_id or "", []),
                )
            )

        return rows

    async def _get_total_sales_rows_from_api_cached(
        self,
        date_from: str,
        date_to: str,
    ) -> list[OzonTotalSalesReportRow]:
        can_use_cache = (
            self._ozon_ad_report_repository is not None
            and date_to < date.today().isoformat()
        )
        if can_use_cache:
            cached_rows = await self._ozon_ad_report_repository.get_total_sales_rows_cache(
                date_from,
                date_to,
            )
            if cached_rows is not None:
                return cached_rows

        api_rows = await self._seller_client.get_analytics_sales_by_sku(
            date_from,
            date_to,
        )
        total_sales_rows = await self._build_total_sales_rows_from_api(
            api_rows,
            date_from,
            date_to,
        )
        if can_use_cache:
            await self._ozon_ad_report_repository.save_total_sales_rows_cache(
                date_from,
                date_to,
                total_sales_rows,
            )

        return total_sales_rows

    def _build_sales_analytics_row(
        self,
        total_sales_row: OzonTotalSalesReportRow | None,
        ad_row: OzonSkuEfficiencyRow | None,
        effective_date: str,
    ) -> OzonSalesAnalyticsRow:
        sku = (
            total_sales_row.sku
            if total_sales_row is not None and total_sales_row.sku is not None
            else ad_row.sku
            if ad_row is not None
            else ""
        )
        offer_id = (
            total_sales_row.offer_id
            if total_sales_row is not None and total_sales_row.offer_id is not None
            else ad_row.offer_id
            if ad_row is not None
            else None
        )
        title = (
            total_sales_row.title
            if total_sales_row is not None and total_sales_row.title is not None
            else ad_row.title
            if ad_row is not None
            else None
        )
        unit_product = self._find_unit_product(
            offer_id,
            effective_date,
        ) or self._find_unit_product_by_sku(sku, effective_date)
        unit_expense = (
            unit_product.expense_with_ozon_commission
            if unit_product is not None
            else None
        )
        unit_price_without_vat = (
            unit_product.price_without_vat
            if unit_product is not None
            else None
        )
        promotions = []
        if total_sales_row is not None and total_sales_row.promotions:
            promotions.extend(total_sales_row.promotions)
        if ad_row is not None and ad_row.promotion_matched:
            promotions.extend(_row_promotions_from_ad_row(ad_row))
        promotion_summary = _summarize_row_promotions(promotions)
        total_orders = total_sales_row.orders if total_sales_row is not None else 0
        total_revenue_with_vat = (
            total_sales_row.ordered_amount_with_vat
            if total_sales_row is not None
            else 0
        )
        total_revenue_without_vat = total_revenue_with_vat / VAT_MULTIPLIER
        average_total_order_revenue_without_vat = (
            total_revenue_without_vat / total_orders if total_orders else None
        )
        ad_orders = ad_row.sku_orders if ad_row is not None else 0
        ad_revenue_without_vat = (
            ad_row.sku_revenue_without_vat if ad_row is not None else 0
        )
        ad_spend_without_vat = ad_row.ad_spend_without_vat if ad_row is not None else 0
        organic_orders = max(total_orders - ad_orders, 0)
        organic_revenue_without_vat = max(
            total_revenue_without_vat - ad_revenue_without_vat,
            0,
        )
        average_organic_order_revenue_without_vat = (
            organic_revenue_without_vat / organic_orders if organic_orders else None
        )
        profit_before_tax = (
            total_revenue_without_vat - ad_spend_without_vat - unit_expense * total_orders
            if unit_expense is not None
            else None
        )
        net_profit = self._economics_calculator.calculate_net_profit(profit_before_tax)

        return OzonSalesAnalyticsRow(
            sku=sku,
            offer_id=offer_id,
            title=title,
            total_orders=total_orders,
            total_revenue_with_vat=total_revenue_with_vat,
            total_revenue_without_vat=total_revenue_without_vat,
            ad_orders=ad_orders,
            ad_revenue_without_vat=ad_revenue_without_vat,
            model_orders=ad_row.model_orders if ad_row is not None else 0,
            model_revenue_without_vat=(
                ad_row.model_revenue_without_vat if ad_row is not None else 0
            ),
            ad_spend_without_vat=ad_spend_without_vat,
            organic_orders=organic_orders,
            organic_revenue_without_vat=organic_revenue_without_vat,
            unit_expense_with_ozon_commission=unit_expense,
            profit_before_tax=profit_before_tax,
            net_profit=net_profit,
            drr_percent=(
                ad_spend_without_vat / total_revenue_without_vat * 100
                if total_revenue_without_vat
                else None
            ),
            unit_economy_price_without_vat=unit_price_without_vat,
            average_total_order_revenue_without_vat=average_total_order_revenue_without_vat,
            total_price_discount_percent=_price_discount_percent(
                average_total_order_revenue_without_vat,
                unit_price_without_vat,
            ),
            average_organic_order_revenue_without_vat=(
                average_organic_order_revenue_without_vat
            ),
            organic_price_discount_percent=_price_discount_percent(
                average_organic_order_revenue_without_vat,
                unit_price_without_vat,
            ),
            promotion_matched=promotion_summary["matched"],
            promotion_count=promotion_summary["count"],
            promotion_action_ids=promotion_summary["action_ids"],
            promotion_titles=promotion_summary["titles"],
            promotion_price_with_vat=promotion_summary["price_with_vat"],
            promotion_discount_percent=promotion_summary["discount_percent"],
            matched_unit_economy=unit_expense is not None,
            has_ad_spend=ad_spend_without_vat > 0,
            has_sales=total_orders > 0 or total_revenue_with_vat > 0,
        )

    def _build_sales_analytics_total(
        self,
        rows: list[OzonSalesAnalyticsRow],
    ) -> OzonSalesAnalyticsRow:
        total_revenue_without_vat = sum(row.total_revenue_without_vat for row in rows)
        ad_spend_without_vat = sum(row.ad_spend_without_vat for row in rows)
        profit_values = [
            row.profit_before_tax
            for row in rows
            if row.profit_before_tax is not None
        ]
        profit_before_tax = sum(profit_values) if profit_values else None
        promotion_summary = _merge_sales_row_promotion_summary(rows)

        return OzonSalesAnalyticsRow(
            sku="TOTAL",
            offer_id=None,
            title="Итого",
            total_orders=sum(row.total_orders for row in rows),
            total_revenue_with_vat=sum(row.total_revenue_with_vat for row in rows),
            total_revenue_without_vat=total_revenue_without_vat,
            ad_orders=sum(row.ad_orders for row in rows),
            ad_revenue_without_vat=sum(row.ad_revenue_without_vat for row in rows),
            model_orders=sum(row.model_orders for row in rows),
            model_revenue_without_vat=sum(row.model_revenue_without_vat for row in rows),
            ad_spend_without_vat=ad_spend_without_vat,
            organic_orders=sum(row.organic_orders for row in rows),
            organic_revenue_without_vat=sum(row.organic_revenue_without_vat for row in rows),
            unit_expense_with_ozon_commission=None,
            profit_before_tax=profit_before_tax,
            net_profit=self._economics_calculator.calculate_net_profit(profit_before_tax),
            drr_percent=(
                ad_spend_without_vat / total_revenue_without_vat * 100
                if total_revenue_without_vat
                else None
            ),
            unit_economy_price_without_vat=None,
            average_total_order_revenue_without_vat=(
                total_revenue_without_vat / sum(row.total_orders for row in rows)
                if sum(row.total_orders for row in rows)
                else None
            ),
            total_price_discount_percent=None,
            average_organic_order_revenue_without_vat=(
                sum(row.organic_revenue_without_vat for row in rows)
                / sum(row.organic_orders for row in rows)
                if sum(row.organic_orders for row in rows)
                else None
            ),
            organic_price_discount_percent=None,
            promotion_matched=promotion_summary["matched"],
            promotion_count=promotion_summary["count"],
            promotion_action_ids=promotion_summary["action_ids"],
            promotion_titles=promotion_summary["titles"],
            promotion_price_with_vat=promotion_summary["price_with_vat"],
            promotion_discount_percent=promotion_summary["discount_percent"],
            matched_unit_economy=len(profit_values) == len(rows),
            has_ad_spend=any(row.has_ad_spend for row in rows),
            has_sales=any(row.has_sales for row in rows),
        )

    def _build_sales_forecast_row(
        self,
        row: OzonSalesAnalyticsRow,
        target_ad_spend_with_vat: float,
        target_drr_percent: float,
    ) -> OzonSalesForecastRow:
        forecast_ad_spend_without_vat = target_ad_spend_with_vat / VAT_MULTIPLIER
        target_drr_decimal = target_drr_percent / 100
        forecast_ad_revenue_without_vat = (
            forecast_ad_spend_without_vat / target_drr_decimal
            if forecast_ad_spend_without_vat > 0
            else 0
        )
        average_revenue_without_vat = _forecast_average_order_revenue_without_vat(row)
        forecast_ad_orders = (
            forecast_ad_revenue_without_vat / average_revenue_without_vat
            if average_revenue_without_vat is not None
            else 0
        )
        forecast_organic_orders = row.organic_orders
        forecast_organic_revenue_without_vat = row.organic_revenue_without_vat
        forecast_total_orders = forecast_ad_orders + forecast_organic_orders
        forecast_total_revenue_without_vat = (
            forecast_ad_revenue_without_vat + forecast_organic_revenue_without_vat
        )
        unit_expense = row.unit_expense_with_ozon_commission
        forecast_ad_profit_before_tax = (
            forecast_ad_revenue_without_vat
            - forecast_ad_spend_without_vat
            - unit_expense * forecast_ad_orders
            if unit_expense is not None
            else None
        )
        forecast_organic_profit_before_tax = (
            forecast_organic_revenue_without_vat
            - unit_expense * forecast_organic_orders
            if unit_expense is not None
            else None
        )
        forecast_total_profit_before_tax = (
            forecast_ad_profit_before_tax + forecast_organic_profit_before_tax
            if forecast_ad_profit_before_tax is not None
            and forecast_organic_profit_before_tax is not None
            else None
        )

        return OzonSalesForecastRow(
            sku=row.sku,
            offer_id=row.offer_id,
            title=row.title,
            target_ad_spend_with_vat=target_ad_spend_with_vat,
            target_drr_percent=target_drr_percent,
            ad_spend_share_percent=0,
            fact_total_orders=row.total_orders,
            fact_total_revenue_without_vat=row.total_revenue_without_vat,
            fact_ad_spend_without_vat=row.ad_spend_without_vat,
            fact_ad_revenue_without_vat=row.ad_revenue_without_vat,
            fact_organic_orders=row.organic_orders,
            fact_organic_revenue_without_vat=row.organic_revenue_without_vat,
            forecast_ad_spend_without_vat=forecast_ad_spend_without_vat,
            forecast_ad_revenue_without_vat=forecast_ad_revenue_without_vat,
            forecast_ad_orders=forecast_ad_orders,
            forecast_organic_orders=forecast_organic_orders,
            forecast_organic_revenue_without_vat=forecast_organic_revenue_without_vat,
            forecast_total_orders=forecast_total_orders,
            forecast_total_revenue_without_vat=forecast_total_revenue_without_vat,
            unit_expense_with_ozon_commission=unit_expense,
            forecast_ad_profit_before_tax=forecast_ad_profit_before_tax,
            forecast_ad_net_profit=self._economics_calculator.calculate_net_profit(
                forecast_ad_profit_before_tax
            ),
            forecast_organic_profit_before_tax=forecast_organic_profit_before_tax,
            forecast_organic_net_profit=self._economics_calculator.calculate_net_profit(
                forecast_organic_profit_before_tax
            ),
            forecast_total_profit_before_tax=forecast_total_profit_before_tax,
            forecast_total_net_profit=self._economics_calculator.calculate_net_profit(
                forecast_total_profit_before_tax
            ),
            matched_unit_economy=unit_expense is not None,
        )

    def _build_sales_forecast_total(
        self,
        rows: list[OzonSalesForecastRow],
    ) -> OzonSalesForecastRow:
        ad_profit_values = [
            row.forecast_ad_profit_before_tax
            for row in rows
            if row.forecast_ad_profit_before_tax is not None
        ]
        organic_profit_values = [
            row.forecast_organic_profit_before_tax
            for row in rows
            if row.forecast_organic_profit_before_tax is not None
        ]
        total_profit_values = [
            row.forecast_total_profit_before_tax
            for row in rows
            if row.forecast_total_profit_before_tax is not None
        ]
        ad_profit_before_tax = sum(ad_profit_values) if ad_profit_values else None
        organic_profit_before_tax = (
            sum(organic_profit_values) if organic_profit_values else None
        )
        total_profit_before_tax = sum(total_profit_values) if total_profit_values else None

        return OzonSalesForecastRow(
            sku="TOTAL",
            offer_id=None,
            title="Итого",
            target_ad_spend_with_vat=sum(row.target_ad_spend_with_vat for row in rows),
            target_drr_percent=(
                sum(row.forecast_ad_spend_without_vat for row in rows)
                / sum(row.forecast_ad_revenue_without_vat for row in rows)
                * 100
                if sum(row.forecast_ad_revenue_without_vat for row in rows)
                else None
            ),
            ad_spend_share_percent=sum(row.ad_spend_share_percent for row in rows),
            fact_total_orders=sum(row.fact_total_orders for row in rows),
            fact_total_revenue_without_vat=sum(
                row.fact_total_revenue_without_vat for row in rows
            ),
            fact_ad_spend_without_vat=sum(row.fact_ad_spend_without_vat for row in rows),
            fact_ad_revenue_without_vat=sum(row.fact_ad_revenue_without_vat for row in rows),
            fact_organic_orders=sum(row.fact_organic_orders for row in rows),
            fact_organic_revenue_without_vat=sum(
                row.fact_organic_revenue_without_vat for row in rows
            ),
            forecast_ad_spend_without_vat=sum(
                row.forecast_ad_spend_without_vat for row in rows
            ),
            forecast_ad_revenue_without_vat=sum(
                row.forecast_ad_revenue_without_vat for row in rows
            ),
            forecast_ad_orders=sum(row.forecast_ad_orders for row in rows),
            forecast_organic_orders=sum(row.forecast_organic_orders for row in rows),
            forecast_organic_revenue_without_vat=sum(
                row.forecast_organic_revenue_without_vat for row in rows
            ),
            forecast_total_orders=sum(row.forecast_total_orders for row in rows),
            forecast_total_revenue_without_vat=sum(
                row.forecast_total_revenue_without_vat for row in rows
            ),
            unit_expense_with_ozon_commission=None,
            forecast_ad_profit_before_tax=ad_profit_before_tax,
            forecast_ad_net_profit=self._economics_calculator.calculate_net_profit(
                ad_profit_before_tax
            ),
            forecast_organic_profit_before_tax=organic_profit_before_tax,
            forecast_organic_net_profit=self._economics_calculator.calculate_net_profit(
                organic_profit_before_tax
            ),
            forecast_total_profit_before_tax=total_profit_before_tax,
            forecast_total_net_profit=self._economics_calculator.calculate_net_profit(
                total_profit_before_tax
            ),
            matched_unit_economy=len(total_profit_values) == len(rows),
        )

    async def _resolve_campaign_ids(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> list[str]:
        if campaign_id != "ALL":
            return [campaign_id]

        campaigns = await self.get_campaigns_for_period(date_from, date_to)
        campaign_ids = [campaign.id for campaign in campaigns]
        if not campaign_ids:
            raise ValueError("No Ozon campaigns were found for selected period")

        return campaign_ids

    async def _create_statistics_report(
        self,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
    ) -> str:
        data = await self._performance_client.request(
            "POST",
            "/api/client/statistics",
            {
                "campaigns": campaign_ids,
                "dateFrom": date_from,
                "dateTo": date_to,
                "groupBy": "NO_GROUP_BY",
            },
        )
        uuid = data.get("UUID")
        if not isinstance(uuid, str):
            raise ValueError("Ozon statistics report UUID is missing")

        return uuid

    async def _create_statistics_report_with_limit_wait(
        self,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
    ) -> str:
        for attempt in range(self._settings.ozon_report_wait_attempts):
            try:
                return await self._create_statistics_report(campaign_ids, date_from, date_to)
            except OzonApiError as exc:
                if (
                    not _is_active_report_limit_error(str(exc))
                    or attempt == self._settings.ozon_report_wait_attempts - 1
                ):
                    raise

                await asyncio.sleep(self._settings.ozon_report_wait_interval_seconds)

        raise TimeoutError("Ozon statistics report active slot was not released in time")

    async def _performance_request_with_active_limit_wait(
        self,
        method: str,
        path: str,
        payload: dict[str, object] | None = None,
    ) -> dict[str, object]:
        for attempt in range(self._settings.ozon_report_wait_attempts):
            try:
                return await self._performance_client.request(method, path, payload)
            except OzonApiError as exc:
                if (
                    not _is_active_report_limit_error(str(exc))
                    or attempt == self._settings.ozon_report_wait_attempts - 1
                ):
                    raise

                await asyncio.sleep(self._settings.ozon_report_wait_interval_seconds)

        raise TimeoutError("Ozon Performance API active slot was not released in time")

    async def _wait_statistics_report(self, uuid: str) -> str:
        for _ in range(self._settings.ozon_report_wait_attempts):
            data = await self._performance_client.request(
                "GET",
                f"/api/client/statistics/{uuid}",
            )
            state = data.get("state")
            if state == "OK":
                return "OK"
            if state not in {"NOT_STARTED", "IN_PROGRESS"}:
                raise ValueError(f"Ozon statistics report failed with state {state}")

            await asyncio.sleep(self._settings.ozon_report_wait_interval_seconds)

        raise TimeoutError("Ozon statistics report was not ready in time")

    async def _download_statistics_report_or_recreate(
        self,
        campaign_key: str,
        campaign_ids: list[str],
        date_from: str,
        date_to: str,
        report_uuid: str,
        allow_recreate: bool,
    ) -> tuple[str, str, str]:
        try:
            return await self._performance_client.download_report(report_uuid), report_uuid, "OK"
        except OzonApiError as exc:
            if not allow_recreate or not _is_missing_report_file_error(exc):
                raise

        fresh_uuid = await self._create_statistics_report_with_limit_wait(
            campaign_ids,
            date_from,
            date_to,
        )
        await self._save_report_status(
            OzonStatisticsReportStatus(
                campaign_id=campaign_key,
                report_uuid=fresh_uuid,
                state="CREATED",
            ),
            date_from,
            date_to,
        )
        fresh_state = await self._wait_statistics_report(fresh_uuid)
        report_csv = await self._performance_client.download_report(fresh_uuid)
        return report_csv, fresh_uuid, fresh_state

    async def _get_campaign_performance_metrics(
        self,
        campaign_id: str,
        date_from: str,
        date_to: str,
    ) -> OzonCampaignPerformanceMetrics:
        data = await self._performance_client.request(
            "GET",
            (
                "/api/client/statistics/campaign/product/json"
                f"?dateFrom={date_from}&dateTo={date_to}&campaignIds={campaign_id}"
            ),
        )
        rows = data.get("rows")
        if not isinstance(rows, list) or not rows:
            return OzonCampaignPerformanceMetrics(
                campaign_id=campaign_id,
                views=0,
                clicks=0,
                to_cart=0,
                orders=0,
                revenue=0,
                ad_spend=0,
                drr_percent=None,
                ctr_percent=None,
                average_cpc=None,
            )

        row = rows[0]
        if not isinstance(row, dict):
            raise ValueError("Ozon campaign statistics row is invalid")

        return OzonCampaignPerformanceMetrics(
            campaign_id=str(row.get("id") or campaign_id),
            title=_optional_str(row.get("title")),
            status=_optional_str(row.get("status")),
            views=parse_int(row.get("views")),
            clicks=parse_int(row.get("clicks")),
            to_cart=parse_int(row.get("toCart")),
            orders=parse_int(row.get("orders")),
            revenue=parse_decimal(row.get("ordersMoney")),
            ad_spend=parse_decimal(row.get("moneySpent")),
            drr_percent=parse_optional_decimal(row.get("drr")),
            ctr_percent=parse_optional_decimal(row.get("ctr")),
            average_cpc=parse_optional_decimal(row.get("clickPrice")),
        )


def _optional_str(value: object) -> str | None:
    return str(value) if value is not None else None


_LIFT_OPTION_LABELS = {
    "none": "Не требуется",
    "stairs": "Подъём по лестнице",
    "lift": "Подъём на лифте",
}


def _optional_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _apply_delivery_actual_cost(
    order: OzonOrderLookupResponse,
    actual_cost: float | None,
) -> OzonOrderLookupResponse:
    if actual_cost is None:
        return order
    delivery_result = None
    net_profit_with_actual = None
    if order.delivery_total_transferred is not None:
        delivery_result = order.delivery_total_transferred - actual_cost
        if order.net_profit_total is not None:
            net_profit_with_actual = order.net_profit_total + delivery_result
    return order.model_copy(
        update={
            "delivery_cost_actual": actual_cost,
            "delivery_result": delivery_result,
            "net_profit_with_delivery_actual": net_profit_with_actual,
        }
    )


def _parse_prr_option(raw: object) -> tuple[str | None, str | None, float | None, str | None]:
    """prr_option comes back as a bare code string from /v3/posting/fbs/list
    but as {code, price, currency_code, floor} from /v3/posting/fbs/get -
    handle both shapes since callers may pass either."""
    if isinstance(raw, dict):
        code = _optional_str(raw.get("code")) or None
        price = _optional_float(raw.get("price"))
        floor = raw.get("floor")
        floor = str(floor) if floor not in (None, "") else None
    elif isinstance(raw, str) and raw:
        code, price, floor = raw, None, None
    else:
        code, price, floor = None, None, None
    label = _LIFT_OPTION_LABELS.get(code) if code else None
    return code, label, price, floor


def _first_string(value: object) -> str | None:
    if isinstance(value, str):
        return value

    if isinstance(value, list):
        for item in value:
            if isinstance(item, str):
                return item

    return None


def _campaign_overlaps_period(
    campaign: OzonCampaign,
    date_from: str,
    date_to: str,
) -> bool:
    period_start = date.fromisoformat(date_from)
    period_end = date.fromisoformat(date_to)
    campaign_start = _parse_optional_date(campaign.raw.get("fromDate")) or _parse_optional_date(
        campaign.raw.get("createdAt")
    )
    if campaign_start is None or campaign_start > period_end:
        return False

    raw_end = _parse_optional_date(campaign.raw.get("toDate"))
    if raw_end is not None:
        return raw_end >= period_start

    state = campaign.state or ""
    if state == "CAMPAIGN_STATE_RUNNING":
        return True

    updated_at = _parse_optional_date(campaign.raw.get("updatedAt"))
    return updated_at is None or updated_at >= period_start


def _parse_optional_date(value: object) -> date | None:
    if not isinstance(value, str) or not value.strip():
        return None

    text = value.strip()
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
        except ValueError:
            return None


def _parse_sku_report_csv(content: str) -> list[dict[str, str]]:
    lines = [line for line in content.splitlines() if line.strip()]
    if len(lines) < 2:
        return []

    current_campaign_id: str | None = None
    current_campaign_title: str | None = None
    headers: list[str] | None = None
    rows: list[dict[str, str]] = []

    for line in lines:
        cells = _parse_csv_line(line)
        campaign_meta = _parse_report_campaign_header(cells)
        if campaign_meta is not None:
            current_campaign_id, current_campaign_title = campaign_meta
            continue

        if _is_sku_report_header(cells):
            headers = [canonicalize_ozon_ad_report_header(cell) for cell in cells]
            continue

        if headers is None:
            continue

        values = [cell.strip() for cell in cells]
        normalized_values = values + [""] * max(len(headers) - len(values), 0)
        row = dict(zip(headers, normalized_values, strict=False))
        if current_campaign_id is not None:
            row["__campaign_id"] = current_campaign_id
        if current_campaign_title is not None:
            row["__campaign_title"] = current_campaign_title
        rows.append(row)

    if rows:
        return rows

    header_index = _find_sku_report_header_index(lines)
    reader = csv.DictReader(StringIO("\n".join(lines[header_index:])), delimiter=";")
    return [
        {
            canonicalize_ozon_ad_report_header(header): value
            for header, value in row.items()
            if header is not None
        }
        for row in reader
        if row
    ]


def _parse_csv_line(line: str) -> list[str]:
    return next(csv.reader(StringIO(line), delimiter=";"), [])


def _is_sku_report_header(cells: list[str]) -> bool:
    return is_ozon_ad_report_header(cells)


def _parse_report_campaign_header(cells: list[str]) -> tuple[str, str] | None:
    text = next(
        (cell.strip() for cell in cells if "Кампания по продвижению товаров" in cell),
        "",
    )
    if not text:
        return None

    campaign_id = _extract_campaign_id(text)
    if campaign_id is None:
        return None

    title = text.split(", период", 1)[0].strip()
    return campaign_id, title


def _extract_campaign_id(text: str) -> str | None:
    if "№" not in text:
        return None

    tail = text.split("№", 1)[1].strip()
    digits = []
    for char in tail:
        if not char.isdigit():
            break
        digits.append(char)

    return "".join(digits) or None


def _find_sku_report_header_index(lines: list[str]) -> int:
    for index, line in enumerate(lines):
        if _is_sku_report_header(_parse_csv_line(line)):
            return index

    raise ValueError("Ozon SKU report header row was not found")


def _build_import_report_uuid(
    campaign_id: str,
    date_from: str,
    date_to: str,
    filename: str,
    content: bytes,
) -> str:
    digest = hashlib.sha256()
    digest.update(campaign_id.encode("utf-8"))
    digest.update(date_from.encode("utf-8"))
    digest.update(date_to.encode("utf-8"))
    digest.update(filename.encode("utf-8"))
    digest.update(content)
    return f"import-{digest.hexdigest()[:24]}"


def _build_composite_report_uuid(
    campaign_id: str,
    date_from: str,
    date_to: str,
) -> str:
    digest = hashlib.sha256(
        f"{campaign_id}:{date_from}:{date_to}".encode("utf-8")
    ).hexdigest()
    return f"composite-{digest[:24]}"


def _build_campaign_group_key(campaign_ids: list[str]) -> str:
    digest = hashlib.sha256(",".join(campaign_ids).encode("utf-8")).hexdigest()
    return f"campaigns-{len(campaign_ids)}-{digest[:16]}"


def _chunked(values: list[str], chunk_size: int) -> list[list[str]]:
    return [values[index : index + chunk_size] for index in range(0, len(values), chunk_size)]


def _summarize_row_promotions(promotions: list[object]) -> dict[str, object]:
    action_ids: list[str] = []
    titles: list[str] = []
    prices: list[float] = []
    discounts: list[float] = []

    for promotion in promotions:
        action_id = str(getattr(promotion, "action_id", "") or "").strip()
        if action_id and action_id not in action_ids:
            action_ids.append(action_id)

        title = str(getattr(promotion, "title", "") or "").strip()
        if title and title not in titles:
            titles.append(title)

        action_price = getattr(promotion, "action_price_with_vat", None)
        max_action_price = getattr(promotion, "max_action_price_with_vat", None)
        price = action_price if action_price is not None else max_action_price
        if price is not None:
            prices.append(float(price))

        discount = getattr(promotion, "discount_percent", None)
        if discount is not None:
            discounts.append(float(discount))

    return {
        "matched": bool(action_ids or titles),
        "count": len(action_ids or titles),
        "action_ids": action_ids,
        "titles": titles,
        "price_with_vat": min(prices) if prices else None,
        "discount_percent": max(discounts) if discounts else None,
    }


def _merge_sales_row_promotion_summary(rows: list[OzonSalesAnalyticsRow]) -> dict[str, object]:
    action_ids: list[str] = []
    titles: list[str] = []
    prices: list[float] = []
    discounts: list[float] = []
    for row in rows:
        for action_id in row.promotion_action_ids:
            if action_id not in action_ids:
                action_ids.append(action_id)
        for title in row.promotion_titles:
            if title not in titles:
                titles.append(title)
        if row.promotion_price_with_vat is not None:
            prices.append(row.promotion_price_with_vat)
        if row.promotion_discount_percent is not None:
            discounts.append(row.promotion_discount_percent)

    return {
        "matched": bool(action_ids or titles),
        "count": len(action_ids or titles),
        "action_ids": action_ids,
        "titles": titles,
        "price_with_vat": min(prices) if prices else None,
        "discount_percent": max(discounts) if discounts else None,
    }


def _row_promotions_from_ad_row(row: OzonSkuEfficiencyRow) -> list[OzonPromotionInfo]:
    return [
        OzonPromotionInfo(
            action_id=action_id,
            title=row.promotion_titles[index] if index < len(row.promotion_titles) else None,
            action_price_with_vat=row.promotion_price_with_vat,
            discount_percent=row.promotion_discount_percent,
        )
        for index, action_id in enumerate(row.promotion_action_ids)
    ]


def _price_discount_percent(
    actual_price_without_vat: float | None,
    unit_price_without_vat: float | None,
) -> float | None:
    if (
        actual_price_without_vat is None
        or unit_price_without_vat is None
        or unit_price_without_vat <= 0
    ):
        return None

    return (1 - actual_price_without_vat / unit_price_without_vat) * 100


def _build_response_segment(
    response: OzonSkuEfficiencyResponse,
) -> OzonSkuEfficiencySegment:
    return OzonSkuEfficiencySegment(
        date_from=response.date_from,
        date_to=response.date_to,
        report_uuid=response.report_uuid,
        report_state=response.report_state,
        unit_economy_version=response.unit_economy_version,
        unit_economy_version_valid_from=response.unit_economy_version_valid_from,
        unit_economy_workbook_path=response.unit_economy_workbook_path,
        rows=response.rows,
        total=response.total,
    )


def _build_composite_warning(
    segments: list[UnitEconomyPeriodSegment],
) -> str:
    segment_text = "; ".join(
        (
            f"{segment.date_from} - {segment.date_to}: "
            f"юнитка от {segment.version.valid_from}"
        )
        for segment in segments
    )
    return (
        "Составной расчёт: период разделён по версиям юнит-экономики. "
        f"{segment_text}. Итоговые продажи, расходы и прибыль сложены backend по сегментам."
    )


def _append_warning(existing: str | None, message: str) -> str:
    return f"{existing} {message}" if existing else message


RUN_TYPE_LABELS: dict[str, str] = {
    "final": "Итог за день",
    "preliminary": "Предварительно",
}


_ORDER_STATUS_LABELS: dict[str, str] = {
    "awaiting_registration": "Ожидает регистрации",
    "acceptance_in_progress": "Идёт приёмка",
    "awaiting_approve": "Ожидает подтверждения",
    "awaiting_packaging": "Ожидает упаковки",
    "awaiting_deliver": "Готов к отгрузке",
    "awaiting_verification": "Ожидает проверки",
    "arbitration": "Арбитраж",
    "client_arbitration": "Клиентский арбитраж",
    "delivering": "Доставляется",
    "driver_pickup": "У курьера",
    "delivered": "Доставлен",
    "cancelled": "Отменён",
    "not_accepted": "Не принят",
    "sent_by_seller": "Отправлен продавцом",
}


def _index_ad_rows(
    rows: list[OzonSkuEfficiencyRow],
) -> dict[str, OzonSkuEfficiencyRow]:
    result: dict[str, OzonSkuEfficiencyRow] = {}
    for row in rows:
        if row.sku == "TOTAL":
            continue

        for key in _row_match_keys(row.offer_id, row.sku):
            result.setdefault(key, row)

    return result


def _index_total_sales_rows(
    rows: list[OzonTotalSalesReportRow],
) -> dict[str, OzonTotalSalesReportRow]:
    result: dict[str, OzonTotalSalesReportRow] = {}
    for row in rows:
        for match_key in _row_match_keys(row.offer_id, row.sku):
            result.setdefault(match_key, row)

    return result


def _total_sales_match_key(row: OzonTotalSalesReportRow) -> str | None:
    keys = _row_match_keys(row.offer_id, row.sku)
    return keys[0] if keys else None


def _first_matching_row(
    rows_by_key: dict[str, OzonSkuEfficiencyRow],
    keys: list[str],
) -> OzonSkuEfficiencyRow | None:
    for key in keys:
        row = rows_by_key.get(key)
        if row is not None:
            return row

    return None


def _first_matching_sales_row(
    rows_by_key: dict[str, OzonSalesAnalyticsRow],
    keys: list[str],
) -> OzonSalesAnalyticsRow | None:
    for key in keys:
        row = rows_by_key.get(key)
        if row is not None:
            return row

    return None


def _with_forecast_budget_shares(
    rows: list[OzonSalesForecastRow],
) -> list[OzonSalesForecastRow]:
    total_ad_spend = sum(row.forecast_ad_spend_without_vat for row in rows)
    return [
        row.model_copy(
            update={
                "ad_spend_share_percent": (
                    row.forecast_ad_spend_without_vat / total_ad_spend * 100
                    if total_ad_spend
                    else 0
                )
            }
        )
        for row in rows
    ]


def _sales_row_key(row: OzonSalesAnalyticsRow) -> str:
    return row.offer_id or row.sku


def _forecast_allocation_share(
    row: OzonSalesAnalyticsRow,
    allocation_total: float,
    use_ad_spend: bool,
) -> float:
    if allocation_total <= 0:
        return 0

    value = row.ad_spend_without_vat if use_ad_spend else row.total_revenue_without_vat
    return value / allocation_total


def _forecast_average_order_revenue_without_vat(
    row: OzonSalesAnalyticsRow,
) -> float | None:
    if row.ad_orders > 0 and row.ad_revenue_without_vat > 0:
        return row.ad_revenue_without_vat / row.ad_orders

    if row.total_orders > 0 and row.total_revenue_without_vat > 0:
        return row.total_revenue_without_vat / row.total_orders

    return None


def _row_match_keys(offer_id: str | None, sku: str | None) -> list[str]:
    keys: list[str] = []
    if offer_id:
        keys.append(f"offer:{offer_id.strip().lower()}")
    if sku:
        keys.append(f"sku:{sku.strip().lower()}")

    return keys


def _resolve_attribution_product(
    sku: str,
    product: OzonCampaignProduct,
) -> ResolvedAttributionProduct:
    unit_economy = product.unit_economy
    offer_id = product.offer_id or (
        unit_economy.offer_id if unit_economy is not None else None
    )
    if unit_economy is None:
        status = (
            "UNIT_ECONOMICS_NOT_FOUND"
            if offer_id is not None
            else "SKU_MAPPING_NOT_FOUND"
        )
    elif unit_economy.expense_with_ozon_commission is None:
        status = "UNIT_COST_NOT_FOUND"
    else:
        status = "OK"

    if product.offer_id:
        mapping_source = "seller_api+unit_economy"
    elif unit_economy is not None:
        mapping_source = "unit_economy_sku"
    else:
        mapping_source = "not_found"

    return ResolvedAttributionProduct(
        sku=sku,
        offer_id=offer_id,
        title=product.title or (
            unit_economy.title if unit_economy is not None else None
        ),
        unit_economy=unit_economy,
        mapping_source=mapping_source,
        status=status,
    )


def _result_matches_unit_version(
    response: OzonSkuEfficiencyResponse,
    selection: UnitEconomyVersionSelection | None,
) -> bool:
    if selection is None:
        return response.unit_economy_version is None

    return response.unit_economy_version == selection.version.version_id


def _result_has_campaign_source_metadata(response: OzonSkuEfficiencyResponse) -> bool:
    if not response.campaign_id.startswith("campaigns-"):
        return True

    return all(row.campaign_ids for row in response.rows)


def _result_uses_direct_sku_profit(response: OzonSkuEfficiencyResponse) -> bool:
    return all(
        row.profit_breakdown is not None
        and _money_values_close(
            row.profit_breakdown.ad_spend_without_vat,
            row.sku_ad_spend_without_vat,
        )
        for row in response.rows
    )


def _result_uses_model_attribution_rule(response: OzonSkuEfficiencyResponse) -> bool:
    for row in response.rows:
        if row.profit_breakdown is None:
            return False

        should_include_model = can_include_model_attribution_in_sku(
            total_ordered_amount_with_vat=row.total_ordered_amount_with_vat,
            attribution_revenue_with_vat=row.revenue_with_vat,
            model_revenue_with_vat=row.model_revenue_with_vat,
            direct_revenue_with_vat=row.direct_revenue_with_vat,
            direct_orders=row.direct_orders,
            model_orders=row.model_orders,
        )
        expected_orders = row.orders if should_include_model else row.direct_orders
        expected_revenue_without_vat = (
            row.revenue_without_vat
            if should_include_model
            else row.direct_revenue_without_vat
        )

        if row.model_attribution_included_in_sku != should_include_model:
            return False
        if row.sku_orders != row.direct_orders:
            return False
        if not _money_values_close(
            row.profit_breakdown.revenue_without_vat,
            expected_revenue_without_vat,
        ):
            return False
        if row.profit_breakdown.orders != expected_orders:
            return False

    return True


def _result_uses_non_negative_profit_tax(response: OzonSkuEfficiencyResponse) -> bool:
    for row in response.rows:
        if row.profit_before_tax is None or row.net_profit is None:
            continue

        expected_tax = max(row.profit_before_tax, 0) * PROFIT_TAX_RATE
        expected_net_profit = row.profit_before_tax - expected_tax
        if not _money_values_close(row.net_profit, expected_net_profit):
            return False

    return True


def _money_values_close(left: float, right: float) -> bool:
    return abs(left - right) < 0.01


def _sum_campaign_expense_csv(csv_text: str) -> float:
    """Sums the "Расход" column of /api/client/statistics/expense's CSV.

    Format: ID;Дата;Название;Расход;Расход бонусов;Расход с абонентского счета
    """
    lines = [line for line in csv_text.splitlines() if line.strip()]
    if not lines:
        return 0.0

    total = 0.0
    for line in lines[1:]:  # skip header
        columns = line.split(";")
        if len(columns) < 4:
            continue
        raw_value = columns[3].strip().replace(" ", "").replace(",", ".")
        try:
            total += float(raw_value)
        except ValueError:
            continue

    return total


def _sum_promotion_report_spend(csv_text: str) -> float:
    """Sums the "Всего" (per-campaign total) rows of the /api/client/statistics

    per-SKU report's "Расход, ₽, с НДС" column (index 8). Each campaign's
    section ends with one such row, so summing only those avoids
    double-counting the per-SKU rows that make up the same total.
    """
    total = 0.0
    for line in csv_text.splitlines():
        if not line.startswith("Всего;"):
            continue
        columns = line.split(";")
        if len(columns) < 9:
            continue
        raw_value = columns[8].strip().replace(" ", "").replace(",", ".")
        try:
            total += float(raw_value)
        except ValueError:
            continue

    return total


def _group_product_sales_rows_by_sku(
    rows: list[ProductSalesRow],
) -> dict[str, list[ProductSalesRow]]:
    grouped: dict[str, list[ProductSalesRow]] = {}
    for row in rows:
        grouped.setdefault(row.sku, []).append(row)
    return grouped


def _build_product_sales_total(
    rows: list[OzonProductSalesRow],
) -> OzonProductSalesTotal:
    revenue_without_vat = sum(row.revenue_without_vat for row in rows)
    profit_rows = [row for row in rows if row.profit_before_ads is not None]
    profit_before_ads = (
        sum(row.profit_before_ads for row in profit_rows) if profit_rows else None
    )
    return OzonProductSalesTotal(
        ordered_units=sum(row.ordered_units for row in rows),
        redeemed_units=sum(row.redeemed_units for row in rows),
        cancelled_units=sum(row.cancelled_units for row in rows),
        revenue_with_vat=sum(row.revenue_with_vat for row in rows),
        revenue_without_vat=revenue_without_vat,
        profit_before_ads=profit_before_ads,
    )


def _filter_promotion_report_by_campaign(
    report: PromotionAnalyticsReport,
    campaign_id: str,
) -> PromotionAnalyticsReport:
    if campaign_id == "ALL":
        return report

    statistics = [
        row for row in report.statistics if row.campaign_id == campaign_id
    ]
    union = [
        row for row in report.union if row.campaign_id == campaign_id
    ]
    if not statistics:
        raise ValueError(
            f"В загруженном отчёте нет кампании {campaign_id}."
        )

    return PromotionAnalyticsReport(
        statistics=statistics,
        union=union,
        period_date_from=report.period_date_from,
        period_date_to=report.period_date_to,
    )


def _result_has_product_identity(response: OzonSkuEfficiencyResponse) -> bool:
    return all(row.offer_id for row in response.rows if row.sku != "TOTAL")


def _result_has_price_diagnostics(response: OzonSkuEfficiencyResponse) -> bool:
    for row in response.rows:
        if row.sku == "TOTAL":
            continue
        if row.profit_breakdown is None or row.profit_breakdown.orders <= 0:
            continue
        if row.average_ad_order_revenue_with_vat is None:
            return False
        if row.unit_economy_price_with_vat is not None and row.ad_price_discount_percent is None:
            return False

    return True


def _result_has_promotion_diagnostics(response: OzonSkuEfficiencyResponse) -> bool:
    return response.promotion_report_state is not None


def _analytics_row_sku(row: dict[str, object]) -> str | None:
    dimensions = row.get("dimensions")
    if not isinstance(dimensions, list) or not dimensions:
        return None

    first_dimension = dimensions[0]
    if not isinstance(first_dimension, dict):
        return None

    sku = first_dimension.get("id")
    return str(sku) if sku is not None else None


def _analytics_row_name(row: dict[str, object]) -> str | None:
    dimensions = row.get("dimensions")
    if not isinstance(dimensions, list) or not dimensions:
        return None

    first_dimension = dimensions[0]
    if not isinstance(first_dimension, dict):
        return None

    name = first_dimension.get("name")
    return str(name) if name is not None else None


def _parse_daily_sales_api_row(row: dict[str, object]) -> dict[str, object] | None:
    dimensions = row.get("dimensions")
    metrics = row.get("metrics")
    if (
        not isinstance(dimensions, list)
        or len(dimensions) < 2
        or not isinstance(metrics, list)
        or len(metrics) < 2
    ):
        return None

    sku_dimension = dimensions[0]
    day_dimension = dimensions[1]
    if not isinstance(sku_dimension, dict) or not isinstance(day_dimension, dict):
        return None

    sku = sku_dimension.get("id")
    day = day_dimension.get("id") or day_dimension.get("name")
    if sku is None or day is None:
        return None

    return {
        "sku": str(sku),
        "title": _optional_str(sku_dimension.get("name")),
        "date": str(day)[:10],
        "ordered_units": metrics[0],
        "ordered_amount_with_vat": metrics[1],
    }


def _first_daily_title(rows: list[dict[str, object]]) -> str | None:
    for row in rows:
        title = row.get("title")
        if title:
            return str(title)

    return None


def _is_active_report_limit_error(message: str) -> bool:
    normalized = message.lower()
    return (
        "превышен лимит активных запросов" in normalized
        or "active requests" in normalized
        or "maximum 1" in normalized
    )


def _is_forbidden_campaign_list_error(message: str) -> bool:
    normalized = message.lower()
    return (
        "generation of this type of report is forbidden" in normalized
        or "forbidden for the transferred list of campaigns" in normalized
        or "invalidargument" in normalized
        and "campaign" in normalized
        and "forbidden" in normalized
    )


def _is_missing_report_file_error(error: OzonApiError) -> bool:
    normalized = str(error).lower()
    return error.status_code == 404 and (
        "nosuchkey" in normalized
        or "no such key" in normalized
    )
