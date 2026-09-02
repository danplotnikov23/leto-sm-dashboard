from io import BytesIO
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.domain.models import DashboardKpi, ProductAnalysis, ShortlistItem


class UnitEconomicsExcelExporter:
    def export(self, analyses: list[ProductAnalysis], kpi: DashboardKpi) -> bytes:
        workbook = Workbook()
        self._configure_formula_recalculation(workbook)
        summary = workbook.active
        summary.title = "Сводка"
        details = workbook.create_sheet("Юнитка")
        assumptions = workbook.create_sheet("Допущения")
        formulas = workbook.create_sheet("Пояснения")

        self._build_summary(summary, kpi)
        self._build_details(details, analyses)
        self._build_assumptions(assumptions)
        self._build_formulas(formulas)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_shortlist(self, items: list[ShortlistItem]) -> bytes:
        workbook = Workbook()
        self._configure_formula_recalculation(workbook)
        summary = workbook.active
        summary.title = "Сводка"
        selection = workbook.create_sheet("Отбор")
        formulas = workbook.create_sheet("Формулы")
        assumptions = workbook.create_sheet("Допущения")

        self._build_shortlist_summary(summary)
        self._build_shortlist_details(selection, items)
        self._build_shortlist_formulas(formulas)
        self._build_assumptions(assumptions)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _build_summary(self, sheet, kpi: DashboardKpi) -> None:
        sheet["A1"] = "Лето СМ Платформа: примерная юнитка"
        sheet["A1"].font = Font(size=16, bold=True, color="17233C")
        sheet["A3"] = "Метрика"
        sheet["B3"] = "Значение"
        rows = [
            ("Товаров", kpi.total_products),
            ("Выгодных", kpi.profitable_products),
            ("Убыточных", kpi.unprofitable_products),
            ("Средняя маржа", f"{kpi.average_margin_percent}%"),
            ("Потенциальная прибыль", kpi.potential_profit),
            ("Риск", kpi.high_risk_products),
            ("Конкурент ниже безубыточной", kpi.competitor_below_break_even),
        ]
        for row_index, row in enumerate(rows, start=4):
            sheet.cell(row=row_index, column=1, value=row[0])
            sheet.cell(row=row_index, column=2, value=row[1])
        self._style_table(sheet, 3, 1, len(rows) + 1, 2)
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 22

    def _build_details(self, sheet, analyses: list[ProductAnalysis]) -> None:
        headers = [
            "Артикул",
            "Схема",
            "Ozon SKU ID",
            "Товар",
            "Тип/категория Ozon",
            "Статус",
            "Остаток",
            "Объем, л",
            "Вес, кг",
            "Штрихкод",
            "Цена до скидки с НДС",
            "Поставить в Ozon с НДС",
            "Цена с картой",
            "Покупатель видит",
            "Соинвест Ozon",
            "Программы партнеров",
            "Начисление доставки",
            "Нам начислят всего",
            "Услуги Ozon",
            "После услуг Ozon",
            "Мин. цена Ozon",
            "Цена без НДС",
            "Торг. наценка",
            "Категория",
            "Закупка с НДС",
            "Закупка без НДС",
            "Закупка + УСН 7%",
            "Конкурент оценка",
            "Тип конкурента",
            "Ссылка конкурента",
            "Поиск на Ozon",
            "Доставка/логистика",
            "Упаковка",
            "Fulfillment",
            "Комиссия Ozon",
            "Эквайринг",
            "DRR",
            "Реклама",
            "Прочие",
            "Себес без комиссии",
            "Себес + комиссия",
            "Расходы до налога",
            "Прибыль до налога",
            "Налоговый режим",
            "НДС применяется",
            "База УСН/налога",
            "УСН 6%",
            "1% свыше 300к",
            "Налоги всего",
            "Чистая прибыль Ozon",
            "Быстрый вывод %",
            "Быстрый вывод",
            "Дизайнер/контент %",
            "Дизайнер/контент",
            "Фулфилмент и забор %",
            "Фулфилмент и забор",
            "Итого доп. расходы бизнеса",
            "Чистая прибыль бизнеса",
            "FBS 1 заказ/день",
            "FBS 5 заказов/день",
            "FBS 10 заказов/день",
            "FBS 20 заказов/день",
            "FBS 40 заказов/день",
            "Маржа",
            "Рекоменд. цена",
            "Рекомендация",
            "DRR 10% прибыль",
            "DRR 12% прибыль",
            "DRR 15% прибыль",
            "Предупреждения",
        ]
        sheet.append(headers)
        for item in analyses:
            economics = item.economics
            scenarios = {scenario.drr_percent: scenario for scenario in economics.drr_scenarios}
            fbs_scenarios = {
                scenario.orders_per_day: scenario for scenario in economics.fbs_batch_scenarios
            }
            sheet.append(
                [
                    item.product.supplier_article,
                    "realFBS/FBS оценка",
                    "",
                    item.product.title,
                    item.product.category,
                    item.product.status,
                    item.product.stock,
                    item.product.dimensions.volume_liters,
                    item.product.weight_kg,
                    item.product.barcode,
                    economics.ozon_price_before_discount_vat_included,
                    economics.real_fbs_price_vat_included,
                    economics.bank_card_price_vat_included,
                    economics.buyer_payment_price_vat_included,
                    economics.ozon_bonus_accrual,
                    economics.partner_program_accrual,
                    economics.delivery_accrual,
                    economics.marketplace_gross_accrual_vat_included,
                    economics.ozon_services_total,
                    economics.expected_payout_after_ozon_services,
                    economics.ozon_min_price_vat_included,
                    economics.sale_price_vat_excluded,
                    economics.markup_percent,
                    item.product.category,
                    economics.purchase_price_vat_included,
                    economics.purchase_price_vat_excluded,
                    economics.tax_only_break_even_price_vat_included,
                    economics.estimated_competitor_price_vat_included,
                    item.competitor.leader.match_type if item.competitor.leader else "",
                    item.competitor.leader_url,
                    self._ozon_search_url(item),
                    economics.logistics,
                    economics.package_cost,
                    economics.fulfillment_processing_cost,
                    economics.ozon_commission,
                    economics.acquiring,
                    (
                        f"{economics.drr_scenarios[1].drr_percent}%"
                        if len(economics.drr_scenarios) > 1
                        else "12%"
                    ),
                    economics.advertising,
                    economics.other_costs,
                    economics.cost_basis_without_commission,
                    economics.cost_basis_with_commission,
                    economics.total_expenses_before_tax,
                    economics.profit_before_tax,
                    economics.tax_regime,
                    "да" if economics.vat_applicable else "нет",
                    economics.tax_income_basis,
                    economics.usn_tax,
                    economics.usn_additional_contribution,
                    economics.profit_tax,
                    economics.net_profit,
                    economics.fast_payout_fee_percent,
                    economics.fast_payout_fee,
                    economics.designer_content_percent,
                    economics.designer_content_cost,
                    economics.business_fulfillment_pickup_percent,
                    economics.business_fulfillment_pickup_cost,
                    economics.business_extra_costs_total,
                    economics.business_net_profit,
                    (
                        fbs_scenarios.get(1).business_net_profit_after_fixed
                        if fbs_scenarios.get(1)
                        else None
                    ),
                    (
                        fbs_scenarios.get(5).business_net_profit_after_fixed
                        if fbs_scenarios.get(5)
                        else None
                    ),
                    (
                        fbs_scenarios.get(10).business_net_profit_after_fixed
                        if fbs_scenarios.get(10)
                        else None
                    ),
                    (
                        fbs_scenarios.get(20).business_net_profit_after_fixed
                        if fbs_scenarios.get(20)
                        else None
                    ),
                    (
                        fbs_scenarios.get(40).business_net_profit_after_fixed
                        if fbs_scenarios.get(40)
                        else None
                    ),
                    economics.margin_percent,
                    economics.recommended_price_vat_included,
                    economics.recommendation,
                    scenarios.get(10.0).net_profit if scenarios.get(10.0) else None,
                    scenarios.get(12.0).net_profit if scenarios.get(12.0) else None,
                    scenarios.get(15.0).net_profit if scenarios.get(15.0) else None,
                    " | ".join(economics.warnings),
                ]
            )
        self._style_table(sheet, 1, 1, len(analyses) + 1, len(headers))
        sheet.freeze_panes = "A2"
        widths = {
            "A": 18,
            "B": 46,
            "C": 28,
            "D": 14,
            "E": 14,
            "F": 16,
            "G": 16,
            "H": 16,
            "AO": 24,
            "AS": 60,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for index in range(11, 42):
            sheet.column_dimensions[get_column_letter(index)].width = max(
                sheet.column_dimensions[get_column_letter(index)].width or 0,
                14,
            )

    def _ozon_search_url(self, item: ProductAnalysis) -> str:
        query = " ".join(
            part
            for part in (
                item.product.title,
                item.product.category or "",
                item.product.brand or "",
            )
            if part
        )
        return f"https://www.ozon.ru/search/?text={quote_plus(query)}"

    def _build_shortlist_summary(self, sheet) -> None:
        sheet["A1"] = "Лето СМ Платформа: отбор товаров"
        sheet["A1"].font = Font(size=16, bold=True, color="17233C")
        rows = [
            ("Метрика", "Значение", "Формула"),
            ("Позиций в отборе", "=COUNTA(Отбор!A4:A10000)", "количество строк отбора"),
            ("План оборота", "=SUM(Отбор!BG4:BG10000)", "цена Ozon * план, шт"),
            ("Факт оборота", "=SUM(Отбор!BH4:BH10000)", "цена Ozon * продано, шт"),
            ("План прибыль бизнеса", "=SUM(Отбор!BI4:BI10000)", "прибыль после рейса * план"),
            ("Факт прибыль бизнеса", "=SUM(Отбор!BJ4:BJ10000)", "прибыль после рейса * факт"),
            ("Маржа плана", "=IF(B4>0,B6/B4,0)", "план прибыль / план оборота"),
            ("Цель оборота / месяц", 15_000_000, "ориентир запуска"),
            ("Выполнение цели", "=IF(B9>0,B4/B9,0)", "план оборота / цель"),
        ]
        for row in rows:
            sheet.append(row)
        self._style_table(sheet, 2, 1, len(rows), 3)
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 22
        sheet.column_dimensions["C"].width = 48
        for row_index in (8, 10):
            sheet.cell(row=row_index, column=2).number_format = "0.00%"
        for row_index in (4, 5, 6, 7, 9):
            sheet.cell(row=row_index, column=2).number_format = "#,##0 ₽"

    def _build_shortlist_details(self, sheet, items: list[ShortlistItem]) -> None:
        headers = [
            "Артикул",
            "Схема реализации",
            "Ozon SKU ID",
            "Название",
            "Тип",
            "Статус",
            "Источник / видимость",
            "На складе Ozon",
            "На моих складах",
            "Объем, л",
            "Вес, кг",
            "Штрихкод",
            "Цена до скидки с НДС, руб.",
            "Цена Ozon с НДС, руб.",
            "Цена без НДС, руб.",
            "Торговая наценка, %",
            "Цена для покупателя",
            "Закупочная цена с НДС, руб.",
            "Отвоз на СЦ Ozon",
            "FBS Ozon, руб.",
            "Фулфилмент, руб.",
            "Дизайнер, руб.",
            "Быстрый вывод, руб.",
            "Вознаграждение Ozon, %",
            "Вознаграждение Ozon, руб.",
            "Реклама, руб.",
            "Эквайринг, руб.",
            "Интеграция, руб.",
            "Прочие расходы, руб.",
            "Налог 6% + резерв 1%",
            "Забор у поставщика",
            "Чистая прибыль после рейса, руб.",
            "Рентабельность, %",
            "Баллы / соинвест Ozon",
            "Начисления всего",
            "Запросить скидку на закуп",
            "Себестоимость без комиссии",
            "Себестоимость + комиссия Ozon",
            "Цена конкурента",
            "Ссылка конкурента",
            "Штук в комплекте",
            "Соинвест Ozon, %",
            "Партнеры, %",
            "Начисление доставки, %",
            "Видимая скидка, %",
            "DRR, %",
            "Упаковка, руб.",
            "Обработка фулфилментом, руб.",
            "Забор поставщика в день, руб.",
            "Доставка на СЦ в день, руб.",
            "Заказов FBS в день",
            "Целевая маржа, %",
            "Безубыточная цена Ozon",
            "Закупка для целевой маржи",
            "Предельная закупка для безубытка",
            "Закупка за штуку",
            "План, шт.",
            "Продано, шт.",
            "План оборота",
            "Факт оборота",
            "План прибыль",
            "Факт прибыль",
            "Источник прайса",
            "Заметка",
            "Группа отбора",
            "Подгруппа отбора",
            "Длина, см",
            "Ширина, см",
            "Высота, см",
        ]
        groups = [
            ("A1:L1", "Товар и источник", "DDE8FA"),
            ("M1:Q1", "Цена и покупатель", "E6EEF9"),
            ("R1:AE1", "Расходы", "F4DDDD"),
            ("AF1:AG1", "Итог", "DDEFE3"),
            ("AH1:AI1", "Начисления Ozon", "DDE8FA"),
            ("AJ1:AL1", "Переговоры и себестоимость", "F7E8C7"),
            ("AM1:AN1", "Конкурент", "FFF2CC"),
            ("AO1:BQ1", "Редактируемые допущения и прогноз", "E2F0D9"),
        ]
        for range_ref, title, color in groups:
            sheet.merge_cells(range_ref)
            cell = sheet[range_ref.split(":")[0]]
            cell.value = title
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True, color="17233C")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.append(headers)
        rates_row: list[object | None] = [None] * len(headers)
        rates_row[21] = 0.04  # V: designer
        rates_row[22] = 0.0245  # W: fast payout
        rates_row[26] = 0.01  # AA: acquiring
        rates_row[29] = 0.07  # AD: USN 6% + 1% reserve
        sheet.append(rates_row)
        for item in items:
            row_index = sheet.max_row + 1
            economics = item.analysis.economics
            sale_price = item.entry.sale_price_vat_included or economics.real_fbs_price_vat_included
            partner_rate = self._safe_rate(economics.partner_program_accrual, sale_price)
            delivery_rate = self._safe_rate(economics.delivery_accrual, sale_price)
            competitor = item.analysis.competitor.leader
            sheet.append(
                [
                    item.analysis.product.supplier_article,
                    "FBS",
                    competitor.sku if competitor else "",
                    item.analysis.product.title,
                    item.analysis.product.category,
                    economics.recommendation,
                    item.entry.source_import_filename or "",
                    "",
                    item.analysis.product.stock,
                    item.analysis.product.dimensions.volume_liters,
                    item.analysis.product.weight_kg,
                    item.analysis.product.barcode,
                    f"=ROUND(N{row_index}/(1-AS{row_index}),2)",
                    sale_price,
                    f"=N{row_index}",
                    f"=IF(R{row_index}=0,0,(O{row_index}-R{row_index})/R{row_index})",
                    f"=MAX(N{row_index}-AH{row_index},0)",
                    economics.purchase_price_vat_included,
                    f"=IF(AY{row_index}>0,AX{row_index}/AY{row_index},0)",
                    economics.logistics,
                    f"=AU{row_index}+AV{row_index}",
                    f"=AI{row_index}*V$3",
                    f"=AI{row_index}*W$3",
                    economics.ozon_commission_percent / 100,
                    f"=N{row_index}*X{row_index}",
                    f"=N{row_index}*AT{row_index}",
                    f"=N{row_index}*AA$3",
                    20,
                    economics.other_costs,
                    f"=AI{row_index}*AD$3",
                    f"=IF(AY{row_index}>0,AW{row_index}/AY{row_index},0)",
                    (
                        f"=AI{row_index}-SUM(R{row_index}:W{row_index})-Y{row_index}"
                        f"-Z{row_index}-AA{row_index}-AB{row_index}-AC{row_index}"
                        f"-AD{row_index}-AE{row_index}"
                    ),
                    f"=IF(N{row_index}=0,0,AF{row_index}/N{row_index})",
                    f"=N{row_index}*AP{row_index}",
                    (
                        f"=Q{row_index}+AH{row_index}+N{row_index}*AQ{row_index}"
                        f"+N{row_index}*AR{row_index}"
                    ),
                    f"=MAX(R{row_index}-BB{row_index},0)",
                    f"=SUM(R{row_index}:W{row_index},Z{row_index}:AE{row_index})",
                    f"=AK{row_index}+Y{row_index}",
                    economics.estimated_competitor_price_vat_included,
                    item.analysis.competitor.leader_url,
                    item.entry.offer_quantity,
                    economics.seller_bonus_percent / 100,
                    partner_rate,
                    delivery_rate,
                    economics.ozon_visible_discount_percent / 100,
                    economics.advertising_drr_percent / 100,
                    economics.package_cost,
                    economics.fulfillment_processing_cost,
                    2500,
                    600,
                    20,
                    economics.target_margin_percent / 100,
                    (
                        f"=IFERROR((R{row_index}+S{row_index}+T{row_index}+U{row_index}"
                        f"+AB{row_index}+AC{row_index}+AE{row_index})/"
                        f"((1+AQ{row_index}+AR{row_index})-X{row_index}-AA$3-AT{row_index}"
                        f"-AD$3-(1+AQ{row_index}+AR{row_index})*(V$3+W$3)),0)"
                    ),
                    (
                        f"=MAX(AI{row_index}-SUM(S{row_index}:W{row_index})-Y{row_index}"
                        f"-Z{row_index}-AA{row_index}-AB{row_index}-AC{row_index}"
                        f"-AD{row_index}-AE{row_index}-N{row_index}*AZ{row_index},0)"
                    ),
                    (
                        f"=MAX(AI{row_index}-SUM(S{row_index}:W{row_index})-Y{row_index}"
                        f"-Z{row_index}-AA{row_index}-AB{row_index}-AC{row_index}"
                        f"-AD{row_index}-AE{row_index},0)"
                    ),
                    f"=R{row_index}/AO{row_index}",
                    item.entry.planned_sales_qty,
                    item.entry.sold_qty,
                    f"=N{row_index}*BE{row_index}",
                    f"=N{row_index}*BF{row_index}",
                    f"=AF{row_index}*BE{row_index}",
                    f"=AF{row_index}*BF{row_index}",
                    item.entry.source_import_filename or "",
                    item.entry.note,
                    item.entry.group_name,
                    item.entry.subgroup_name,
                    item.entry.length_cm
                    if item.entry.length_cm is not None
                    else item.analysis.product.dimensions.length_cm,
                    item.entry.width_cm
                    if item.entry.width_cm is not None
                    else item.analysis.product.dimensions.width_cm,
                    item.entry.height_cm
                    if item.entry.height_cm is not None
                    else item.analysis.product.dimensions.height_cm,
                ]
            )
        rows_count = max(len(items) + 3, 4)
        self._style_table(sheet, 2, 1, rows_count - 1, len(headers))
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A2:BQ{rows_count}"
        self._style_shortlist_sheet(sheet, rows_count)

    def _style_shortlist_sheet(self, sheet, rows_count: int) -> None:
        widths = {
            "A": 16,
            "B": 16,
            "C": 16,
            "D": 46,
            "E": 25,
            "F": 24,
            "G": 28,
            "AN": 42,
            "BK": 28,
            "BL": 34,
            "BM": 20,
            "BN": 20,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for index in range(8, 70):
            column = get_column_letter(index)
            sheet.column_dimensions[column].width = max(
                sheet.column_dimensions[column].width or 0,
                13,
            )
        percent_columns = ("P", "X", "AG", "AP", "AQ", "AR", "AS", "AT", "AZ")
        money_columns = (
            "M",
            "N",
            "O",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "Y",
            "Z",
            "AA",
            "AB",
            "AC",
            "AD",
            "AE",
            "AF",
            "AH",
            "AI",
            "AJ",
            "AK",
            "AL",
            "AM",
            "AU",
            "AV",
            "AW",
            "AX",
            "BA",
            "BB",
            "BC",
            "BD",
            "BG",
            "BH",
            "BI",
            "BJ",
        )
        for row in range(3, rows_count + 1):
            for column in money_columns:
                sheet[f"{column}{row}"].number_format = "#,##0 ₽"
            for column in percent_columns:
                sheet[f"{column}{row}"].number_format = "0.00%"
        positive_fill = PatternFill("solid", fgColor="E8F9F0")
        negative_fill = PatternFill("solid", fgColor="FFE8E6")
        input_fill = PatternFill("solid", fgColor="FFF8D9")
        for row in range(4, rows_count + 1):
            for column in (
                "N",
                "R",
                "AM",
                "AN",
                "AO",
                "AP",
                "AQ",
                "AR",
                "AS",
                "AT",
                "AU",
                "AV",
                "AW",
                "AX",
                "AY",
                "AZ",
                "BE",
                "BF",
                "BL",
                "BM",
                "BN",
                "BO",
                "BP",
                "BQ",
            ):
                sheet[f"{column}{row}"].fill = input_fill
        drr_validation = DataValidation(
            type="list",
            formula1='"8%,10%,12%,15%"',
            allow_blank=False,
        )
        drr_validation.error = "Выберите DRR: 8%, 10%, 12% или 15%."
        drr_validation.errorTitle = "Некорректный DRR"
        sheet.add_data_validation(drr_validation)
        drr_validation.add(f"AT4:AT{rows_count}")

        comments = {
            "N2": "Цена, которую продавец задает в Ozon. Меняйте ее для сценарного расчета.",
            "Q2": "Сумма, которую платит покупатель после баллов/соинвеста Ozon.",
            "R2": "Полная закупка комплекта с НДС поставщика. Это редактируемое поле.",
            "S2": "600 руб. за рейс на СЦ Ozon, разделенные на число заказов FBS за день.",
            "T2": "Тариф доставки FBS Ozon из backend-расчета по объему и весу.",
            "U2": "Упаковочные материалы плюс обработка фулфилментом.",
            "V2": "4% от начислений всего согласно принятой бизнес-модели.",
            "W2": "2,45% от начислений всего за быстрый вывод средств.",
            "AD2": "УСН 6% плюс резерв 1% после превышения годового порога.",
            "AE2": "2 500 руб. за забор у поставщика, разделенные на заказы дня.",
            "AF2": "Итог после Ozon, налогов, подготовки товара и доли фиксированного рейса.",
            "AJ2": "Разница между текущей закупкой и закупкой, нужной для целевой маржи.",
            "BA2": "Минимальная цена Ozon, при которой прибыль после всех расходов равна нулю.",
            "BB2": "Максимальная закупка, позволяющая получить заданную целевую маржу.",
            "BC2": "Максимальная закупка без убытка при текущей цене Ozon.",
        }
        for cell_ref, comment_text in comments.items():
            sheet[cell_ref].comment = Comment(comment_text, "Лето СМ Платформа")
        for column in ("AF", "AJ", "BI", "BJ"):
            range_ref = f"{column}4:{column}{rows_count}"
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill),
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=positive_fill),
            )

    def _max_competitor_orders(self, item: ProductAnalysis) -> int | None:
        orders = [
            offer.orders_count for offer in item.competitor.offers if offer.orders_count is not None
        ]
        return max(orders) if orders else None

    def _safe_rate(self, value: float, basis: float) -> float:
        if basis <= 0:
            return 0.0
        return value / basis

    def _build_shortlist_formulas(self, sheet) -> None:
        rows = [
            (
                "Редактируемые поля",
                (
                    "Желтые ячейки: цена Ozon, закупка, конкурент, проценты, упаковка, "
                    "обработка, рейс, план/факт, группа, габариты и заметка."
                ),
            ),
            ("Покупатель видит", "=Цена Ozon - Соинвест Ozon, но не ниже 0."),
            ("Начисления всего", "=Покупатель видит + соинвест + партнеры + доставка."),
            (
                "Услуги Ozon",
                ("=Комиссия + логистика + эквайринг. Внешний fulfillment считается отдельно."),
            ),
            (
                "Прибыль после рейса",
                (
                    "=Начисления всего - закупка - комиссия - FBS Ozon - эквайринг "
                    "- упаковка и обработка - реклама - налоги - доля рейса."
                ),
            ),
            (
                "Бизнес-расходы",
                (
                    "Быстрый вывод 2,45% и дизайнер 4% считаются от начислений всего. "
                    "Процентный резерв фулфилмента 5% отключен: используются реальные "
                    "операции и фиксированный рейс."
                ),
            ),
            (
                "Фиксированный рейс FBS",
                (
                    "Фикс дня 3 100 ₽: 2 500 ₽ забор + 600 ₽ сдача на СЦ Ozon. "
                    "Каждая часть делится на введенное число заказов FBS за день."
                ),
            ),
            (
                "Переговоры с поставщиком",
                (
                    "Закупка для целевой маржи показывает допустимую цену поставщика; "
                    "«Запросить скидку» — разница с текущей закупкой."
                ),
            ),
            (
                "План/факт",
                "План и факт оборота/прибыли считаются от введенной цены и количества.",
            ),
        ]
        sheet.append(["Блок", "Формула / смысл"])
        for row in rows:
            sheet.append(row)
        self._style_table(sheet, 1, 1, len(rows) + 1, 2)
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 110
        for row in sheet.iter_rows(min_row=2, max_row=len(rows) + 1):
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

    def _build_assumptions(self, sheet) -> None:
        rows = [
            ("Показатель", "Значение", "Комментарий"),
            ("Налоговый режим", "ИП УСН 6%", "Основной сценарий MVP под текущий режим."),
            ("НДС", "не применяется", "Для ИП УСН 6% в текущем сценарии НДС не вычитается."),
            ("НДС сценарий", "22%", "Оставлен в архитектуре для будущего режима ОСНО/УСН+НДС."),
            ("УСН", "6%", "Считается от начислений Ozon, а не от прибыли."),
            (
                "1% свыше 300к",
                "1%",
                "В юнитке учтен как предельная оценка после превышения 300 000 ₽ дохода в год.",
            ),
            ("DRR базовый", "12%", "В экспорте также есть сценарии 10%, 12%, 15%."),
            (
                "Видимая скидка Ozon",
                "30%",
                "Нужна для цены до скидки; не расход, если фактическая цена уже после скидки.",
            ),
            (
                "Цена с картой",
                "8%",
                "Сценарий цены покупателя с картой/банком; не списывается без подтверждения.",
            ),
            (
                "Баллы за скидки",
                "45%",
                "Положительное начисление из отчетов Ozon. Не вычитается как расход.",
            ),
            (
                "Программы партнеров",
                "0.5%",
                "Положительное начисление из блока продаж отчета Ozon.",
            ),
            ("Упаковка", "25 ₽", "Оценка для MVP, заменить справочником упаковки."),
            (
                "Обработка фулфилментом",
                "45 ₽",
                "Стартовая оценка. В отборе редактируется по фактическим операциям.",
            ),
            (
                "FBS до 1 литра",
                "63 ₽",
                (
                    "Сверено 17.06.2026 с calculator.ozon.ru: обработка 30 ₽, "
                    "логистика 17 ₽, доставка до места выдачи 16 ₽."
                ),
            ),
            (
                "FBS от 1 до 3 литров",
                "139 ₽",
                (
                    "Сверено 17.06.2026 с calculator.ozon.ru для 2,7555 л / 0,81 кг: "
                    "обработка 30 ₽, логистика 84 ₽, последняя миля 25 ₽."
                ),
            ),
            (
                "Фиксированный рейс FBS",
                "3 100 ₽/день",
                "Забор у поставщика 2 500 ₽ + доставка на СЦ Ozon 600 ₽.",
            ),
            (
                "Быстрый вывод средств Ozon",
                "2.45%",
                "Дополнительный расход бизнеса от начислений всего, прибыль Ozon не меняет.",
            ),
            (
                "Дизайнер / контент",
                "4%",
                "Дополнительный расход бизнеса от начислений всего, прибыль Ozon не меняет.",
            ),
            (
                "Фулфилмент и забор товара",
                "0%",
                (
                    "Процентный резерв отключен, чтобы не считать расход дважды. "
                    "Используются обработка по товару и фиксированный рейс FBS."
                ),
            ),
        ]
        for row in rows:
            sheet.append(row)
        self._style_table(sheet, 1, 1, len(rows), 3)
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 90
        for row in sheet.iter_rows(min_row=2, max_row=len(rows), min_col=3, max_col=3):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    def _build_formulas(self, sheet) -> None:
        rows = [
            (
                "Цена до скидки",
                "Зачеркнутая/стартовая цена Ozon: realFBS / (1 - видимая скидка).",
            ),
            (
                "Цена realFBS",
                "Фактическая цена реализации: выручка, комиссия, DRR и эквайринг.",
            ),
            (
                "Цена с картой",
                "Цена покупателя при банковской скидке; не расход без подтверждения.",
            ),
            (
                "Оплата покупателя",
                (
                    "Сколько платит покупатель после применения баллов. "
                    "В отчете Ozon это строка «Выручка»."
                ),
            ),
            ("Баллы начислено", "Строка «Баллы за скидки» из отчета начислений Ozon."),
            ("Начисления всего", "Оплата покупателя + баллы + программы партнеров + доставка."),
            (
                "Чистая прибыль Ozon",
                (
                    "Начисления всего без НДС, если НДС применим, - закупка - "
                    "комиссия - логистика - "
                    "эквайринг - упаковка - fulfillment - реклама - прочие - налог."
                ),
            ),
            (
                "Дополнительные расходы бизнеса",
                (
                    "Считаются отдельно от текущей прибыли: быстрый вывод 2,45%, "
                    "дизайнер/контент 4%. Процентный резерв фулфилмента отключен."
                ),
            ),
            (
                "Чистая прибыль бизнеса",
                (
                    "Чистая прибыль Ozon - быстрый вывод - дизайнер/контент - "
                    "фактические расходы подготовки и рейса FBS."
                ),
            ),
            ("УСН 6%", "Для текущего ИП считается от начислений Ozon, не от прибыли."),
            (
                "1% свыше 300к",
                "В юнитке как переменная оценка: начисления Ozon * 1%.",
            ),
            (
                "Себес без комиссии",
                "Закупка без НДС + логистика + эквайринг + упаковка + fulfillment + прочие.",
            ),
            ("Себес + комиссия", "Себес без комиссии + комиссия Ozon."),
            ("DRR", "Сценарии 10%, 12%, 15%. Базовый расчет сейчас использует 12%."),
            (
                "Конкурент оценка",
                (
                    "Если сохранен ручной конкурент, используется его цена. "
                    "Если нет, поле остается оценочным и должно быть заменено данными Ozon/отчета."
                ),
            ),
            (
                "Тип конкурента",
                "exact — тот же товар, analog — аналог, reference — ценовой ориентир.",
            ),
        ]
        sheet.append(["Показатель", "Пояснение"])
        for row in rows:
            sheet.append(row)
        self._style_table(sheet, 1, 1, len(rows) + 1, 2)
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 110
        for row in sheet.iter_rows(min_row=2, max_row=len(rows) + 1):
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

    def _style_table(self, sheet, start_row: int, start_col: int, rows: int, cols: int) -> None:
        header_fill = PatternFill("solid", fgColor="005BFF")
        header_font = Font(color="FFFFFF", bold=True)
        for column in range(start_col, start_col + cols):
            cell = sheet.cell(row=start_row, column=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(
            min_row=start_row + 1,
            max_row=start_row + rows - 1,
            min_col=start_col,
            max_col=start_col + cols - 1,
        ):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.auto_filter.ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(start_col + cols - 1)}{start_row + rows - 1}"
        )

    def _configure_formula_recalculation(self, workbook: Workbook) -> None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
